"""
general_input_tab.py
====================================================================
"General Input Information" tab for the Quantitative Risk Assessment
System.

This is a faithful PyQt5 re-implementation of the standalone
"도로터널 FDS 계산기 / Road Tunnel FDS Calculator"
(Road_tunnel_fds_calculator.html).

It produces an FDS input deck from a tunnel cross-section, wall-cell
geometry, fire scenario and inlet boundary condition, and shows a live
cross-section preview + the generated .fds text.

The widget is completely self-contained: drop it into the existing
QTabWidget as the FIRST tab.  See INTEGRATION notes at the bottom of
this file.

Author: ported to PyQt5 for the QRA System.
"""

from __future__ import annotations

import math
import os
import re
import zipfile

from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QGridLayout, QLabel, QLineEdit,
    QComboBox, QPushButton, QGroupBox, QTableWidget, QTableWidgetItem,
    QPlainTextEdit, QScrollArea, QFileDialog, QFrame, QSizePolicy,
    QApplication, QMessageBox, QHeaderView, QListWidget, QListWidgetItem,
    QToolButton, QTabWidget,
)
from PyQt5.QtCore import Qt, QEvent
from PyQt5.QtGui import QFont, QDoubleValidator, QIntValidator, QCursor, QColor

# --- optional matplotlib preview (degrades gracefully if unavailable) ---
try:
    from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
    from matplotlib.figure import Figure
    from matplotlib.patches import Rectangle
    _HAS_MPL = True
except Exception:  # pragma: no cover
    _HAS_MPL = False

# --- optional openpyxl (only needed for the Statistics Excel upload) -----
try:
    import openpyxl as _openpyxl
    _HAS_OPENPYXL = True
except Exception:  # pragma: no cover
    _openpyxl = None
    _HAS_OPENPYXL = False

# --- optional Scenario panel (radios + node-connection diagram) ----------
try:
    from scenario_panel import ScenarioPanel
    _HAS_SCENARIO = True
except Exception:  # pragma: no cover
    ScenarioPanel = None
    _HAS_SCENARIO = False


# ====================================================================
#  Reference data (identical to the HTML calculator)
# ====================================================================
FIRE = [
    [10, 0.0005], [20, 0.002], [30, 0.0045], [40, 0.008], [50, 0.0125],
    [60, 0.0179], [70, 0.0244], [80, 0.0319], [90, 0.0404], [100, 0.0498],
    [110, 0.0603], [120, 0.0717], [130, 0.0842], [140, 0.0977], [150, 0.1121],
    [160, 0.1276], [170, 0.144], [180, 0.1614], [190, 0.1799], [200, 0.1993],
    [210, 0.2197], [220, 0.2412], [230, 0.2636], [240, 0.287], [250, 0.3114],
    [260, 0.3368], [270, 0.3632], [280, 0.3906], [290, 0.419], [300, 0.4484],
    [310, 0.4788], [320, 0.5102], [330, 0.5426], [340, 0.576], [350, 0.6104],
    [360, 0.6457], [370, 0.6821], [380, 0.7195], [390, 0.7578], [400, 0.7972],
    [410, 0.8376], [420, 0.8789], [430, 0.9213], [440, 0.9646], [448, 1],
]

# Default velocity-ramp profile (used when no .VRR profile is supplied)
VR = [
    [0, 0.16], [20, 0.15], [40, 0.13], [60, 0.11], [80, 0.09], [100, 0.07],
    [120, 0.04], [140, 0.01], [160, -0.02], [180, -0.05], [200, -0.08],
    [220, -0.11], [240, -0.14], [260, -0.17], [280, -0.2], [300, -0.24],
    [320, -0.28], [340, -0.33], [360, -0.4], [380, -0.48], [400, -0.56],
    [420, -0.66], [440, -0.78], [460, -0.91], [480, -1.05], [500, -1.19],
    [520, -1.34], [540, -1.47], [560, -1.53], [580, -1.53], [600, -1.5],
    [620, -1.46], [640, -1.42], [660, -1.39], [680, -1.37], [700, -1.37],
    [720, -1.37], [740, -1.37], [760, -1.38], [780, -1.4], [800, -1.43],
    [820, -1.45], [840, -1.46], [860, -1.44], [880, -1.43], [900, -1.42],
]

# preset -> [HRRPUA, fire-car width, fire-car height Z2, width offset -, width offset +]
PRESETS = {
    "20":  [533.333, 3.75, 2.5, 1, 1],
    "30":  [923.077, 3.75, 2.5, 1, 1],
    "100": [900.901, 9.25, 3,   1, 1.5],
}

# Default LEFT wall-cell coordinates [X1, X2, Y1, Y2] for the first rows.
DEFAULT_CELLS = [
    [0, 2, 0, 1], [0, 1, 0, 1], [0, 1, 7, 9], [0, 2, 9, 11],
    [0, 3, 11, 12], [0, 4, 12, 13], [0, 5, 13, 14], [0, 8, 14, 15],
]

N_ROWS = 30  # number of wall-cell rows


# ====================================================================
#  Small numeric helpers (match the JS behaviour)
# ====================================================================
def _jsround(x: float) -> int:
    """Round half up, like JavaScript Math.round."""
    try:
        return int(math.floor(float(x) + 0.5))
    except Exception:
        return 0


def fmt(x) -> str:
    """Format a number the way the calculator does:
    integers print without a decimal point, others to <= 4 dp (trimmed)."""
    try:
        x = float(x)
    except Exception:
        x = 0.0
    if not math.isfinite(x):
        x = 0.0
    if abs(x) < 1e-9:
        x = 0.0
    if round(x * 1e6) / 1e6 == round(x):
        return str(int(round(x)))
    v = round(x * 1e4) / 1e4
    s = ("%.4f" % v).rstrip("0").rstrip(".")
    return s if s not in ("", "-0") else "0"


def safe_name(s: str) -> str:
    return "".join(c if (c.isalnum() or c in "_-") else "_" for c in str(s))


def decide_case(data) -> str:
    """First non-zero velocity -> positive => CASE 3, negative => CASE 4."""
    for _, v in data:
        if abs(v) > 1e-9:
            return "3" if v > 0 else "4"
    return "3"


# ====================================================================
#  The tab widget
# ====================================================================
class GeneralInputTab(QWidget):
    """Self-contained FDS-deck calculator tab."""

    MONO = "Consolas, 'DejaVu Sans Mono', 'Courier New', monospace"

    def __init__(self, parent=None):
        super().__init__(parent)
        self.sym = True            # left -> right auto symmetry
        self._building = False     # re-entrancy guard
        self._last_fds = ""
        self._last_fname = "TN.fds"
        # --- .VRR batch-import state ---
        self.files = []            # list of dicts: name/title/mw/profile/single
        self.active = -1           # index of active file (-1 = manual)
        self.sign = 1              # velocity-profile sign multiplier (+1/-1)
        # --- DXF cross-section state ---
        self.dxf_poly = None       # list of {'x','y'} in metres, origin-shifted
        self.dxf_w = 0.0
        self.dxf_h = 0.0
        self._last_geom_sig = None # (W,H,nW,nH) used to re-apply DXF bands
        self._build_ui()
        self._init_defaults()
        self.update_all()

    # ----------------------------------------------------------------
    #  UI construction
    # ----------------------------------------------------------------
    def _num_field(self, default="", decimals=True):
        e = QLineEdit(str(default))
        if decimals:
            v = QDoubleValidator()
            v.setNotation(QDoubleValidator.StandardNotation)
            e.setValidator(v)
        else:
            e.setValidator(QIntValidator())
        e.setStyleSheet(
            "QLineEdit{background:#0b1119;color:#e8eef6;border:1px solid #243246;"
            "border-radius:6px;padding:6px;font-family:%s;}" % self.MONO)
        e.textChanged.connect(self.update_all)
        return e

    def _label(self, text, unit=""):
        lab = QLabel(text + (f"  ({unit})" if unit else ""))
        lab.setStyleSheet("color:#5a6b7d;font-size:11px;")
        return lab

    def _build_ui(self):
        outer = QVBoxLayout(self)
        outer.setContentsMargins(6, 6, 6, 6)

        header = QLabel("General Input Information")
        header.setStyleSheet(
            "font-size:16px;font-weight:bold;color:#2c3e50;padding:8px;")
        outer.addWidget(header)

        # body inside a scroll area: a full-width Basic Tunnel Information
        # panel on top, then two side-by-side rows beneath it.
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        body = QWidget()
        scroll.setWidget(body)
        body_v = QVBoxLayout(body)
        body_v.setSpacing(12)
        outer.addWidget(scroll)

        # The body is split into two stacked sections:
        #   * info_box : Basic Tunnel Information, Traffic Information, Statistics
        #   * fds_box  : the FDS deck calculator (cross-section grid, preview,
        #                wall cells, fire scenario, inlet BC, VRR import, output)
        # When integrated, the launcher calls take_fds_section() to move fds_box
        # into its own "FDS" tab; standalone they remain stacked here.
        self.body_v = body_v
        self._info_box = QWidget()
        info_v = QVBoxLayout(self._info_box)
        info_v.setContentsMargins(0, 0, 0, 0)
        info_v.setSpacing(12)
        self._fds_box = QWidget()
        fds_v = QVBoxLayout(self._fds_box)
        fds_v.setContentsMargins(0, 0, 0, 0)
        fds_v.setSpacing(12)
        body_v.addWidget(self._info_box)
        body_v.addWidget(self._fds_box)

        # ---- 00 Basic Tunnel Information  (full width) ----------------
        # When this tab runs inside the full QRA System the launcher moves
        # the EVC/FED "Tunnel Basic Specifications", "Fire Location,
        # Evacuation Zone", the Fire Placement Mode controls and "Fire
        # Point Mapping" groups into this panel.  Standalone, a placeholder
        # explains where they come from.
        self.basic_info_group = QGroupBox("Basic Tunnel Information")
        self.basic_info_group.setStyleSheet("QGroupBox{font-weight:bold;}")
        self.basic_info_row = QHBoxLayout(self.basic_info_group)
        self.basic_info_row.setContentsMargins(8, 6, 8, 8)
        self.basic_info_row.setSpacing(10)
        self._basic_info_placeholder = QLabel(
            "Tunnel basic specification, fire-location and fire-placement "
            "controls appear here when this tab is loaded inside the QRA "
            "System (moved from the EVC/FED Analysis tab).")
        self._basic_info_placeholder.setWordWrap(True)
        self._basic_info_placeholder.setStyleSheet(
            "color:#5a6b7d;font-size:11px;padding:6px;")
        self.basic_info_row.addWidget(self._basic_info_placeholder)
        # (added to the info sub-tabs below)

        # ---- 00b Traffic Information  (full width, under Basic Tunnel) -
        # The launcher moves the EVC/FED "Traffic Volume & Vehicle Specs"
        # table (renamed "Traffic Volume") and the pre-fire driving-state
        # control frame into this panel, side by side (20% / 80%).
        self.traffic_info_group = QGroupBox("Traffic Information")
        self.traffic_info_group.setStyleSheet("QGroupBox{font-weight:bold;}")
        self.traffic_info_row = QHBoxLayout(self.traffic_info_group)
        self.traffic_info_row.setContentsMargins(8, 6, 8, 8)
        self.traffic_info_row.setSpacing(10)
        self._traffic_info_placeholder = QLabel(
            "Traffic volume table and pre-fire driving-state controls appear "
            "here when this tab is loaded inside the QRA System (moved from "
            "the EVC/FED Analysis tab).")
        self._traffic_info_placeholder.setWordWrap(True)
        self._traffic_info_placeholder.setStyleSheet(
            "color:#5a6b7d;font-size:11px;padding:6px;")
        self.traffic_info_row.addWidget(self._traffic_info_placeholder)
        # (added to the info sub-tabs below)

        # ---- 00c Statistics  (full width, under Traffic Information) --
        self.stats_group = QGroupBox("Statistics")
        self.stats_group.setStyleSheet("QGroupBox{font-weight:bold;}")
        self._build_stats_panel(self.stats_group)
        # (added to the info sub-tabs below)

        # ---- 00d Scenario  (full width, under Statistics) ------------
        self.scenario_group = QGroupBox("Scenario")
        self.scenario_group.setStyleSheet("QGroupBox{font-weight:bold;}")
        _sv = QVBoxLayout(self.scenario_group)
        _sv.setContentsMargins(6, 6, 6, 6)
        if _HAS_SCENARIO:
            try:
                self.scenario_panel = ScenarioPanel()
                _sv.addWidget(self.scenario_panel)
            except Exception as _exc:
                self.scenario_panel = None
                _lbl = QLabel("Scenario panel unavailable: %s" % _exc)
                _lbl.setStyleSheet("color:#c0392b;font-size:11px;")
                _sv.addWidget(_lbl)
        else:
            self.scenario_panel = None
            _lbl = QLabel("Scenario panel needs scenario_panel.py beside this module.")
            _lbl.setStyleSheet("color:#5a6b7d;font-size:11px;")
            _sv.addWidget(_lbl)
        # (added to the info sub-tabs below, together with Project Directory)

        # ---- 00e Project Directory  (full width, under Scenario) ------
        # The launcher moves the Directory-Setup tab's "Project
        # Configuration" and "Directory Structure Preview" groups in here,
        # side by side (each ~half the width).
        self.project_dir_group = QGroupBox("Project Directory")
        self.project_dir_group.setStyleSheet("QGroupBox{font-weight:bold;}")
        self.project_dir_row = QHBoxLayout(self.project_dir_group)
        self.project_dir_row.setContentsMargins(8, 6, 8, 8)
        self.project_dir_row.setSpacing(10)
        self._project_dir_placeholder = QLabel(
            "Project Configuration and Directory Structure Preview appear "
            "here when this tab is loaded inside the QRA System (moved from "
            "the Directory Setup tab).")
        self._project_dir_placeholder.setWordWrap(True)
        self._project_dir_placeholder.setStyleSheet(
            "color:#5a6b7d;font-size:11px;padding:6px;")
        self.project_dir_row.addWidget(self._project_dir_placeholder)

        # ---- Assemble the four panels into sub-tabs -------------------
        # Basic Tunnel Information | Traffic Information | Statistics |
        # Scenario (which also carries the Project Directory panel).
        def _tab_page(widget):
            sa = QScrollArea()
            sa.setWidgetResizable(True)
            sa.setFrameShape(QFrame.NoFrame)
            sa.setWidget(widget)
            return sa

        # the tab labels already name each section, so drop the redundant
        # group-box titles (keep the frame)
        for _g in (self.basic_info_group, self.traffic_info_group,
                   self.stats_group, self.scenario_group):
            _g.setTitle("")

        # Scenario tab = Scenario panel + Project Directory stacked
        scen_page = QWidget()
        scen_v = QVBoxLayout(scen_page)
        scen_v.setContentsMargins(0, 0, 0, 0)
        scen_v.setSpacing(12)
        scen_v.addWidget(self.scenario_group)
        scen_v.addWidget(self.project_dir_group)
        scen_v.addStretch(1)

        self.info_tabs = QTabWidget()
        self.info_tabs.setDocumentMode(True)
        self.info_tabs.addTab(_tab_page(self.basic_info_group),
                              "Basic Tunnel Information")
        self.info_tabs.addTab(_tab_page(self.traffic_info_group),
                              "Traffic Information")
        self.info_tabs.addTab(_tab_page(self.stats_group), "Statistics")
        self.info_tabs.addTab(_tab_page(scen_page), "Scenario")
        self.info_tabs.setMinimumHeight(560)
        info_v.addWidget(self.info_tabs)

        # ---- Row 1: Cross-section / grid  +  preview  (side by side) --
        row1 = QHBoxLayout()
        row1.setSpacing(12)
        grid_col = QVBoxLayout()
        preview_col = QVBoxLayout()
        row1.addLayout(grid_col, 5)
        row1.addLayout(preview_col, 6)
        fds_v.addLayout(row1)

        # ---- Row 2: remaining inputs  +  FDS output ------------------
        row2 = QHBoxLayout()
        row2.setSpacing(12)
        left_col = QVBoxLayout()
        right_col = QVBoxLayout()
        row2.addLayout(left_col, 5)
        row2.addLayout(right_col, 6)
        fds_v.addLayout(row2)

        # ---- 01 Cross-section / grid ----------------------------------
        g1 = QGroupBox("01  Tunnel Cross-section / Grid  (shared by all files)")
        g1.setStyleSheet("QGroupBox{font-weight:bold;}")
        grid1 = QGridLayout(g1)
        self.in_W = self._num_field(10.86)
        self.in_nW = self._num_field(20, decimals=False)
        self.in_H = self._num_field(6.73)
        self.in_nH = self._num_field(15, decimals=False)
        self.in_Lh = self._num_field(720)
        self.in_dx = self._num_field(1)
        self.in_breath = self._num_field(1.8)
        self.lbl_breath_snap = QLineEdit("—")
        self.lbl_breath_snap.setReadOnly(True)
        self.lbl_breath_snap.setStyleSheet(
            "QLineEdit{background:#0e1620;color:#4cb3d6;border:1px dashed #243246;"
            "border-radius:6px;padding:6px;font-family:%s;}" % self.MONO)

        fields1 = [
            ("Width (W)", "m", self.in_W),
            ("Width cell count (J)", "", self.in_nW),
            ("Height (H)", "m", self.in_H),
            ("Height cell count (K)", "", self.in_nH),
            ("Half tunnel length (½)", "m", self.in_Lh),
            ("Length cell size Δx", "m", self.in_dx),
            ("Breath line (SLCF)", "m", self.in_breath),
            ("Breath cell snap", "m", self.lbl_breath_snap),
        ]
        for i, (name, unit, field) in enumerate(fields1):
            r, c = divmod(i, 2)
            cell = QVBoxLayout()
            cell.addWidget(self._label(name, unit))
            cell.addWidget(field)
            grid1.addLayout(cell, r, c)
        grid_col.addWidget(g1)

        # derived chips
        self.derived_lbl = QLabel("")
        self.derived_lbl.setWordWrap(True)
        self.derived_lbl.setStyleSheet(
            "background:#0e1620;color:#cdd9e6;border:1px solid #243246;"
            "border-radius:8px;padding:8px;font-family:%s;font-size:11px;" % self.MONO)
        grid_col.addWidget(self.derived_lbl)

        # ---- 02 Wall-cell coordinates --------------------------------
        g2 = QGroupBox("02  Wall-cell Coordinates  (all-zero rows are skipped)")
        g2.setStyleSheet("QGroupBox{font-weight:bold;}")
        v2 = QVBoxLayout(g2)

        symbar = QHBoxLayout()
        self.btn_sym_auto = QPushButton("Left → Right auto-symmetry")
        self.btn_sym_free = QPushButton("Edit left / right independently")
        for b in (self.btn_sym_auto, self.btn_sym_free):
            b.setCheckable(True)
        self.btn_sym_auto.setChecked(True)
        self.btn_sym_auto.clicked.connect(lambda: self.set_sym(True))
        self.btn_sym_free.clicked.connect(lambda: self.set_sym(False))
        symbar.addWidget(self.btn_sym_auto)
        symbar.addWidget(self.btn_sym_free)
        symbar.addStretch(1)
        v2.addLayout(symbar)

        self.table = QTableWidget(N_ROWS, 8)
        self.table.setHorizontalHeaderLabels(
            ["L·X1", "L·X2", "L·Y1", "L·Y2", "R·X1", "R·X2", "R·Y1", "R·Y2"])
        self.table.verticalHeader().setDefaultSectionSize(22)
        self.table.setVerticalHeaderLabels([f"{i+1}" for i in range(N_ROWS)])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table.setStyleSheet(
            "QTableWidget{font-family:%s;font-size:11px;}" % self.MONO)
        self.table.setMaximumHeight(360)
        self.table.cellChanged.connect(self._on_cell_changed)
        self._setup_cell_spinner()
        v2.addWidget(self.table)

        note = QLabel("Auto-symmetry:  X1ʀ = J − X2,   X2ʀ = J − X1   (right cells follow left).")
        note.setStyleSheet("color:#5a6b7d;font-size:11px;")
        v2.addWidget(note)

        # ---- DXF cross-section import ----
        dxfbar = QHBoxLayout()
        self.btn_dxf = QPushButton("📐 Load DXF section…")
        self.btn_dxf.clicked.connect(self._on_dxf_load)
        self.btn_dxf.setStyleSheet(
            "QPushButton{background:#6c5ce7;color:white;font-weight:bold;"
            "border-radius:6px;padding:6px 12px;}")
        self.btn_dxf_clear = QPushButton("✕ clear DXF")
        self.btn_dxf_clear.clicked.connect(self._on_dxf_clear)
        self.btn_dxf_clear.setStyleSheet(
            "QPushButton{background:#3a4250;color:#cdd9e6;border-radius:6px;padding:6px 12px;}")
        self.btn_dxf_clear.hide()
        dxfbar.addWidget(self.btn_dxf)
        dxfbar.addWidget(self.btn_dxf_clear)
        dxfbar.addStretch(1)
        v2.addLayout(dxfbar)
        self.dxf_info = QLabel("Import a closed cross-section polyline (LWPOLYLINE / POLYLINE). "
                               "Sets W·H and fills wall cells by point-in-polygon.")
        self.dxf_info.setWordWrap(True)
        self.dxf_info.setStyleSheet("color:#8a7ddb;font-size:11px;")
        v2.addWidget(self.dxf_info)
        left_col.addWidget(g2)

        # ---- 04 Fire scenario ----------------------------------------
        g4 = QGroupBox("04  Fire Scenario / Fire Vehicle")
        g4.setStyleSheet("QGroupBox{font-weight:bold;}")
        grid4 = QGridLayout(g4)
        self.cmb_preset = QComboBox()
        self.cmb_preset.addItems(["20 MW", "30 MW", "100 MW", "Custom"])
        self.cmb_preset.currentIndexChanged.connect(self._on_preset_changed)
        self.in_hrr = self._num_field(533.333)
        self.in_bhl = self._num_field(3.75)
        self.in_bz = self._num_field(2.5)
        self.in_bo1 = self._num_field(1)
        self.in_bo2 = self._num_field(1)
        fields4 = [
            ("Preset", "", self.cmb_preset),
            ("HRRPUA", "kW/m²", self.in_hrr),
            ("Fire-car width", "m", self.in_bhl),
            ("Fire-car height Z2", "m", self.in_bz),
            ("Width offset −", "m", self.in_bo1),
            ("Width offset +", "m", self.in_bo2),
        ]
        for i, (name, unit, field) in enumerate(fields4):
            r, c = divmod(i, 3)
            cell = QVBoxLayout()
            cell.addWidget(self._label(name, unit))
            cell.addWidget(field)
            grid4.addLayout(cell, r, c)
        # mark the manual fire fields so we can flip preset->Custom on edit
        for f in (self.in_hrr, self.in_bhl, self.in_bz, self.in_bo1, self.in_bo2):
            f.textEdited.connect(self._fire_field_edited)
        left_col.addWidget(g4)

        # ---- 05 Inlet boundary condition -----------------------------
        g5 = QGroupBox("05  Inlet Boundary Condition")
        g5.setStyleSheet("QGroupBox{font-weight:bold;}")
        grid5 = QGridLayout(g5)
        self.cmb_bc = QComboBox()
        self.cmb_bc.addItems([
            "CASE 1 · Initial velocity + OPEN BC",
            "CASE 2 · Pressure Dynamic BC",
            "CASE 3 · Velocity Profile (Vr_Ramp)",
            "CASE 4 · Velocity Profile (−)",
        ])
        self.cmb_bc.setCurrentIndex(2)  # CASE 3 default
        self.cmb_bc.currentIndexChanged.connect(self.update_all)
        self.in_twfin = self._num_field(1200)
        self.in_v0 = self._num_field(6.02)
        cell = QVBoxLayout()
        cell.addWidget(self._label("Active CASE"))
        cell.addWidget(self.cmb_bc)
        grid5.addLayout(cell, 0, 0, 1, 2)
        cell = QVBoxLayout()
        cell.addWidget(self._label("TWFIN", "s"))
        cell.addWidget(self.in_twfin)
        grid5.addLayout(cell, 1, 0)
        cell = QVBoxLayout()
        cell.addWidget(self._label("V0 (CASE 1)", "m/s"))
        cell.addWidget(self.in_v0)
        grid5.addLayout(cell, 1, 1)
        left_col.addWidget(g5)

        # ---- 03 .VRR batch import -----------------------------------
        g3 = QGroupBox("03  Wind-velocity Batch Import  (.VRR / .txt profiles)")
        g3.setStyleSheet("QGroupBox{font-weight:bold;}")
        v3 = QVBoxLayout(g3)

        topbar = QHBoxLayout()
        self.btn_add_vrr = QPushButton("➕ Add .VRR files…")
        self.btn_add_vrr.clicked.connect(self._on_add_vrr)
        self.btn_add_vrr.setStyleSheet(
            "QPushButton{background:#16a085;color:white;font-weight:bold;"
            "border-radius:6px;padding:6px 12px;}")
        self.btn_clear_vrr = QPushButton("🗑 Clear all")
        self.btn_clear_vrr.clicked.connect(self._clear_files)
        self.btn_clear_vrr.setStyleSheet(
            "QPushButton{background:#3a4250;color:#cdd9e6;border-radius:6px;padding:6px 12px;}")
        topbar.addWidget(self.btn_add_vrr)
        topbar.addWidget(self.btn_clear_vrr)
        topbar.addStretch(1)
        topbar.addWidget(self._label("Single-col Δt", "s"))
        self.in_dt = self._num_field(20)
        self.in_dt.setMaximumWidth(70)
        topbar.addWidget(self.in_dt)
        topbar.addWidget(self._label("Sign"))
        self.cmb_sign = QComboBox()
        self.cmb_sign.addItems(["+1", "−1"])
        self.cmb_sign.currentIndexChanged.connect(self._on_sign_changed)
        self.cmb_sign.setMaximumWidth(64)
        topbar.addWidget(self.cmb_sign)
        v3.addLayout(topbar)

        self.queue = QListWidget()
        self.queue.setMaximumHeight(150)
        self.queue.setStyleSheet(
            "QListWidget{background:#0b1119;color:#e8eef6;border:1px solid #243246;"
            "border-radius:6px;font-family:%s;font-size:11px;}"
            "QListWidget::item:selected{background:#16a085;color:white;}" % self.MONO)
        self.queue.itemClicked.connect(self._on_queue_clicked)
        v3.addWidget(self.queue)

        qbtns = QHBoxLayout()
        self.btn_remove_vrr = QPushButton("− Remove selected")
        self.btn_remove_vrr.clicked.connect(self._remove_file)
        self.btn_remove_vrr.setStyleSheet(
            "QPushButton{background:#3a4250;color:#cdd9e6;border-radius:6px;padding:5px 10px;}")
        qbtns.addWidget(self.btn_remove_vrr)
        qbtns.addStretch(1)
        v3.addLayout(qbtns)

        self.lock_note = QLabel("")
        self.lock_note.setWordWrap(True)
        self.lock_note.setStyleSheet(
            "background:#1a2233;color:#f0c674;border:1px solid #4a3f1e;"
            "border-radius:6px;padding:6px;font-size:11px;")
        self.lock_note.hide()
        v3.addWidget(self.lock_note)

        # velocity-profile chart (active file)
        if _HAS_MPL:
            self.vfig = Figure(figsize=(5, 1.9), facecolor="#0b1017")
            self.vcanvas = FigureCanvas(self.vfig)
            self.vcanvas.setMinimumHeight(150)
            self.vax = self.vfig.add_subplot(111)
            self.vcanvas.hide()
            v3.addWidget(self.vcanvas)
        else:
            self.vcanvas = None
        left_col.addWidget(g3)

        left_col.addStretch(1)

        # ---- RIGHT: preview + FDS output -----------------------------
        gp = QGroupBox("▣  Cross-section Preview")
        gp.setStyleSheet("QGroupBox{font-weight:bold;}")
        vp = QVBoxLayout(gp)
        if _HAS_MPL:
            self.fig = Figure(figsize=(5, 3.4), facecolor="#0b1017")
            self.canvas = FigureCanvas(self.fig)
            self.canvas.setMinimumHeight(300)
            self.ax = self.fig.add_subplot(111)
            vp.addWidget(self.canvas)
        else:
            self.canvas = None
            warn = QLabel("(matplotlib not available — preview disabled)")
            warn.setStyleSheet("color:#8aa0b8;font-size:11px;padding:20px;")
            vp.addWidget(warn)
        preview_col.addWidget(gp)

        g6 = QGroupBox("06  FDS Output")
        g6.setStyleSheet("QGroupBox{font-weight:bold;}")
        v6 = QVBoxLayout(g6)
        btnbar = QHBoxLayout()
        self.btn_copy = QPushButton("📋 Copy")
        self.btn_save = QPushButton("💾 Save .fds")
        self.btn_save_all = QPushButton("🗜 Save All (ZIP)")
        self.btn_reset = QPushButton("↺ Reset")
        self.btn_copy.clicked.connect(self._copy)
        self.btn_save.clicked.connect(self._save)
        self.btn_save_all.clicked.connect(self._save_all_zip)
        self.btn_reset.clicked.connect(self._reset)
        for b, color in ((self.btn_copy, "#e67e22"),
                         (self.btn_save, "#3a9ec2"),
                         (self.btn_save_all, "#16a085"),
                         (self.btn_reset, "#7f8c8d")):
            b.setStyleSheet(
                "QPushButton{background:%s;color:white;font-weight:bold;"
                "border-radius:6px;padding:7px 14px;}" % color)
            btnbar.addWidget(b)
        self.btn_save_all.setEnabled(False)
        self.stat_lbl = QLabel("")
        self.stat_lbl.setStyleSheet("color:#3fb98a;font-family:%s;" % self.MONO)
        btnbar.addWidget(self.stat_lbl)
        btnbar.addStretch(1)
        v6.addLayout(btnbar)

        self.code = QPlainTextEdit()
        self.code.setReadOnly(True)
        self.code.setStyleSheet(
            "QPlainTextEdit{background:#080c12;color:#cdd9e6;border:1px solid #243246;"
            "border-radius:8px;padding:10px;font-family:%s;font-size:11px;}" % self.MONO)
        self.code.setMinimumHeight(360)
        v6.addWidget(self.code)
        right_col.addWidget(g6)

    # ----------------------------------------------------------------
    #  Basic Tunnel Information panel (populated by the launcher)
    # ----------------------------------------------------------------
    def _clear_basic_placeholder(self):
        if self._basic_info_placeholder is not None:
            self._basic_info_placeholder.hide()
            self.basic_info_row.removeWidget(self._basic_info_placeholder)
            self._basic_info_placeholder.deleteLater()
            self._basic_info_placeholder = None

    def mount_basic_info(self, group_widget):
        """Add an external group box (e.g. the EVC/FED 'Tunnel Basic
        Specifications' panel) as the first column of the full-width Basic
        Tunnel Information panel.  Reparenting preserves every child widget,
        its value and all signal connections, so the main app keeps reading
        the same self.tbi_* / self.evc_* objects."""
        if group_widget is None:
            return False
        old = group_widget.parentWidget()
        if old is not None and old.layout() is not None:
            old.layout().removeWidget(group_widget)
        self._clear_basic_placeholder()
        self.basic_info_row.insertWidget(0, group_widget)
        group_widget.show()
        return True

    def add_basic_info_widget(self, widget):
        """Append a widget (typically a group box) as another column in the
        Basic Tunnel Information panel.  The caller is responsible for having
        detached it from any previous layout."""
        if widget is None:
            return False
        self._clear_basic_placeholder()
        self.basic_info_row.addWidget(widget)
        widget.show()
        return True

    def add_basic_info_stack(self, widgets):
        """Append a vertical stack of widgets as a single column in the
        Basic Tunnel Information panel."""
        widgets = [w for w in (widgets or []) if w is not None]
        if not widgets:
            return False
        self._clear_basic_placeholder()
        box = QWidget()
        v = QVBoxLayout(box)
        v.setContentsMargins(0, 0, 0, 0)
        v.setSpacing(6)
        for w in widgets:
            v.addWidget(w)
            w.show()
        v.addStretch(1)
        self.basic_info_row.addWidget(box)
        return True

    # ----------------------------------------------------------------
    #  Traffic Information panel (populated by the launcher)
    # ----------------------------------------------------------------
    def mount_traffic_info(self, control_widget, table_widget,
                           control_pct=20, table_pct=80):
        """Place the moved pre-fire driving-state control (left) and the
        Traffic Volume table (right) side by side in the Traffic Information
        panel at the given width percentages.  Reparenting preserves every
        child widget, value and signal, so the main app keeps reading the
        same self.evc_* / self.tbi_* objects."""
        if self._traffic_info_placeholder is not None:
            self._traffic_info_placeholder.hide()
            self.traffic_info_row.removeWidget(self._traffic_info_placeholder)
            self._traffic_info_placeholder.deleteLater()
            self._traffic_info_placeholder = None
        if control_widget is not None:
            self.traffic_info_row.addWidget(control_widget, control_pct)
            control_widget.show()
        if table_widget is not None:
            self.traffic_info_row.addWidget(table_widget, table_pct)
            table_widget.show()
        return True

    def mount_project_directory(self, config_widget, preview_widget,
                                config_pct=50, preview_pct=50):
        """Place the 'Project Configuration' (left) and 'Directory Structure
        Preview' (right) groups side by side, each covering half the width.
        Reparenting preserves every child widget, value and signal, so the
        main app keeps reading the same self.project_name_input /
        self.dir_tree objects."""
        if self._project_dir_placeholder is not None:
            self._project_dir_placeholder.hide()
            self.project_dir_row.removeWidget(self._project_dir_placeholder)
            self._project_dir_placeholder.deleteLater()
            self._project_dir_placeholder = None
        if config_widget is not None:
            self.project_dir_row.addWidget(config_widget, config_pct)
            config_widget.show()
        if preview_widget is not None:
            self.project_dir_row.addWidget(preview_widget, preview_pct)
            preview_widget.show()
        # Make the panel 50% taller than its natural height.
        try:
            self.project_dir_group.layout().activate()
            nat = self.project_dir_group.sizeHint().height()
            if nat > 0:
                self.project_dir_group.setMinimumHeight(int(nat * 1.5))
        except Exception:
            pass
        return True

    def take_fds_section(self):
        """Detach and return the FDS-calculator section (everything below the
        Statistics panel: cross-section grid, preview, wall cells, fire
        scenario, inlet BC, VRR import and FDS output) so it can be placed in
        its own tab.  The widgets stay driven by this object's methods, so the
        live preview and deck generation keep working after reparenting.
        After this call the General Input tab shows only the info panels."""
        box = getattr(self, "_fds_box", None)
        if box is None:
            return None
        try:
            self.body_v.removeWidget(box)
        except Exception:
            pass
        self._fds_box = None
        return box

    # ----------------------------------------------------------------
    #  Statistics panel  (fire-accident statistics, Excel-uploadable)
    # ----------------------------------------------------------------
    STAT_CATS = ["Passenger Car", "Bus", "Truck", "Total"]

    def _build_stats_panel(self, group):
        self._stats_building = False
        self.stats_fire = {}       # year -> [pcar, bus, truck, total]
        self.stats_mileage = {}    # year -> [...]
        self.stats_rate = {}       # year -> [...]  (optional parsed Case/Year)
        self.stats_selected = set()

        v = QVBoxLayout(group)
        v.setContentsMargins(8, 6, 8, 8)
        v.setSpacing(6)

        bar = QHBoxLayout()
        bar.addWidget(self._label("Metric"))
        self.stats_metric_combo = QComboBox()
        self.stats_metric_combo.addItems(
            ["Fire Cases", "Mileage", "Rate (Cases / Mileage \u00d7100)"])
        self.stats_metric_combo.currentIndexChanged.connect(self._on_stats_metric_changed)
        bar.addWidget(self.stats_metric_combo)
        bar.addSpacing(12)
        self.stats_upload_btn = QPushButton("\U0001F4C2 Upload Excel\u2026")
        self.stats_upload_btn.clicked.connect(self._load_stats_excel)
        self.stats_upload_btn.setStyleSheet(
            "QPushButton{background:#2980b9;color:white;font-weight:bold;"
            "border-radius:6px;padding:6px 12px;}")
        bar.addWidget(self.stats_upload_btn)
        self.stats_sample_btn = QPushButton("Load sample")
        self.stats_sample_btn.clicked.connect(
            lambda: (self._seed_default_stats(), self._refresh_stats()))
        self.stats_sample_btn.setStyleSheet(
            "QPushButton{background:#3a4250;color:#cdd9e6;border-radius:6px;padding:6px 12px;}")
        bar.addWidget(self.stats_sample_btn)
        bar.addStretch(1)
        for txt, mode in (("All", "all"), ("None", "none"),
                          ("Last 5", "last5"), ("Last 10", "last10")):
            b = QPushButton(txt)
            b.setStyleSheet(
                "QPushButton{background:#34495e;color:#e8eef6;border-radius:5px;"
                "padding:5px 10px;}")
            b.clicked.connect(lambda _=False, m=mode: self._stats_select(m))
            bar.addWidget(b)
        v.addLayout(bar)

        if not _HAS_OPENPYXL:
            note = QLabel("\u2139 Excel upload needs the 'openpyxl' package "
                          "(pip install openpyxl).  Sample data is shown below.")
            note.setStyleSheet("color:#c08a2d;font-size:11px;")
            v.addWidget(note)

        self.stats_table = QTableWidget(0, 6)
        self.stats_table.setHorizontalHeaderLabels(["Apply", "Year"] + self.STAT_CATS)
        self.stats_table.verticalHeader().setVisible(False)
        self.stats_table.setStyleSheet(
            "QTableWidget{background:#ffffff;color:#1a252f;gridline-color:#cdd9e6;"
            "font-size:12px;border:1px solid #95a5a6;}"
            "QHeaderView::section{background:#eef2f7;color:#1a252f;font-weight:bold;"
            "padding:4px;border:1px solid #bdc3c7;}")
        hdr = self.stats_table.horizontalHeader()
        hdr.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        hdr.setSectionResizeMode(1, QHeaderView.ResizeToContents)
        for c in range(2, 6):
            hdr.setSectionResizeMode(c, QHeaderView.Stretch)
        rh = self.stats_table.verticalHeader().defaultSectionSize() or 28
        # show at least 7 data rows (plus header) at a glance, scroll beyond
        self.stats_table.setMinimumHeight(rh * 8 + 34)
        self.stats_table.setMaximumHeight(rh * 12 + 34)
        self.stats_table.itemChanged.connect(self._on_stats_item_changed)
        v.addWidget(self.stats_table)

        self.stats_status = QLabel("")
        self.stats_status.setStyleSheet("color:#5a6b7d;font-size:11px;")
        v.addWidget(self.stats_status)

        self._seed_default_stats()
        self._refresh_stats()

    def _seed_default_stats(self):
        fire = {
            2012: [1739, 351, 1392, 3482], 2013: [1622, 320, 1374, 3316],
            2014: [1477, 300, 1401, 3178], 2015: [1539, 268, 1485, 3292],
            2016: [1529, 281, 1415, 3225], 2017: [1524, 290, 1354, 3168],
            2018: [1608, 283, 1397, 3288], 2019: [1401, 238, 1334, 2973],
            2020: [1338, 210, 1290, 2838], 2021: [1334, 184, 1298, 2816],
            2022: [1390, 196, 1291, 2877], 2023: [1363, 251, 1353, 2967],
            2024: [1425, 226, 1397, 3048],
        }
        mile = {
            2012: [194705, 18335, 59682, 272722], 2013: [194281, 21564, 61570, 277415],
            2014: [205518, 20737, 60472, 286727], 2015: [212721, 19818, 62372, 294911],
            2016: [224024, 19857, 63688, 307569], 2017: [231274, 19318, 65374, 315966],
            2018: [239854, 18646, 68574, 327074], 2019: [242696, 17904, 68954, 329554],
            2020: [246981, 16065, 68978, 332024], 2021: [273294, 13797, 69280, 356371],
            2022: [249049, 14132, 72449, 335630], 2023: [252925, 14029, 64181, 331135],
            2024: [260456, 13742, 64534, 338732],
        }
        self.stats_fire = {y: [float(x) for x in v] for y, v in fire.items()}
        self.stats_mileage = {y: [float(x) for x in v] for y, v in mile.items()}
        self.stats_rate = {}
        self.stats_selected = set(self.stats_fire.keys())

    def _stats_all_years(self):
        ys = set(self.stats_fire) | set(self.stats_mileage) | set(self.stats_rate)
        return sorted(ys)

    def _metric_key(self):
        return ("Fire Cases", "Mileage", "Rate")[self.stats_metric_combo.currentIndex()]

    def _metric_series(self, year):
        key = self._metric_key()
        if key == "Fire Cases":
            return self.stats_fire.get(year)
        if key == "Mileage":
            return self.stats_mileage.get(year)
        f = self.stats_fire.get(year)
        m = self.stats_mileage.get(year)
        if f and m:
            return [(f[i] / m[i] * 100.0 if m[i] else 0.0) for i in range(4)]
        return self.stats_rate.get(year)

    def _avg_series(self):
        rows = [self._metric_series(y) for y in self._stats_all_years()
                if y in self.stats_selected]
        rows = [r for r in rows if r]
        if not rows:
            return None
        return [sum(r[i] for r in rows) / len(rows) for i in range(4)]

    def _refresh_stats(self):
        years = self._stats_all_years()
        self.stats_selected = {y for y in self.stats_selected if y in years} or set(years)
        self._stats_building = True
        try:
            self.stats_table.setRowCount(len(years) + (1 if years else 0))
            for r, y in enumerate(years):
                chk = QTableWidgetItem()
                chk.setFlags(Qt.ItemIsUserCheckable | Qt.ItemIsEnabled | Qt.ItemIsSelectable)
                chk.setCheckState(Qt.Checked if y in self.stats_selected else Qt.Unchecked)
                chk.setTextAlignment(Qt.AlignCenter)
                self.stats_table.setItem(r, 0, chk)
                yi = QTableWidgetItem(str(y))
                yi.setTextAlignment(Qt.AlignCenter)
                yi.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
                self.stats_table.setItem(r, 1, yi)
                series = self._metric_series(y)
                for c in range(4):
                    it = QTableWidgetItem(fmt(series[c]) if series else "")
                    it.setTextAlignment(Qt.AlignCenter)
                    it.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
                    self.stats_table.setItem(r, 2 + c, it)
            if years:
                self._fill_avg_row(len(years))
        finally:
            self._stats_building = False
        self._update_stats_status()

    def _fill_avg_row(self, r):
        avg = self._avg_series()
        bold = QFont()
        bold.setBold(True)
        blank = QTableWidgetItem("")
        blank.setFlags(Qt.ItemIsEnabled)
        self.stats_table.setItem(r, 0, blank)
        yi = QTableWidgetItem("Average (selected)")
        yi.setFlags(Qt.ItemIsEnabled)
        yi.setFont(bold)
        self.stats_table.setItem(r, 1, yi)
        for c in range(4):
            it = QTableWidgetItem(fmt(avg[c]) if avg else "")
            it.setTextAlignment(Qt.AlignCenter)
            it.setFlags(Qt.ItemIsEnabled)
            it.setFont(bold)
            it.setBackground(QColor(234, 244, 251))
            self.stats_table.setItem(r, 2 + c, it)

    def _recompute_stats_summary(self):
        years = self._stats_all_years()
        sel = set()
        for r, y in enumerate(years):
            it = self.stats_table.item(r, 0)
            if it is not None and it.checkState() == Qt.Checked:
                sel.add(y)
        self.stats_selected = sel
        self._stats_building = True
        try:
            if years:
                self._fill_avg_row(len(years))
        finally:
            self._stats_building = False
        self._update_stats_status()

    def _on_stats_item_changed(self, item):
        if self._stats_building:
            return
        if item.column() == 0:
            self._recompute_stats_summary()

    def _on_stats_metric_changed(self, _i):
        self._refresh_stats()

    def _stats_select(self, mode):
        years = self._stats_all_years()
        if mode == "all":
            sel = set(years)
        elif mode == "none":
            sel = set()
        elif mode == "last5":
            sel = set(years[-5:])
        elif mode == "last10":
            sel = set(years[-10:])
        else:
            sel = set(self.stats_selected)
        self.stats_selected = sel
        self._stats_building = True
        try:
            for r, y in enumerate(years):
                it = self.stats_table.item(r, 0)
                if it is not None:
                    it.setCheckState(Qt.Checked if y in sel else Qt.Unchecked)
            if years:
                self._fill_avg_row(len(years))
        finally:
            self._stats_building = False
        self._update_stats_status()

    def _update_stats_status(self):
        years = self._stats_all_years()
        self.stats_status.setText(
            "Showing %s \u00b7 %d year(s) loaded \u00b7 %d selected"
            % (self._metric_key(), len(years), len(self.stats_selected)))

    def get_selected_statistics(self):
        """Public accessor: the data the user has chosen to apply."""
        years = sorted(y for y in self._stats_all_years() if y in self.stats_selected)
        return dict(
            metric=self._metric_key(),
            years=years,
            fire={y: self.stats_fire.get(y) for y in years},
            mileage={y: self.stats_mileage.get(y) for y in years},
            average=self._avg_series(),
        )

    # ----------------------------------------------------------------
    #  Statistics Excel import
    # ----------------------------------------------------------------
    def _load_stats_excel(self):
        if not _HAS_OPENPYXL:
            QMessageBox.information(
                self, "openpyxl required",
                "Reading Excel files needs the 'openpyxl' package.\n\n"
                "Install it with:  pip install openpyxl")
            return
        path, _ = QFileDialog.getOpenFileName(
            self, "Upload fire-accident statistics", "",
            "Excel files (*.xlsx *.xlsm);;All files (*)")
        if not path:
            return
        try:
            res = self._parse_stats_workbook(path)
        except Exception as e:
            QMessageBox.warning(self, "Could not read file", str(e))
            return
        if not res or not (res["Fire"] or res["Mileage"] or res["Rate"]):
            QMessageBox.warning(
                self, "No statistics found",
                "Could not find Fire case / Mileage rows in this workbook.\n\n"
                "Expected a column of years with Passenger-car / Bus / Truck / "
                "Total values, grouped under 'Fire case' and 'Mileage' labels.")
            return
        self.stats_fire = res["Fire"]
        self.stats_mileage = res["Mileage"]
        self.stats_rate = res["Rate"]
        self.stats_selected = set(self._stats_all_years())
        self._refresh_stats()
        self.stats_status.setText(self.stats_status.text() + "  \u2713 " + os.path.basename(path))

    def _parse_stats_workbook(self, path):
        wb = _openpyxl.load_workbook(path, data_only=True)
        for ws in wb.worksheets:
            res = self._parse_stats_sheet(ws)
            if res and (res["Fire"] or res["Mileage"] or res["Rate"]):
                return res
        return None

    @staticmethod
    def _stats_block_of(text):
        t = text.lower()
        if "fire" in t or "\ud654\uc7ac" in t:        # 화재
            return "Fire"
        if "mileage" in t or "\uc8fc\ud589" in t or "\uac70\ub9ac" in t:  # 주행 / 거리
            return "Mileage"
        if "case/year" in t or "case / year" in t or "rate" in t \
                or ("case" in t and "year" in t):
            return "Rate"
        return None

    def _parse_stats_sheet(self, ws):
        year_hdrs = {"\ub144\ub3c4", "\uc5f0\ub3c4", "year", "yr"}  # 년도 / 연도
        maxr = min(ws.max_row, 2000)
        maxc = min(ws.max_column, 60)
        year_col = None
        header_row = 0
        for r in range(1, maxr + 1):
            for c in range(1, maxc + 1):
                v = ws.cell(r, c).value
                if isinstance(v, str) and v.strip().lower() in year_hdrs:
                    year_col, header_row = c, r
                    break
            if year_col:
                break
        if year_col is None:        # fallback: a column with several 4-digit years
            for c in range(1, maxc + 1):
                cnt = sum(1 for r in range(1, maxr + 1)
                          if isinstance(ws.cell(r, c).value, (int, float))
                          and 1900 <= int(ws.cell(r, c).value) <= 2100)
                if cnt >= 3:
                    year_col = c
                    break
        if year_col is None:
            return None

        fire, mile, rate = {}, {}, {}
        cur = None
        start = header_row + 1 if header_row else 1
        for r in range(start, maxr + 1):
            for c in range(1, year_col):     # update current block from label cells
                v = ws.cell(r, c).value
                if isinstance(v, str) and v.strip():
                    b = self._stats_block_of(v)
                    if b:
                        cur = b
                        break
            yv = ws.cell(r, year_col).value
            if cur and isinstance(yv, (int, float)) and 1900 <= int(yv) <= 2100:
                vals = []
                for k in range(1, 5):
                    cv = ws.cell(r, year_col + k).value
                    try:
                        vals.append(float(cv))
                    except (TypeError, ValueError):
                        vals.append(0.0)
                {"Fire": fire, "Mileage": mile, "Rate": rate}[cur][int(yv)] = vals
        return {"Fire": fire, "Mileage": mile, "Rate": rate}

    # ----------------------------------------------------------------
    #  Defaults / table seeding
    # ----------------------------------------------------------------
    def _init_defaults(self):
        self._building = True
        for i in range(N_ROWS):
            left = DEFAULT_CELLS[i] if i < len(DEFAULT_CELLS) else [0, 0, 0, 0]
            for k in range(4):
                self._set_cell(i, k, left[k], editable=True)
            for k in range(4):  # right columns 4..7
                self._set_cell(i, 4 + k, 0, editable=not self.sym)
        self._building = False
        self.set_sym(True, silent=True)

    def _set_cell(self, row, col, value, editable=True):
        item = QTableWidgetItem(str(int(round(float(value or 0)))))
        item.setTextAlignment(Qt.AlignCenter)
        if not editable:
            item.setFlags(item.flags() & ~Qt.ItemIsEditable)
            item.setForeground(Qt.gray)
        self.table.setItem(row, col, item)

    def _cell_val(self, row, col):
        it = self.table.item(row, col)
        if it is None:
            return 0
        try:
            return _jsround(float(it.text()))
        except Exception:
            return 0

    # ----------------------------------------------------------------
    #  Symmetry handling
    # ----------------------------------------------------------------
    def set_sym(self, on, silent=False):
        self.sym = bool(on)
        self.btn_sym_auto.setChecked(self.sym)
        self.btn_sym_free.setChecked(not self.sym)
        # enable/disable right columns
        self._building = True
        for row in range(N_ROWS):
            for col in range(4, 8):
                it = self.table.item(row, col)
                if it is None:
                    self._set_cell(row, col, 0, editable=not self.sym)
                    it = self.table.item(row, col)
                fl = it.flags()
                if self.sym:
                    it.setFlags(fl & ~Qt.ItemIsEditable)
                    it.setForeground(Qt.gray)
                else:
                    it.setFlags(fl | Qt.ItemIsEditable)
                    it.setForeground(Qt.white)
        self._building = False
        if not silent:
            self.update_all()

    def _on_cell_changed(self, row, col):
        if self._building:
            return
        self.update_all()

    # ----------------------------------------------------------------
    #  Hover spinner overlay (▲/▼ to nudge a cell value)
    # ----------------------------------------------------------------
    def _setup_cell_spinner(self):
        self._spin_row = -1
        self._spin_col = -1

        self.spinner = QWidget(self.table.viewport())
        self.spinner.setFixedSize(17, 22)
        self.spinner.setToolTip("Click ▲/▼ or scroll to change this cell")
        lay = QVBoxLayout(self.spinner)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(0)

        self.btn_up = QToolButton(self.spinner)
        self.btn_up.setArrowType(Qt.UpArrow)
        self.btn_dn = QToolButton(self.spinner)
        self.btn_dn.setArrowType(Qt.DownArrow)
        for b in (self.btn_up, self.btn_dn):
            b.setFixedSize(17, 11)
            b.setAutoRepeat(True)
            b.setAutoRepeatDelay(300)
            b.setAutoRepeatInterval(70)
            b.setFocusPolicy(Qt.NoFocus)
            b.setStyleSheet(
                "QToolButton{background:#2b3a4f;border:1px solid #45597a;margin:0;}"
                "QToolButton:hover{background:#3d72a4;}"
                "QToolButton:pressed{background:#2f6285;}"
                "QToolButton:disabled{background:#222b38;border-color:#2c3a4c;}")
            lay.addWidget(b)
        self.btn_up.clicked.connect(lambda: self._spin_step(+1))
        self.btn_dn.clicked.connect(lambda: self._spin_step(-1))
        self.spinner.hide()

        # observe mouse movement over the table to show/move the spinner
        self.table.setMouseTracking(True)
        self.table.viewport().setMouseTracking(True)
        self.table.viewport().installEventFilter(self)
        self.spinner.installEventFilter(self)
        # hide it whenever the table scrolls (positions would go stale)
        self.table.verticalScrollBar().valueChanged.connect(self._force_hide_spinner)
        self.table.horizontalScrollBar().valueChanged.connect(self._force_hide_spinner)

    def eventFilter(self, obj, event):
        try:
            if obj is self.table.viewport():
                et = event.type()
                if et == QEvent.MouseMove:
                    self._update_spinner(event.pos())
                elif et == QEvent.Leave:
                    self._hide_spinner()
            elif obj is self.spinner:
                et = event.type()
                if et == QEvent.Wheel:
                    self._spin_step(1 if event.angleDelta().y() > 0 else -1)
                    return True
                elif et == QEvent.Leave:
                    # left the spinner; hide only if not over the table
                    vp = self.table.viewport()
                    if not vp.rect().contains(vp.mapFromGlobal(QCursor.pos())):
                        self._force_hide_spinner()
        except RuntimeError:
            # underlying C++ widget already deleted (e.g. during teardown)
            return False
        return super().eventFilter(obj, event)

    def _update_spinner(self, pos):
        idx = self.table.indexAt(pos)
        if not idx.isValid():
            self._hide_spinner()
            return
        it = self.table.item(idx.row(), idx.column())
        if it is None:
            self._hide_spinner()
            return
        self._spin_row, self._spin_col = idx.row(), idx.column()
        rect = self.table.visualItemRect(it)
        sw, sh = self.spinner.width(), self.spinner.height()
        x = rect.right() - sw - 2
        y = rect.top() + (rect.height() - sh) // 2
        self.spinner.move(max(0, x), max(0, y))
        editable = bool(it.flags() & Qt.ItemIsEditable)
        self.btn_up.setEnabled(editable)
        self.btn_dn.setEnabled(editable)
        self.spinner.show()
        self.spinner.raise_()

    def _hide_spinner(self):
        # keep it visible if the pointer is actually hovering the spinner
        if self.spinner.isVisible():
            if self.spinner.rect().contains(self.spinner.mapFromGlobal(QCursor.pos())):
                return
        self._force_hide_spinner()

    def _force_hide_spinner(self, *_):
        self.spinner.hide()
        self._spin_row = self._spin_col = -1

    def _spin_step(self, delta):
        r, c = self._spin_row, self._spin_col
        if r < 0 or c < 0:
            return
        it = self.table.item(r, c)
        if it is None or not (it.flags() & Qt.ItemIsEditable):
            return
        try:
            val = int(round(float(it.text())))
        except Exception:
            val = 0
        val = max(0, val + int(delta))
        it.setText(str(val))   # fires cellChanged -> update_all (+ symmetry)

    # ----------------------------------------------------------------
    #  Fire preset handling
    # ----------------------------------------------------------------
    def _on_preset_changed(self, idx):
        key = {0: "20", 1: "30", 2: "100"}.get(idx)
        if key:
            p = PRESETS[key]
            for f, val in zip(
                (self.in_hrr, self.in_bhl, self.in_bz, self.in_bo1, self.in_bo2), p):
                f.blockSignals(True)
                f.setText(fmt(val))
                f.blockSignals(False)
        self.update_all()

    def _fire_field_edited(self, *_):
        # user typed in a fire field -> switch preset to Custom
        if self.cmb_preset.currentIndex() != 3:
            self.cmb_preset.blockSignals(True)
            self.cmb_preset.setCurrentIndex(3)
            self.cmb_preset.blockSignals(False)
        self.update_all()

    # ----------------------------------------------------------------
    #  numeric getters
    # ----------------------------------------------------------------
    def _num(self, field, default=0.0):
        try:
            return float(field.text())
        except Exception:
            return default

    def _bc_value(self):
        return str(self.cmb_bc.currentIndex() + 1)

    # ----------------------------------------------------------------
    #  Core computation (mirrors compute() in the HTML)
    # ----------------------------------------------------------------
    def compute(self):
        W = self._num(self.in_W)
        H = self._num(self.in_H)
        nW = max(1, _jsround(self._num(self.in_nW)))
        nH = max(1, _jsround(self._num(self.in_nH)))
        Lh = self._num(self.in_Lh)
        Lfull = Lh * 2
        dx = self._num(self.in_dx) or 1
        nLen = max(1, _jsround(Lfull / dx))
        cW = W / nW if nW else 0
        cH = H / nH if nH else 0

        # auto-symmetry: recompute & write back right cells from left
        if self.sym:
            self._building = True
            for i in range(N_ROWS):
                lv = [self._cell_val(i, k) for k in range(4)]
                rv = [nW - lv[1], nW - lv[0], lv[2], lv[3]]
                for k in range(4):
                    self._set_cell(i, 4 + k, rv[k], editable=False)
            self._building = False

        left = [[self._cell_val(i, k) for k in range(4)] for i in range(N_ROWS)]
        right = [[self._cell_val(i, 4 + k) for k in range(4)] for i in range(N_ROWS)]

        breath = self._num(self.in_breath) or 1.8
        kBreath = max(0, _jsround(breath / cH)) if cH > 0 else 0
        breathSnap = kBreath * cH if cH > 0 else breath

        return dict(W=W, H=H, nW=nW, nH=nH, Lh=Lh, Lfull=Lfull, dx=dx,
                    nLen=nLen, cW=cW, cH=cH, breath=breath,
                    kBreath=kBreath, breathSnap=breathSnap,
                    left=left, right=right)

    # ----------------------------------------------------------------
    #  FDS deck generation (mirrors buildFDS())
    # ----------------------------------------------------------------
    def _cfg_manual(self):
        title = {0: "20MW", 1: "30MW", 2: "100MW"}.get(
            self.cmb_preset.currentIndex(), "TN")
        return dict(
            title=title,
            hrr=self._num(self.in_hrr),
            hl=self._num(self.in_bhl),
            bz=self._num(self.in_bz),
            o1=self._num(self.in_bo1),
            o2=self._num(self.in_bo2),
            profile=None,
            bc=self._bc_value(),
            twfin=self._num(self.in_twfin) or 1200,
            v0=self._num(self.in_v0),
            breath=self._num(self.in_breath) or 1.8,
            sign=1,
        )

    @staticmethod
    def _obst_line(m, c, tag):
        return ("&OBST XB= 0.0000, %s, %s, %s, %s, %s, COLOR='GRAY' / %s"
                % (fmt(m["Lfull"]), fmt(m["cW"] * c[0]), fmt(m["cW"] * c[1]),
                   fmt(m["cH"] * c[2]), fmt(m["cH"] * c[3]), tag))

    def build_fds(self, m, cfg):
        T = cfg["twfin"] or 1200
        hrr, hl, bz = cfg["hrr"], cfg["hl"], cfg["bz"]
        o1, o2, v0 = cfg["o1"], cfg["o2"], cfg["v0"]
        sg = cfg.get("sign", 1) or 1
        prof = ([[t, v * sg] for t, v in cfg["profile"]]
                if cfg.get("profile") else None)
        bc = decide_case(prof) if prof else cfg["bc"]

        fx1, fx2 = m["Lh"] - hl, m["Lh"] + hl
        fy1, fy2 = m["W"] / 2 - o1, m["W"] / 2 + o2
        pby = m["W"] / 2
        breath = cfg["breath"] or 1.8
        pbz = (max(0, _jsround(breath / m["cH"])) * m["cH"]) if m["cH"] > 0 else breath

        def cm(c):
            return "" if c == bc else "/"

        L = []
        L.append("&HEAD CHID='%s', TITLE='%s' /" % (safe_name(cfg["title"]), cfg["title"]))
        L.append("")
        L.append("&MESH IJK=%d,%d,%d,XB=0.000,%s,0.0000,%s,0.0000,%s,SYNCHRONIZE=.TRUE./"
                 % (m["nLen"], m["nW"], m["nH"], fmt(m["Lfull"]), fmt(m["W"]), fmt(m["H"])))
        L.append("")
        L.append("&TIME TWFIN = %.1f/" % T)
        L.append("")
        L.append("&MISC TMPA= 20.0, GVEC=1.0, 0.0, 1.0, RAMP_GX='X GRADE', RAMP_GZ='Z GRADE' /")
        L.append("&RAMP ID='X GRADE', X=    0.0, F=0.0/")
        L.append("&RAMP ID='X GRADE', X= %s, F=0.0 /" % fmt(m["Lfull"]))
        L.append("&RAMP ID='Z GRADE', X=0.0, F=-9.81/")
        L.append("&RAMP ID='Z GRADE', X=%s, F=-9.81/" % fmt(m["Lfull"]))
        L.append("")
        L.append("&REAC ID='FIRE', SOOT_YIELD=0.133, CO_YIELD=0.168, MASS_EXTINCTION_COEFFICIENT=8700. /")
        L.append("&RADI RADIATIVE_FRACTION = 0.3/")
        L.append("")
        L.append("/**********TUNNELBLOCK**********/")
        for i, c in enumerate(m["left"]):
            if c[0] != c[1] and c[2] != c[3]:
                L.append(self._obst_line(m, c, "*L%d" % (i + 1)))
        for i, c in enumerate(m["right"]):
            if c[0] != c[1] and c[2] != c[3]:
                L.append(self._obst_line(m, c, "*R%d" % (i + 1)))
        L.append("")
        L.append("&SURF ID='BURNER1', HRRPUA=%s, RAMP_Q='Fire_Ramp', COLOR='RASPBERRY'/, PART_ID='smoke' /" % fmt(hrr))
        L.append("&SURF ID='BURNER2', HRRPUA=%s, RAMP_Q='Fire_Ramp', COLOR='RASPBERRY'/, PART_ID='smoke' /" % fmt(hrr))
        L.append("")
        L.append("&VENT XB= %s, %s, %s, %s, 0.0, %s, SURF_ID='BURNER1'/ fire"
                 % (fmt(fx1), fmt(fx2), fmt(fy1), fmt(fy1), fmt(bz)))
        L.append("&VENT XB= %s, %s, %s, %s, 0.0, %s, SURF_ID='BURNER2'/ fire"
                 % (fmt(fx1), fmt(fx2), fmt(fy2), fmt(fy2), fmt(bz)))
        L.append("&OBST XB= %s, %s, %s, %s, 0.0, %s, COLOR='FIREBRICK'/ fire car"
                 % (fmt(fx1), fmt(fx2), fmt(fy1), fmt(fy2), fmt(bz)))
        L.append("")
        L.append("/ Fire intensity RAMP (Fire_Ramp)")
        full = list(FIRE)
        t = 458
        while t <= 1508:
            full.append([t, 1])
            t += 10
        full.append([1800, 1])
        for tt, ff in full:
            L.append("&RAMP ID='Fire_Ramp',T=%sF=%s/"
                     % (("%.1f" % tt).rjust(7) + ",", ("%.4f" % ff).rjust(7)))
        L.append("")
        L.append("///Tunnel Inlet Definition%s"
                 % (("  (velocity applied · CASE%s)" % bc) if prof else ""))
        L.append("//CASE 1 ***** initial velocity + OPEN BC *****")
        L.append("%s&MISC V0 = %s " % (cm("1"), fmt(v0)))
        L.append("%s&VENT MB='XMAX', SURF_ID='OPEN'/ RIGHT" % cm("1"))
        L.append("%s&VENT MB='XMIN', SURF_ID='OPEN'/ LEFT" % cm("1"))
        L.append("")
        L.append("//CASE 2 ***** Pressure Dynamic BC *****")
        L.append("%s&VENT MB='YMAX', SURF_ID='OPEN'/ RIGHT" % cm("2"))
        L.append("%s&VENT MB='XMIN', SURF_ID='OPEN', DYNAMIC_PRESSURE=1.000, PRESSURE_RAMP='Wind'/ LEFT" % cm("2"))
        L.append("%s&RAMP ID='Wind',T=    0,   F= 116.64450 /" % cm("2"))
        L.append("%s&RAMP ID='Wind',T=  300,   F= 116.64450 /" % cm("2"))
        L.append("")
        ramp_data = prof if prof else VR
        vel3 = "1.00" if prof else "-1.00"
        L.append("//CASE 3 ***** Velocity Profile (Vr_Ramp) *****")
        L.append("%s&VENT MB='XMIN', SURF_ID='LEFT_PORTAL'/ LEFT" % cm("3"))
        L.append("%s&SURF ID='LEFT_PORTAL', VEL = %s, RGB=1,0,0, RAMP_V ='Vr_Ramp'/" % (cm("3"), vel3))
        L.append("%s&VENT MB='XMAX', SURF_ID='OPEN' /" % cm("3"))
        if bc == "3":
            for tt, vv in ramp_data:
                L.append("&RAMP ID='Vr_Ramp', T= %s , F= %s/" % (fmt(tt), fmt(vv)))
        L.append("")
        L.append("//CASE 4 ***** Velocity Profile (−) *****")
        L.append("%s&VENT MB='XMIN', SURF_ID='OPEN' /" % cm("4"))
        L.append("%s&VENT MB='XMAX', SURF_ID='RIGHT_PORTAL'/ RIGHT" % cm("4"))
        L.append("%s&SURF ID='RIGHT_PORTAL', VEL = 1.00, RGB=1,0,0, RAMP_V ='Vr_Ramp'/" % cm("4"))
        if bc == "4":
            for tt, vv in ramp_data:
                L.append("&RAMP ID='Vr_Ramp', T= %s , F= %s/" % (fmt(tt), fmt(vv)))
        L.append("")
        L.append("///Tunnel Inlet Definition - END")
        L.append("")
        L.append("&DUMP DT_PL3D=30.0, DT_BNDF=30, DT_SLCF=30.0, PLOT3D_QUANTITY(2)='VELOCITY', "
                 "PLOT3D_QUANTITY(3)='carbon monoxide', PLOT3D_QUANTITY(4)='carbon dioxide', "
                 "PLOT3D_QUANTITY(5)='soot density', WRITE_XYZ=.TRUE./")
        L.append("")
        for i, q in enumerate(["soot density", "carbon dioxide", "carbon monoxide",
                               "TEMPERATURE", "INTEGRATED INTENSITY", "oxygen"]):
            L.append("&SLCF PBZ = %s, QUANTITY='%s'/%s"
                     % (fmt(pbz), q, " breath level" if i == 0 else ""))
        L.append("")
        for i, q in enumerate(["soot density", "carbon dioxide", "carbon monoxide",
                               "TEMPERATURE", "INTEGRATED INTENSITY"]):
            L.append("&SLCF PBY = %s, QUANTITY='%s'/%s"
                     % (fmt(pby), q, " tunnel center" if i == 0 else ""))
        L.append("")
        L.append("&TAIL /")
        return "\n".join(L)

    # ----------------------------------------------------------------
    #  Master update: recompute -> derived chips -> preview -> FDS text
    # ----------------------------------------------------------------
    def update_all(self, *_):
        if self._building:
            return
        # re-apply DXF wall bands when the grid geometry changes
        if self.dxf_poly is not None:
            sig = (self._num(self.in_W), self._num(self.in_H),
                   _jsround(self._num(self.in_nW)), _jsround(self._num(self.in_nH)))
            if sig != self._last_geom_sig:
                self._last_geom_sig = sig
                self._apply_dxf_bands()
        try:
            m = self.compute()
        except Exception:
            return
        self._render_derived(m)
        self._draw(m)
        if 0 <= self.active < len(self.files):
            cfg = self._cfg_for_file(self.files[self.active])
        else:
            cfg = self._cfg_manual()
        fds = self.build_fds(m, cfg)
        self._last_fds = fds
        self._last_fname = safe_name(cfg["title"]) + ".fds"
        self.code.setPlainText(fds)

    def _render_derived(self, m):
        total = m["nLen"] * m["nW"] * m["nH"]
        chips = [
            ("Cell width", "%s m" % fmt(m["cW"])),
            ("Cell height", "%s m" % fmt(m["cH"])),
            ("Length Δx", "%s m" % fmt(m["dx"])),
            ("Full length", "%s m" % fmt(m["Lfull"])),
            ("MESH cells", "%d·%d·%d" % (m["nLen"], m["nW"], m["nH"])),
            ("Total cells", "{:,}".format(total)),
            ("Clear area ≈", "%s m²" % fmt(m["W"] * m["H"])),
            ("Breath snap", "%s m  (cell %d, target %s m)"
             % (fmt(m["breathSnap"]), m["kBreath"], fmt(m["breath"]))),
        ]
        txt = "   ".join("%s: %s" % (k, v) for k, v in chips)
        self.derived_lbl.setText(txt)
        self.lbl_breath_snap.setText(
            "%s  (cell %d · target %s m)" % (fmt(m["breathSnap"]), m["kBreath"], fmt(m["breath"])))

    # ----------------------------------------------------------------
    #  Cross-section preview
    # ----------------------------------------------------------------
    def _draw(self, m):
        if not _HAS_MPL or self.canvas is None:
            return
        ax = self.ax
        ax.clear()
        nW, nH = m["nW"], m["nH"]
        W, H = m["W"], m["H"]
        ax.set_facecolor("#0f1722")

        # occupancy grid (cell indices)
        occ = [[False] * nH for _ in range(nW)]

        def fill(c):
            if all(v == 0 for v in c):
                return
            for i in range(int(c[0]), int(c[1])):
                for j in range(int(c[2]), int(c[3])):
                    if 0 <= i < nW and 0 <= j < nH:
                        occ[i][j] = True

        for c in m["left"]:
            fill(c)
        for c in m["right"]:
            fill(c)

        cw = W / nW if nW else 0
        ch = H / nH if nH else 0
        # wall cells
        for i in range(nW):
            for j in range(nH):
                if occ[i][j]:
                    ax.add_patch(Rectangle((i * cw, j * ch), cw, ch,
                                           facecolor="#5b6675", edgecolor="#404b59",
                                           linewidth=0.5))
        # grid lines
        for i in range(nW + 1):
            ax.axvline(i * cw, color="#1f3b52",
                       lw=1 if i % 5 == 0 else 0.4, zorder=0)
        for j in range(nH + 1):
            ax.axhline(j * ch, color="#1f3b52",
                       lw=1 if j % 5 == 0 else 0.4, zorder=0)
        # centerline
        ax.axvline(W / 2, color="#2f6285", lw=0.9, ls="--")
        # fire vehicle
        o1 = self._num(self.in_bo1)
        o2 = self._num(self.in_bo2)
        bz = self._num(self.in_bz)
        fy1, fy2 = W / 2 - o1, W / 2 + o2
        ax.add_patch(Rectangle((fy1, 0), fy2 - fy1, bz,
                               facecolor="#f4823c", edgecolor="#c2410c",
                               alpha=0.92, zorder=3))
        ax.text(W / 2, bz + H * 0.02, "fire", color="#ffb347",
                ha="center", va="bottom", fontsize=8)

        # DXF cross-section outline (drawn in metres, already origin-shifted)
        if self.dxf_poly:
            xs = [p["x"] for p in self.dxf_poly] + [self.dxf_poly[0]["x"]]
            ys = [p["y"] for p in self.dxf_poly] + [self.dxf_poly[0]["y"]]
            ax.plot(xs, ys, color="#a29bfe", lw=1.6, zorder=5)

        ax.set_xlim(-W * 0.02, W * 1.02)
        ax.set_ylim(-H * 0.02, H * 1.05)
        ax.set_aspect("equal", adjustable="box")
        ax.set_xlabel("W = %s m" % fmt(W), color="#8aa0b8", fontsize=8)
        ax.set_ylabel("H = %s m" % fmt(H), color="#8aa0b8", fontsize=8)
        for spine in ax.spines.values():
            spine.set_color("#243246")
        ax.tick_params(colors="#5a6b7d", labelsize=7)
        self.fig.tight_layout()
        self.canvas.draw_idle()

    # ----------------------------------------------------------------
    #  DXF cross-section import
    # ----------------------------------------------------------------
    @staticmethod
    def _dxf_bulge_arc(p0, p1, bulge, seg=16):
        """Tessellate a bulged DXF segment into points (excluding p0)."""
        x0, y0 = p0
        x1, y1 = p1
        if abs(bulge) < 1e-9:
            return [(x1, y1)]
        theta = 4.0 * math.atan(bulge)            # included angle
        dx, dy = x1 - x0, y1 - y0
        chord = math.hypot(dx, dy)
        if chord < 1e-12:
            return [(x1, y1)]
        radius = chord / (2.0 * math.sin(abs(theta) / 2.0))
        # midpoint of chord
        mx, my = (x0 + x1) / 2.0, (y0 + y1) / 2.0
        # distance from midpoint to centre
        h = math.sqrt(max(0.0, radius * radius - (chord / 2.0) ** 2))
        # unit normal (sign depends on bulge direction)
        nx, ny = -dy / chord, dx / chord
        sgn = 1.0 if bulge > 0 else -1.0
        cx, cy = mx - sgn * h * nx, my - sgn * h * ny
        a0 = math.atan2(y0 - cy, x0 - cx)
        a1 = math.atan2(y1 - cy, x1 - cx)
        # sweep direction
        if bulge > 0:
            if a1 < a0:
                a1 += 2 * math.pi
        else:
            if a1 > a0:
                a1 -= 2 * math.pi
        pts = []
        n = max(2, int(seg))
        for k in range(1, n + 1):
            a = a0 + (a1 - a0) * k / n
            pts.append((cx + radius * math.cos(a), cy + radius * math.sin(a)))
        return pts

    def _dxf_flatten(self, verts, closed):
        """Turn a list of {x,y,b} vertices into a flat ring of (x,y)."""
        ring = []
        n = len(verts)
        if n == 0:
            return ring
        segs = n if closed else n - 1
        ring.append((verts[0]["x"], verts[0]["y"]))
        for i in range(segs):
            a = verts[i]
            b = verts[(i + 1) % n]
            pts = self._dxf_bulge_arc((a["x"], a["y"]), (b["x"], b["y"]), a.get("b", 0.0))
            ring.extend(pts)
        return ring

    def _dxf_parse(self, text):
        """Parse DXF text, return the largest closed polygon as a dict
        {ring,x0,x1,y0,y1} (ring is a list of (x,y)), or None."""
        lines = text.replace("\r\n", "\n").replace("\r", "\n").split("\n")
        pairs = []
        i = 0
        while i + 1 < len(lines):
            raw_code = lines[i].strip()
            val = lines[i + 1]
            try:
                code = int(raw_code)
            except Exception:
                i += 2
                continue
            pairs.append((code, val.strip()))
            i += 2

        polys = []          # each: {'verts':[{x,y,b}], 'closed':bool}
        cur = None          # current LWPOLYLINE / POLYLINE being built
        kind = None         # 'LW' | 'POLY' | None
        in_vertex = False   # inside a VERTEX sub-entity (old-style POLYLINE)
        pend_x = None

        def flush():
            nonlocal cur, kind, in_vertex
            if cur and len(cur["verts"]) >= 2:
                polys.append(cur)
            cur = None
            kind = None
            in_vertex = False

        for code, val in pairs:
            if code == 0:
                v = val.upper()
                if v == "VERTEX" and kind == "POLY":
                    cur["verts"].append({"x": 0.0, "y": 0.0, "b": 0.0})
                    in_vertex = True
                    continue
                if v == "SEQEND":
                    in_vertex = False
                    continue
                # any other new entity terminates the current poly
                flush()
                if v == "LWPOLYLINE":
                    cur = {"verts": [], "closed": False}
                    kind = "LW"
                    pend_x = None
                elif v == "POLYLINE":
                    cur = {"verts": [], "closed": False}
                    kind = "POLY"
                    pend_x = None
                continue

            if cur is None:
                continue

            if kind == "LW":
                if code == 10:
                    pend_x = float(val)
                elif code == 20:
                    cur["verts"].append({"x": pend_x or 0.0, "y": float(val), "b": 0.0})
                elif code == 42 and cur["verts"]:
                    cur["verts"][-1]["b"] = float(val)
                elif code == 70 and (int(float(val)) & 1):
                    cur["closed"] = True
            elif kind == "POLY":
                if code == 70 and (int(float(val)) & 1):
                    cur["closed"] = True
                if in_vertex and cur["verts"]:
                    if code == 10:
                        cur["verts"][-1]["x"] = float(val)
                    elif code == 20:
                        cur["verts"][-1]["y"] = float(val)
                    elif code == 42:
                        cur["verts"][-1]["b"] = float(val)
        flush()

        best = None
        best_area = -1.0
        for p in polys:
            if not p["closed"]:
                continue
            ring = self._dxf_flatten(p["verts"], True)
            if len(ring) < 3:
                continue
            xs = [pt[0] for pt in ring]
            ys = [pt[1] for pt in ring]
            area = (max(xs) - min(xs)) * (max(ys) - min(ys))
            if area > best_area:
                best_area = area
                best = dict(ring=ring, x0=min(xs), x1=max(xs),
                            y0=min(ys), y1=max(ys))
        return best

    @staticmethod
    def _pip(x, y, ring):
        """Ray-cast point-in-polygon test. ring is a list of (x,y)."""
        inside = False
        n = len(ring)
        j = n - 1
        for i in range(n):
            xi, yi = ring[i]
            xj, yj = ring[j]
            if ((yi > y) != (yj > y)) and \
               (x < (xj - xi) * (y - yi) / ((yj - yi) or 1e-12) + xi):
                inside = not inside
            j = i
        return inside

    @staticmethod
    def _merge_bands(counts, nH):
        """Merge contiguous rows that share the same nonzero count into
        [count, y1, y2] bands."""
        bands = []
        j = 0
        while j < nH:
            c = counts[j]
            k = j
            while k + 1 < nH and counts[k + 1] == c:
                k += 1
            if c > 0:
                bands.append([c, j, k + 1])
            j = k + 1
        return bands

    def _apply_dxf_bands(self):
        """Fill the wall-cell table from the loaded DXF polygon for the
        current grid.  Returns the number of bands written."""
        if not self.dxf_poly:
            return 0
        ring = [(p["x"], p["y"]) for p in self.dxf_poly]
        W = self._num(self.in_W)
        H = self._num(self.in_H)
        nW = max(1, _jsround(self._num(self.in_nW)))
        nH = max(1, _jsround(self._num(self.in_nH)))
        cW = W / nW if nW else 0
        cH = H / nH if nH else 0
        left_counts = [0] * nH
        right_counts = [0] * nH
        for j in range(nH):
            yc = (j + 0.5) * cH
            inside_idx = [i for i in range(nW)
                          if self._pip((i + 0.5) * cW, yc, ring)]
            if inside_idx:
                first = inside_idx[0]
                last = inside_idx[-1]
                left_counts[j] = first             # cells of wall on the left
                right_counts[j] = nW - 1 - last     # cells of wall on the right
            else:
                left_counts[j] = nW                 # whole row is wall
                right_counts[j] = nW
        sym = all(left_counts[j] == right_counts[j] for j in range(nH))

        self._building = True
        # silent symmetry set
        self.sym = sym
        self.btn_sym_auto.setChecked(sym)
        self.btn_sym_free.setChecked(not sym)
        # clear all rows
        for r in range(N_ROWS):
            for c in range(8):
                self._set_cell(r, c, 0, editable=(c < 4) or (not sym))
        left_bands = self._merge_bands(left_counts, nH)
        for idx, (cnt, y1, y2) in enumerate(left_bands[:N_ROWS]):
            for k, v in enumerate((0, cnt, y1, y2)):
                self._set_cell(idx, k, v, editable=True)
        if not sym:
            right_bands = self._merge_bands(right_counts, nH)
            for idx, (cnt, y1, y2) in enumerate(right_bands[:N_ROWS]):
                for k, v in enumerate((nW - cnt, nW, y1, y2)):
                    self._set_cell(idx, 4 + k, v, editable=True)
        self._building = False
        return len(left_bands)

    def _on_dxf_load(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Load DXF cross-section", "", "DXF files (*.dxf);;All files (*)")
        if not path:
            return
        try:
            with open(path, "r", encoding="utf-8", errors="ignore") as fh:
                text = fh.read()
        except Exception as e:
            QMessageBox.warning(self, "DXF load failed", str(e))
            return
        best = self._dxf_parse(text)
        if not best:
            QMessageBox.warning(
                self, "No section found",
                "No closed LWPOLYLINE / POLYLINE was found in this DXF.")
            return
        self.dxf_w = best["x1"] - best["x0"]
        self.dxf_h = best["y1"] - best["y0"]
        self.dxf_poly = [{"x": x - best["x0"], "y": y - best["y0"]}
                         for (x, y) in best["ring"]]
        # adopt DXF bounding box as W/H without triggering update mid-flight
        self._building = True
        self.in_W.setText(fmt(round(self.dxf_w * 1e4) / 1e4))
        self.in_H.setText(fmt(round(self.dxf_h * 1e4) / 1e4))
        self._building = False
        self._last_geom_sig = None      # force band re-apply
        nb = self._apply_dxf_bands()
        self.btn_dxf_clear.show()
        self.dxf_info.setText(
            "✓ %s · section W %s × H %s m · %d wall band(s)"
            % (os.path.basename(path), fmt(self.dxf_w), fmt(self.dxf_h), nb))
        self.dxf_info.setStyleSheet("color:#3fb98a;font-size:11px;")
        self.update_all()

    def _on_dxf_clear(self):
        self.dxf_poly = None
        self.dxf_w = self.dxf_h = 0.0
        self._last_geom_sig = None
        self.btn_dxf_clear.hide()
        self.dxf_info.setText(
            "Import a closed cross-section polyline (LWPOLYLINE / POLYLINE). "
            "Sets W·H and fills wall cells by point-in-polygon.")
        self.dxf_info.setStyleSheet("color:#8a7ddb;font-size:11px;")
        self.update_all()

    # ----------------------------------------------------------------
    #  .VRR profile parsing
    # ----------------------------------------------------------------
    @staticmethod
    def _is_num(v):
        return isinstance(v, float) and math.isfinite(v)

    def _parse_text(self, text):
        """Parse a .VRR/.txt body into an array of rows.  Rows are either
        [T, F] pairs (from 'T=.. F=..') or lists of mixed float/str tokens."""
        aoa = []
        for raw in text.replace("\r\n", "\n").replace("\r", "\n").split("\n"):
            ln = raw.strip()
            if not ln:
                continue
            # skip comment lines that contain no digits
            if re.match(r"^[#/!*;%]", ln) and not re.search(r"\d", ln):
                continue
            tf = re.findall(
                r"\b([TF])\s*=\s*(-?\d+(?:\.\d+)?(?:[eE][+-]?\d+)?)", ln)
            if len(tf) >= 2:
                t = f = None
                for code, val in tf:
                    if code == "T":
                        t = float(val)
                    else:
                        f = float(val)
                if t is not None and f is not None:
                    aoa.append([t, f])
                    continue
            toks = re.split(r"[\s,;\t]+", ln)
            row = []
            for p in toks:
                try:
                    row.append(float(p))
                except Exception:
                    row.append(p)
            aoa.append(row)
        return aoa

    def _vrr_to_profile(self, text):
        """Return (profile, single) where profile is a list of [t, v] or
        None, and single indicates a synthesised time axis."""
        aoa = self._parse_text(text)
        rows = [r for r in aoa if any(self._is_num(x) for x in r)]
        if not rows:
            return None, False
        maxc = max(len(r) for r in rows)
        num_cols = []
        for c in range(maxc):
            cnt = sum(1 for r in rows if c < len(r) and self._is_num(r[c]))
            if cnt >= max(2, int(len(rows) * 0.5)):
                num_cols.append(c)
        if len(num_cols) >= 2:
            tc, vc = num_cols[0], num_cols[1]
            pairs = [[r[tc], r[vc]] for r in rows
                     if tc < len(r) and vc < len(r)
                     and self._is_num(r[tc]) and self._is_num(r[vc])]
            return (pairs or None), False
        if len(num_cols) == 1:
            vc = num_cols[0]
            dt = self._num(self.in_dt) or 20
            pairs = []
            i = 0
            for r in rows:
                if vc < len(r) and self._is_num(r[vc]):
                    pairs.append([i * dt, r[vc]])
                    i += 1
            return (pairs or None), True
        return None, False

    @staticmethod
    def _mw_from_title(title):
        m = re.match(r"^\s*(\d{3})", title)
        if not m:
            return None
        n = int(m.group(1))
        return n if n in (20, 30, 100) else None

    @staticmethod
    def _case_for_file(f):
        if f.get("profile"):
            return decide_case(f["profile"])
        return "3"

    def _cfg_for_file(self, f):
        preset = PRESETS.get(str(f.get("mw")), PRESETS["20"])
        return dict(
            title=f["title"],
            hrr=preset[0], hl=preset[1], bz=preset[2],
            o1=preset[3], o2=preset[4],
            profile=f.get("profile"),
            bc="3",
            twfin=self._num(self.in_twfin) or 1200,
            v0=self._num(self.in_v0),
            breath=self._num(self.in_breath) or 1.8,
            sign=self.sign,
        )

    # ----------------------------------------------------------------
    #  .VRR queue handling
    # ----------------------------------------------------------------
    def _on_add_vrr(self):
        paths, _ = QFileDialog.getOpenFileNames(
            self, "Add velocity profile files", "",
            "Velocity profiles (*.vrr *.txt *.csv *.dat);;All files (*)")
        if not paths:
            return
        added = 0
        for path in paths:
            try:
                with open(path, "r", encoding="utf-8", errors="ignore") as fh:
                    text = fh.read()
            except Exception:
                continue
            title = os.path.splitext(os.path.basename(path))[0]
            profile, single = self._vrr_to_profile(text)
            self.files.append(dict(
                name=os.path.basename(path), title=title,
                mw=self._mw_from_title(title), profile=profile, single=single))
            added += 1
        if added and self.active < 0:
            self.active = 0
        self._render_queue()
        self.update_all()

    def _render_queue(self):
        self.queue.blockSignals(True)
        self.queue.clear()
        for i, f in enumerate(self.files):
            mw = ("%dMW" % f["mw"]) if f.get("mw") else "—"
            if f.get("profile"):
                case = "CASE" + self._case_for_file(f)
                npts = "%dpts" % len(f["profile"])
            else:
                case = "no profile"
                npts = "0pts"
            prefix = "▶ " if i == self.active else "   "
            it = QListWidgetItem("%s[%s] %s · %s · %s" % (prefix, mw, f["title"], case, npts))
            self.queue.addItem(it)
            if i == self.active:
                it.setSelected(True)
        self.queue.blockSignals(False)

        has = len(self.files) > 0
        self.btn_save_all.setEnabled(has)
        self._lock_bc(has)
        if has and 0 <= self.active < len(self.files):
            self._draw_vchart(self.files[self.active])
        elif self.vcanvas is not None:
            self.vcanvas.hide()

    def _on_queue_clicked(self, item):
        self.active = self.queue.row(item)
        f = self.files[self.active]
        # reflect the file's MW preset in the fire panel
        if f.get("mw"):
            idx = {20: 0, 30: 1, 100: 2}.get(f["mw"])
            if idx is not None:
                self.cmb_preset.setCurrentIndex(idx)   # triggers preset fill
        self._render_queue()
        self.update_all()

    def _remove_file(self):
        row = self.queue.currentRow()
        if row < 0 or row >= len(self.files):
            return
        self.files.pop(row)
        if not self.files:
            self.active = -1
        elif self.active >= len(self.files):
            self.active = len(self.files) - 1
        self._render_queue()
        self.update_all()

    def _clear_files(self):
        self.files = []
        self.active = -1
        self._render_queue()
        self.update_all()

    def _lock_bc(self, locked):
        """When files are loaded the inlet CASE is auto-determined per file."""
        self.cmb_bc.setEnabled(not locked)
        self.in_v0.setEnabled(not locked)
        if locked:
            self.lock_note.setText(
                "⛔ Batch mode: inlet CASE is auto-selected per file from its "
                "velocity-profile sign (CASE 3 / CASE 4).  Manual CASE & V0 are "
                "disabled.  The output shows the active file's deck; use "
                "“Save All (ZIP)” to export every file.")
            self.lock_note.show()
        else:
            self.lock_note.hide()

    def _on_sign_changed(self, idx):
        self.sign = 1 if idx == 0 else -1
        if 0 <= self.active < len(self.files):
            self._draw_vchart(self.files[self.active])
        self.update_all()

    # ----------------------------------------------------------------
    #  Velocity-profile chart
    # ----------------------------------------------------------------
    def _draw_vchart(self, f):
        if self.vcanvas is None:
            return
        prof = f.get("profile")
        if not prof:
            self.vcanvas.hide()
            return
        self.vcanvas.show()
        ax = self.vax
        ax.clear()
        ax.set_facecolor("#0f1722")
        ts = [p[0] for p in prof]
        vs = [p[1] * self.sign for p in prof]
        ax.plot(ts, vs, color="#16d6a8", lw=1.4)
        ax.axhline(0, color="#3a4b5c", lw=0.7)
        ax.scatter([ts[0]], [vs[0]], color="#2ecc71", s=22, zorder=5)
        ax.scatter([ts[-1]], [vs[-1]], color="#e74c3c", s=22, zorder=5)
        ax.set_title("velocity profile · %s  (CASE%s)"
                     % (f["title"], self._case_for_file(f)),
                     color="#8aa0b8", fontsize=8)
        for spine in ax.spines.values():
            spine.set_color("#243246")
        ax.tick_params(colors="#5a6b7d", labelsize=7)
        ax.set_xlabel("t (s)", color="#5a6b7d", fontsize=7)
        ax.set_ylabel("v (m/s)", color="#5a6b7d", fontsize=7)
        self.vfig.tight_layout()
        self.vcanvas.draw_idle()

    # ----------------------------------------------------------------
    #  Batch ZIP export
    # ----------------------------------------------------------------
    @staticmethod
    def _fds_folder_path(title):
        """Mirror the HTML folder routing for batch export."""
        top_m = re.match(r"^(020|030|100)", title)
        if not top_m:
            return "기타"
        top = top_m.group(1)
        m = re.match(r"^(?:020|030|100)([CN])(FV0|FVM|FVP|NV0|NVC)", title)
        if not m:
            return top
        mid = "cong" if m.group(1) == "C" else "norm"
        group = m.group(2)
        return "%s/%s/%s" % (top, mid, group)

    def _save_all_zip(self):
        if not self.files:
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Save all FDS decks (ZIP)", "fds_batch.zip",
            "ZIP archive (*.zip)")
        if not path:
            return
        try:
            m = self.compute()
            written = 0
            skipped = 0
            with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as zf:
                for f in self.files:
                    if not f.get("profile"):
                        skipped += 1
                        continue
                    cfg = self._cfg_for_file(f)
                    fds = self.build_fds(m, cfg)
                    folder = self._fds_folder_path(f["title"])
                    arc = "%s/%s.fds" % (folder, safe_name(f["title"]))
                    zf.writestr(arc, fds)
                    written += 1
            msg = "✓ wrote %d deck(s)" % written
            if skipped:
                msg += " · %d skipped (no profile)" % skipped
            self.stat_lbl.setText(msg)
        except Exception as e:
            QMessageBox.warning(self, "ZIP export failed", str(e))

    # ----------------------------------------------------------------
    #  Output actions
    # ----------------------------------------------------------------
    def _copy(self):
        QApplication.clipboard().setText(self._last_fds)
        self.stat_lbl.setText("✓ copied")

    def _save(self):
        path, _ = QFileDialog.getSaveFileName(
            self, "Save FDS file", self._last_fname, "FDS files (*.fds);;All files (*)")
        if not path:
            return
        try:
            with open(path, "w", encoding="utf-8") as fh:
                fh.write(self._last_fds)
            self.stat_lbl.setText("✓ saved")
        except Exception as e:
            QMessageBox.warning(self, "Save failed", str(e))

    def _reset(self):
        self._building = True
        # clear batch-import + DXF state
        self.files = []
        self.active = -1
        self.dxf_poly = None
        self.dxf_w = self.dxf_h = 0.0
        self._last_geom_sig = None
        for f, val in zip(
            (self.in_W, self.in_nW, self.in_H, self.in_nH, self.in_Lh,
             self.in_dx, self.in_breath),
            (10.86, 20, 6.73, 15, 720, 1, 1.8)):
            f.setText(fmt(val))
        self.cmb_preset.setCurrentIndex(0)
        for f, val in zip(
            (self.in_hrr, self.in_bhl, self.in_bz, self.in_bo1, self.in_bo2),
            PRESETS["20"]):
            f.setText(fmt(val))
        self.in_twfin.setText("1200")
        self.in_v0.setText("6.02")
        self.cmb_bc.setCurrentIndex(2)
        self.cmb_sign.setCurrentIndex(0)
        self.sign = 1
        self._init_defaults()
        self._building = False
        self.btn_dxf_clear.hide()
        self.dxf_info.setText(
            "Import a closed cross-section polyline (LWPOLYLINE / POLYLINE). "
            "Sets W·H and fills wall cells by point-in-polygon.")
        self.dxf_info.setStyleSheet("color:#8a7ddb;font-size:11px;")
        self._render_queue()
        self.set_sym(True)
        self.update_all()
        self.stat_lbl.setText("↺ reset")


# ====================================================================
#  Stand-alone test harness
# ====================================================================
if __name__ == "__main__":
    import sys
    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    w = QWidget()
    lay = QVBoxLayout(w)
    tabw = __import__("PyQt5.QtWidgets", fromlist=["QTabWidget"]).QTabWidget()
    tabw.addTab(GeneralInputTab(), "General Input Information")
    lay.addWidget(tabw)
    w.resize(1500, 950)
    w.show()
    sys.exit(app.exec_())

# ====================================================================
#  INTEGRATION INTO qra_main_app.py
# --------------------------------------------------------------------
#  1. Put this file next to qra_main_app.py.
#  2. At the top of qra_main_app.py add:
#         from general_input_tab import GeneralInputTab
#  3. Inside QRAMainWindow.init_ui(), make this the FIRST tab by adding
#     ONE line BEFORE the existing create_tab1_directory_setup() call:
#
#         # --- General Input Information (FDS deck builder) ---
#         self.general_input_tab = GeneralInputTab()
#         self.tabs.addTab(self.general_input_tab, "General Input Information")
#
#         self.create_tab1_directory_setup()   # (existing line, now 2nd tab)
#         ...
#
#  NOTE: _on_main_tab_changed() keys "Tab 4 (EVC/FED)" off index == 3.
#  Because a new first tab shifts every later tab by one, change that
#  guard from `if index == 3` to `if index == 4` so the EVC/FED
#  directory auto-creation still fires on the right tab.
# ====================================================================