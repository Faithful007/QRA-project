"""
fn_curve_subtab_widget.py
=========================
FN-Curve sub-tab for Tab 7 (Risk Calculations).

Mirrors the FNCurve2 sheet of FNCV_ROAD.xlsm:

  Input (columns L–R, rows 3..N) is a list of accident scenarios sorted
  by fatalities (R) in DESCENDING order. Each row has:
      L = fire position (1..6)
      M = scenario code (e.g. "100C+CNVC")
      O = frequency per year   f_i
      R = number of fatalities N_i

  Cumulative frequency (column S) is the CCDF F(N >= n), built top-down:
      S_3 = O_3
      S_i = S_{i-1} + O_i           (i = 4, 5, ...)
  Because rows are sorted by N descending, S_i is the sum of frequencies
  of all scenarios with N >= R_i — i.e. the complementary cumulative
  distribution function evaluated at N = R_i.

  FN-Curve plot points (columns T,U) form a staircase. Walking the data
  bottom-up (smallest N to largest N), each pair of consecutive source
  rows emits TWO plot points:
      horizontal:  (N_curr, F_curr) -> (N_next, F_curr)
      vertical:    (N_next, F_curr) -> (N_next, F_next)
  giving the classic descending log-log staircase used in QRA
  societal-risk plots.

Public API
----------
    FNCurveSubTabWidget(workbook_path: Path | str | None = None,
                        parent: QWidget | None = None)
        Build the widget. If `workbook_path` is None, look next to this
        file for FNCV_ROAD.xlsm.

    .reload_from_workbook(path)        Re-read the workbook and redraw.
    .set_scenarios(rows)               Inject scenarios programmatically:
                                       rows is an iterable of dicts with keys
                                       'pos', 'name', 'freq', 'fatalities'
                                       (extra keys are preserved for the table).
    .get_fn_points()    -> list[(N, F)] sorted by N ascending  (CCDF curve points)
    .get_staircase()    -> (xs, ys) ready to plot on a log-log axis
"""
from __future__ import annotations

from pathlib import Path
from typing import Iterable, List, Optional, Sequence, Tuple

from PyQt5.QtCore import Qt
from PyQt5.QtGui import QColor
from PyQt5.QtWidgets import (
    QAbstractItemView,
    QCheckBox,
    QDoubleSpinBox,
    QFileDialog,
    QFrame,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QLineEdit,
    QMessageBox,
    QPushButton,
    QSplitter,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)


# ── Matplotlib (optional but required for the chart half) ────────────────────
try:
    import matplotlib
    matplotlib.use("Qt5Agg")
    from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
    from matplotlib.figure import Figure
    _HAVE_MPL = True
except Exception:                                          # pragma: no cover
    _HAVE_MPL = False


# ─────────────────────────────────────────────────────────────────────────────
# Core algorithm — mirrors FNCurve2 columns L..U exactly
# ─────────────────────────────────────────────────────────────────────────────
def build_fn_curve(scenarios: Sequence[dict]) -> Tuple[
    List[dict], List[Tuple[float, float]], Tuple[List[float], List[float]]
]:
    """Replicate the FNCurve2 sheet construction.

    Parameters
    ----------
    scenarios : sequence of dicts with keys 'fatalities' (N) and 'freq' (f).
                Other keys are passed through untouched.

    Returns
    -------
    sorted_rows : list of input dicts, sorted by N descending, each with an
                  added 'cumul_freq' key (the column-S value).
    fn_points   : list of (N, F(>=N)) tuples, one per UNIQUE N value,
                  sorted by N ascending.    These are the "knots" of the
                  CCDF and the data plotted in column T,U of the sheet.
    staircase   : (xs, ys) lists ready to feed matplotlib's plot() to draw
                  the staircase the same way the sheet does.
    """
    # Keep only scenarios that contribute to the curve: positive frequency
    # AND at least one fatality. The Excel sheet skips empty/zero rows the
    # same way (rows 15-22 of FNCurve2 only carry T/U because their O is 0).
    rows = [
        dict(s) for s in scenarios
        if (s.get("freq") or 0.0) > 0.0
           and (s.get("fatalities") or 0.0) > 0.0
    ]

    # 1. Sort by fatalities DESCENDING, like FN_CURVE_CREATE3 in the VBA.
    rows.sort(key=lambda r: -float(r["fatalities"]))

    # 2. Build the CCDF column-S: top-down running sum of frequencies.
    #    Because the rows are sorted by N descending, this running sum at
    #    row i is exactly F(N >= N_i) = sum of freq of all events whose N >= N_i.
    cumul = 0.0
    for r in rows:
        cumul += float(r["freq"])
        r["cumul_freq"] = cumul

    # 3. Collapse rows that share the same N into one knot, taking the
    #    LARGEST cumulative frequency at that N (i.e. the value AFTER all
    #    rows with that N have been added). This is what the staircase
    #    drops to when crossing that N from the left.
    knots: dict[float, float] = {}
    for r in rows:
        n = float(r["fatalities"])
        f = r["cumul_freq"]
        # Because rows are processed in N-descending order, the first time
        # we see a given N has the SMALLEST cumul; subsequent occurrences of
        # the same N have larger cumul. We want the LARGEST = the value
        # after the last duplicate row was summed in.
        if (n not in knots) or (f > knots[n]):
            knots[n] = f
    fn_points = sorted(knots.items(), key=lambda p: p[0])   # ascending N

    # 4. Staircase: walk fn_points left-to-right (smallest N first); at each
    #    knot the curve is HORIZONTAL out to the next knot's N, then VERTICAL
    #    DOWN to the next knot's F. Same convention as FNCurve2 columns T,U.
    xs: List[float] = []
    ys: List[float] = []
    for i, (n_curr, f_curr) in enumerate(fn_points):
        if i == 0:
            xs.append(n_curr)
            ys.append(f_curr)
        if i < len(fn_points) - 1:
            n_next, f_next = fn_points[i + 1]
            # Horizontal segment to next N at the current F level
            xs.append(n_next)
            ys.append(f_curr)
            # Vertical drop down to the next F level at that same N
            xs.append(n_next)
            ys.append(f_next)
    return rows, fn_points, (xs, ys)


# ─────────────────────────────────────────────────────────────────────────────
# Workbook loader — reads the FNCurve2 sheet using the SAME columns as VBA
# ─────────────────────────────────────────────────────────────────────────────
def _load_scenarios_from_workbook(workbook_path: Path) -> List[dict]:
    """Read scenarios from the FNCurve2 sheet of FNCV_ROAD.xlsm.

    Expected sheet layout (matches the workbook in the project root):

        Row 1: headers
        Row 2: header row 2
        Row 3..N (left block, columns A..K): per-vehicle/category aggregates.
                                             We DON'T plot these.
        Row 25..(end): the per-scenario list that drives the plot.
            A = Fire Point
            B = Scenario code
            C = Description
            E = Frequency / yr
            F = Frequency / veh-km
            H = Fatalities  (사망자수)
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:                              # pragma: no cover
        raise RuntimeError(
            "openpyxl is required to load FNCV_ROAD.xlsm — pip install openpyxl"
        ) from exc

    wb = load_workbook(str(workbook_path), data_only=True, keep_vba=True)
    if "FNCurve2" not in wb.sheetnames:
        raise RuntimeError(
            f"Workbook {workbook_path.name!r} has no FNCurve2 sheet "
            f"(found: {wb.sheetnames})."
        )
    ws = wb["FNCurve2"]

    scenarios: List[dict] = []
    # Scenario rows start at row 25 in the project's workbook. We probe
    # for the first row whose A column is a number to be defensive against
    # template revisions.
    start_row = None
    for row in range(20, ws.max_row + 1):
        a = ws.cell(row=row, column=1).value
        if isinstance(a, (int, float)):
            start_row = row
            break
    if start_row is None:
        return scenarios

    for row in range(start_row, ws.max_row + 1):
        pos  = ws.cell(row=row, column=1).value          # A
        name = ws.cell(row=row, column=2).value          # B
        desc = ws.cell(row=row, column=3).value          # C
        freq = ws.cell(row=row, column=5).value          # E  freq /yr
        fat  = ws.cell(row=row, column=8).value          # H  사망자수
        if pos is None and name is None and freq is None and fat is None:
            continue
        try:
            f = float(freq) if freq is not None else 0.0
        except (TypeError, ValueError):
            f = 0.0
        try:
            n = float(fat) if fat is not None else 0.0
        except (TypeError, ValueError):
            n = 0.0
        scenarios.append({
            "pos":        pos,
            "name":       str(name) if name is not None else "",
            "desc":       str(desc) if desc is not None else "",
            "freq":       f,
            "fatalities": n,
        })
    return scenarios


# ─────────────────────────────────────────────────────────────────────────────
# Widget
# ─────────────────────────────────────────────────────────────────────────────
class FNCurveSubTabWidget(QWidget):
    """FN-Curve sub-tab for the Risk Calculations tab.

    Reads scenarios from FNCV_ROAD.xlsm (FNCurve2 sheet) on construction,
    builds the CCDF F(N>=n), shows the per-knot data table on the left
    and the log-log staircase chart on the right.
    """

    DEFAULT_WORKBOOK_NAME = "FNCV_ROAD.xlsm"

    def __init__(
        self,
        workbook_path: Optional[Path | str] = None,
        parent: Optional[QWidget] = None,
    ) -> None:
        super().__init__(parent)

        self._workbook_path: Optional[Path] = None
        self._scenarios: List[dict] = []
        self._sorted_rows: List[dict] = []
        self._fn_points: List[Tuple[float, float]] = []
        self._stair_xs: List[float] = []
        self._stair_ys: List[float] = []

        self._build_ui()

        # Load workbook
        if workbook_path is None:
            workbook_path = Path(__file__).resolve().parent / self.DEFAULT_WORKBOOK_NAME
        self.reload_from_workbook(workbook_path, silent=True)

    # ── UI construction ───────────────────────────────────────────────────
    def _build_ui(self) -> None:
        outer = QVBoxLayout(self)
        outer.setContentsMargins(6, 6, 6, 6)
        outer.setSpacing(6)

        # Header / status bar
        header_row = QHBoxLayout()
        title = QLabel("FN-Curve  —  Societal Risk  F(N ≥ n)")
        title.setStyleSheet("font-weight:bold; font-size:13px; color:#2c3e50;")
        header_row.addWidget(title)
        header_row.addStretch()

        self._reload_btn = QPushButton("📂  Load workbook…")
        self._reload_btn.setStyleSheet(
            "QPushButton { background:#2980b9; color:white; padding:4px 10px; border-radius:3px; }"
            "QPushButton:hover { background:#3498db; }"
        )
        self._reload_btn.clicked.connect(self._on_browse_workbook)
        header_row.addWidget(self._reload_btn)

        self._save_png_btn = QPushButton("💾  Save PNG")
        self._save_png_btn.setStyleSheet(
            "QPushButton { background:#27ae60; color:white; padding:4px 10px; border-radius:3px; }"
            "QPushButton:hover { background:#2ecc71; }"
        )
        self._save_png_btn.clicked.connect(self._on_save_png)
        header_row.addWidget(self._save_png_btn)

        outer.addLayout(header_row)

        self._status_lbl = QLabel("Loading workbook…")
        self._status_lbl.setStyleSheet("color:#666; font-size:11px;")
        outer.addWidget(self._status_lbl)

        # Splitter: data table (1/3) on the left, FN curve chart (2/3) on the right
        splitter = QSplitter(Qt.Horizontal)
        splitter.setChildrenCollapsible(False)
        outer.addWidget(splitter, 1)

        # ── Left: data table (mirrors columns L,M,R,O,S of the sheet) ─────
        left_panel = QWidget()
        left_v = QVBoxLayout(left_panel)
        left_v.setContentsMargins(0, 0, 0, 0)
        left_v.setSpacing(4)

        left_v.addWidget(self._info_label(
            "Sorted scenarios (by fatalities descending). "
            "Cumul. Freq. = F(N ≥ n) — sum of freq of all events with N ≥ this row's N."
        ))

        self._table = QTableWidget(0, 5)
        self._table.setHorizontalHeaderLabels([
            "Pos", "Scenario", "Fatalities (N)", "Freq /yr (f)", "Cumul. Freq. F(≥N)"
        ])
        self._table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self._table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self._table.setAlternatingRowColors(True)
        self._table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self._table.horizontalHeader().setStretchLastSection(True)
        self._table.verticalHeader().setDefaultSectionSize(20)
        left_v.addWidget(self._table, 1)

        splitter.addWidget(left_panel)

        # ── Right: chart (2/3) — controls are in the full-width panel above ──
        right_panel = QWidget()
        right_v = QVBoxLayout(right_panel)
        right_v.setContentsMargins(0, 0, 0, 0)
        right_v.setSpacing(4)

        # ── Toolbar (3 rows of controls) sits ABOVE the chart so it can never
        # be clipped when the parent window is short. ────────────────────────
        toolbar = QFrame()
        toolbar.setFrameShape(QFrame.NoFrame)
        toolbar.setStyleSheet(
            "QFrame { background:#eef1f4; border-bottom:1px solid #cdd4dc; }"
        )
        toolbar_v = QVBoxLayout(toolbar)
        toolbar_v.setContentsMargins(6, 4, 6, 4)
        toolbar_v.setSpacing(3)
        # Toolbar is inserted into outer ABOVE the splitter so it spans the
        # full width (both the data-table third and the chart two-thirds).
        # outer already holds [header_row, status_lbl, splitter] at this point,
        # so insertWidget(2, …) places the toolbar just before the splitter.
        outer.insertWidget(2, toolbar)  # full-width control panel above splitter

        # Labels row (row 0): editable chart title + axis labels
        labels_ctrl = QHBoxLayout()
        labels_ctrl.setSpacing(6)
        labels_ctrl.addWidget(QLabel("Title:"))
        self._title_edit = QLineEdit("F-N Curve  (Societal Risk)")
        self._title_edit.setMinimumWidth(180)
        self._title_edit.setToolTip("Chart title (press Enter or tab away to apply).")
        self._title_edit.editingFinished.connect(self._redraw)
        labels_ctrl.addWidget(self._title_edit, 1)

        labels_ctrl.addWidget(QLabel("X label:"))
        self._xlabel_edit = QLineEdit("Number of Fatalities  N")
        self._xlabel_edit.setMinimumWidth(150)
        self._xlabel_edit.setToolTip("Label for the horizontal axis.")
        self._xlabel_edit.editingFinished.connect(self._redraw)
        labels_ctrl.addWidget(self._xlabel_edit, 1)

        labels_ctrl.addWidget(QLabel("Y label:"))
        self._ylabel_edit = QLineEdit("Cumulative Frequency  F(N ≥ n)  [events / year]")
        self._ylabel_edit.setMinimumWidth(180)
        self._ylabel_edit.setToolTip("Label for the vertical axis.")
        self._ylabel_edit.editingFinished.connect(self._redraw)
        labels_ctrl.addWidget(self._ylabel_edit, 1)

        toolbar_v.addLayout(labels_ctrl)

        # Legend placement state. Set on first draw, then updated by the
        # drag-end callback so the legend stays where the user dropped it
        # across subsequent redraws within the same session.
        self._legend_loc: object = "upper right"   # default per user request

        # ALARP controls (row 1)
        ctrl = QHBoxLayout()
        ctrl.setSpacing(6)
        ctrl.addWidget(QLabel("ALARP upper:"))
        self._alarp_upper = QDoubleSpinBox()
        self._alarp_upper.setRange(1e-12, 1.0)
        self._alarp_upper.setDecimals(10)
        self._alarp_upper.setValue(1e-4)
        self._alarp_upper.setFixedWidth(110)
        ctrl.addWidget(self._alarp_upper)

        ctrl.addWidget(QLabel("lower:"))
        self._alarp_lower = QDoubleSpinBox()
        self._alarp_lower.setRange(1e-12, 1.0)
        self._alarp_lower.setDecimals(10)
        self._alarp_lower.setValue(1e-6)
        self._alarp_lower.setFixedWidth(110)
        ctrl.addWidget(self._alarp_lower)

        self._show_alarp_cb = QCheckBox("Show ALARP")
        self._show_alarp_cb.setChecked(True)
        self._show_alarp_cb.toggled.connect(self._redraw)
        ctrl.addWidget(self._show_alarp_cb)

        self._show_grid_cb = QCheckBox("Grid")
        self._show_grid_cb.setChecked(True)
        self._show_grid_cb.toggled.connect(self._redraw)
        ctrl.addWidget(self._show_grid_cb)

        self._alarp_upper.valueChanged.connect(self._redraw)
        self._alarp_lower.valueChanged.connect(self._redraw)

        redraw_btn = QPushButton("🔁  Redraw")
        redraw_btn.setStyleSheet(
            "QPushButton { background:#34495e; color:white; padding:3px 9px; border-radius:3px; }"
        )
        redraw_btn.clicked.connect(self._redraw)
        ctrl.addWidget(redraw_btn)
        ctrl.addStretch()
        toolbar_v.addLayout(ctrl)

        # ── Second control row: HSE criterion lines (slope -1 per decade) ───
        # Standard UK HSE societal-risk criteria are F = C / N, anchored at N=1:
        #   HSE High Level (intolerable threshold):     C = 1e-3
        #   HSE Low Level  (broadly acceptable line):   C = 1e-5
        # The user can override the anchors below.
        hse_ctrl = QHBoxLayout()
        hse_ctrl.setSpacing(6)
        hse_ctrl.addWidget(QLabel("HSE High @ N=1:"))
        self._hse_high = QDoubleSpinBox()
        self._hse_high.setRange(1e-12, 1.0)
        self._hse_high.setDecimals(10)
        self._hse_high.setValue(1e-3)
        self._hse_high.setFixedWidth(110)
        self._hse_high.setToolTip(
            "Anchor of the HSE High-Level (intolerable) criterion at N=1.\n"
            "The line is F = anchor / N (slope -1 per decade on log-log)."
        )
        hse_ctrl.addWidget(self._hse_high)

        hse_ctrl.addWidget(QLabel("HSE Low @ N=1:"))
        self._hse_low = QDoubleSpinBox()
        self._hse_low.setRange(1e-12, 1.0)
        self._hse_low.setDecimals(10)
        self._hse_low.setValue(1e-5)
        self._hse_low.setFixedWidth(110)
        self._hse_low.setToolTip(
            "Anchor of the HSE Low-Level (broadly acceptable) criterion at N=1.\n"
            "The line is F = anchor / N (slope -1 per decade on log-log)."
        )
        hse_ctrl.addWidget(self._hse_low)

        self._show_hse_cb = QCheckBox("Show HSE")
        self._show_hse_cb.setChecked(True)
        self._show_hse_cb.setToolTip(
            "Overlay the HSE High & Low criterion lines (UK risk-neutral, slope -1)."
        )
        self._show_hse_cb.toggled.connect(self._redraw)
        hse_ctrl.addWidget(self._show_hse_cb)

        self._hse_high.valueChanged.connect(self._redraw)
        self._hse_low.valueChanged.connect(self._redraw)

        hse_ctrl.addStretch()
        toolbar_v.addLayout(hse_ctrl)

        # ── Third control row: axis ranges  ─────────────────────────────────
        # On log axes, "interval" means a tick multiplier — e.g. 10 = one major
        # tick per decade (the standard), 100 = every two decades. Internally
        # we use ceil(log10(interval)) as the LogLocator step.
        axis_ctrl = QHBoxLayout()
        axis_ctrl.setSpacing(6)

        axis_ctrl.addWidget(QLabel("X (N):"))
        axis_ctrl.addWidget(QLabel("start"))
        self._x_start = QDoubleSpinBox()
        self._x_start.setRange(1e-6, 1e9)
        self._x_start.setDecimals(6)
        self._x_start.setValue(1.0)
        self._x_start.setFixedWidth(95)
        axis_ctrl.addWidget(self._x_start)

        axis_ctrl.addWidget(QLabel("stop"))
        self._x_stop = QDoubleSpinBox()
        self._x_stop.setRange(1e-6, 1e9)
        self._x_stop.setDecimals(6)
        self._x_stop.setValue(1.0e4)
        self._x_stop.setFixedWidth(95)
        axis_ctrl.addWidget(self._x_stop)

        axis_ctrl.addWidget(QLabel("step"))
        self._x_step = QDoubleSpinBox()
        self._x_step.setRange(1.001, 1e6)
        self._x_step.setDecimals(3)
        self._x_step.setValue(10.0)
        self._x_step.setFixedWidth(70)
        self._x_step.setToolTip(
            "Major-tick multiplier on the log X axis.\n"
            "10 = every decade (standard).  100 = every two decades."
        )
        axis_ctrl.addWidget(self._x_step)

        axis_ctrl.addSpacing(14)

        axis_ctrl.addWidget(QLabel("Y (F):"))
        axis_ctrl.addWidget(QLabel("start"))
        self._y_start = QDoubleSpinBox()
        self._y_start.setRange(1e-15, 1.0)
        self._y_start.setDecimals(12)
        self._y_start.setValue(1.0e-9)
        self._y_start.setFixedWidth(110)
        axis_ctrl.addWidget(self._y_start)

        axis_ctrl.addWidget(QLabel("stop"))
        self._y_stop = QDoubleSpinBox()
        self._y_stop.setRange(1e-15, 1.0)
        self._y_stop.setDecimals(12)
        self._y_stop.setValue(1.0)
        self._y_stop.setFixedWidth(110)
        axis_ctrl.addWidget(self._y_stop)

        axis_ctrl.addWidget(QLabel("step"))
        self._y_step = QDoubleSpinBox()
        self._y_step.setRange(1.001, 1e6)
        self._y_step.setDecimals(3)
        self._y_step.setValue(10.0)
        self._y_step.setFixedWidth(70)
        self._y_step.setToolTip(
            "Major-tick multiplier on the log Y axis.\n"
            "10 = every decade (standard).  100 = every two decades."
        )
        axis_ctrl.addWidget(self._y_step)

        self._auto_axes_cb = QCheckBox("Auto")
        self._auto_axes_cb.setChecked(False)
        self._auto_axes_cb.setToolTip(
            "When checked, the chart auto-fits to the data and ignores the\n"
            "start/stop/step boxes above. Uncheck to use the manual ranges."
        )
        self._auto_axes_cb.toggled.connect(self._redraw)
        axis_ctrl.addWidget(self._auto_axes_cb)

        fit_btn = QPushButton("Fit to data")
        fit_btn.setStyleSheet(
            "QPushButton { background:#7f8c8d; color:white; padding:3px 9px; border-radius:3px; }"
            "QPushButton:hover { background:#95a5a6; }"
        )
        fit_btn.setToolTip(
            "Snap the start/stop boxes to the data extent (one decade of padding)."
        )
        fit_btn.clicked.connect(self._fit_axes_to_data)
        axis_ctrl.addWidget(fit_btn)

        for sb in (self._x_start, self._x_stop, self._x_step,
                   self._y_start, self._y_stop, self._y_step):
            sb.valueChanged.connect(self._redraw)

        axis_ctrl.addStretch()
        toolbar_v.addLayout(axis_ctrl)

        # ── Chart canvas comes AFTER the toolbar so the toolbar is always
        # visible, and the chart absorbs whatever vertical space remains. ──
        if _HAVE_MPL:
            self._fig = Figure(figsize=(7, 5))
            self._fig.patch.set_facecolor("#f8f9fa")
            self._canvas = FigureCanvas(self._fig)
            self._ax = self._fig.add_subplot(111)
            self._canvas.setMinimumHeight(280)
            right_v.addWidget(self._canvas, 1)   # stretch=1 → fills remainder
        else:                                                # pragma: no cover
            self._canvas = None
            self._ax = None
            no_mpl = QLabel(
                "matplotlib is not available — install it to see the FN curve "
                "(pip install matplotlib)."
            )
            no_mpl.setAlignment(Qt.AlignCenter)
            no_mpl.setStyleSheet("color:#7f8c8d; font-size:13px;")
            right_v.addWidget(no_mpl, 1)

        splitter.addWidget(right_panel)
        splitter.setStretchFactor(0, 1)  # left  = 1/3 (data table)
        splitter.setStretchFactor(1, 2)  # right = 2/3 (FN curve chart)
        splitter.setSizes([367, 733])

    @staticmethod
    def _info_label(text: str) -> QLabel:
        lbl = QLabel(text)
        lbl.setWordWrap(True)
        lbl.setStyleSheet("color:#555; font-size:11px; padding:2px;")
        return lbl

    # ── Public API ────────────────────────────────────────────────────────
    def reload_from_workbook(
        self,
        workbook_path: Path | str,
        silent: bool = False,
    ) -> bool:
        """Reload scenarios from the given workbook and redraw.

        Returns True on success, False otherwise.
        """
        path = Path(workbook_path)
        if not path.exists():
            msg = (
                f"Workbook not found:\n{path}\n\n"
                f"Click 'Load workbook…' to pick the correct FNCV_ROAD.xlsm."
            )
            self._status_lbl.setText(f"⚠  {path.name} not found.")
            self._status_lbl.setStyleSheet("color:#c0392b; font-size:11px;")
            if not silent:
                QMessageBox.warning(self, "FN-Curve", msg)
            return False
        try:
            scenarios = _load_scenarios_from_workbook(path)
        except Exception as exc:
            self._status_lbl.setText(f"⚠  Error loading {path.name}: {exc}")
            self._status_lbl.setStyleSheet("color:#c0392b; font-size:11px;")
            if not silent:
                QMessageBox.warning(self, "FN-Curve", str(exc))
            return False

        self._workbook_path = path
        self.set_scenarios(scenarios)
        return True

    def set_scenarios(self, scenarios: Iterable[dict]) -> None:
        """Inject scenarios programmatically and rebuild the curve.

        Each scenario must be a dict with at minimum:
            'fatalities' : float (N)
            'freq'       : float (events / year)
        Other keys ('pos', 'name', 'desc', …) are kept for the table.
        """
        self._scenarios = [dict(s) for s in scenarios]
        self._recompute()
        self._populate_table()
        self._redraw()

        n_total   = len(self._scenarios)
        n_used    = len(self._sorted_rows)
        n_knots   = len(self._fn_points)
        if n_used == 0:
            self._status_lbl.setText(
                f"⚠  Loaded {n_total} scenario(s) but none have non-zero "
                f"frequency AND fatalities — FN-curve is empty. "
                f"Run the EVC analysis to populate fatalities (column H of FNCurve2)."
            )
            self._status_lbl.setStyleSheet("color:#c0392b; font-size:11px;")
        else:
            self._status_lbl.setText(
                f"✅  {n_used} contributing scenario(s) → {n_knots} unique N "
                f"value(s).  Total freq F(N≥0) = {self._sorted_rows[-1]['cumul_freq']:.4e} /yr."
            )
            self._status_lbl.setStyleSheet("color:#27ae60; font-size:11px;")

    def get_fn_points(self) -> List[Tuple[float, float]]:
        """Return [(N, F(>=N)), ...] sorted by N ascending."""
        return list(self._fn_points)

    def get_staircase(self) -> Tuple[List[float], List[float]]:
        """Return (xs, ys) ready to plot the staircase on a log-log axis."""
        return list(self._stair_xs), list(self._stair_ys)

    # ── Internal: recompute ───────────────────────────────────────────────
    def _recompute(self) -> None:
        sorted_rows, fn_points, (xs, ys) = build_fn_curve(self._scenarios)
        self._sorted_rows = sorted_rows
        self._fn_points = fn_points
        self._stair_xs = xs
        self._stair_ys = ys

    # ── Internal: populate table ──────────────────────────────────────────
    def _populate_table(self) -> None:
        self._table.setRowCount(0)

        def _ci(text, bg=None, align=Qt.AlignCenter):
            it = QTableWidgetItem(str(text))
            it.setTextAlignment(align)
            if bg is not None:
                it.setBackground(bg)
            return it

        zebra = QColor("#fbfcfd")
        for r in self._sorted_rows:
            ri = self._table.rowCount()
            self._table.insertRow(ri)
            self._table.setItem(ri, 0, _ci(r.get("pos", "")))
            self._table.setItem(ri, 1, _ci(r.get("name", "")))
            self._table.setItem(ri, 2, _ci(f"{float(r['fatalities']):.2f}"))
            self._table.setItem(ri, 3, _ci(f"{float(r['freq']):.4e}"))
            self._table.setItem(
                ri, 4,
                _ci(f"{float(r['cumul_freq']):.4e}", bg=zebra),
            )

    # ── Internal: snap manual axis boxes to the data extent ──────────────
    def _fit_axes_to_data(self) -> None:
        """Set the start/stop spinboxes to one decade of padding around the data."""
        if not self._stair_xs or not self._stair_ys:
            return
        xs = self._stair_xs
        ys = [y for y in self._stair_ys if y > 0.0]
        if not ys:
            return
        # Pad by a decade on each side, then snap to nice powers of 10.
        import math
        x_lo = 10 ** math.floor(math.log10(min(xs)) - 0.0)
        x_hi = 10 ** math.ceil(math.log10(max(xs)) + 0.0)
        y_lo = 10 ** math.floor(math.log10(min(ys)) - 0.0)
        y_hi = 10 ** math.ceil(math.log10(max(ys)) + 0.0)
        # Block signals so we redraw only once at the end.
        for sb in (self._x_start, self._x_stop, self._y_start, self._y_stop):
            sb.blockSignals(True)
        self._x_start.setValue(max(x_lo, self._x_start.minimum()))
        self._x_stop.setValue(min(x_hi, self._x_stop.maximum()))
        self._y_start.setValue(max(y_lo, self._y_start.minimum()))
        self._y_stop.setValue(min(y_hi, self._y_stop.maximum()))
        for sb in (self._x_start, self._x_stop, self._y_start, self._y_stop):
            sb.blockSignals(False)
        self._auto_axes_cb.setChecked(False)
        self._redraw()

    # ── Internal: redraw chart ────────────────────────────────────────────
    def _redraw(self) -> None:
        if not _HAVE_MPL or self._ax is None:
            return
        ax = self._ax
        ax.clear()

        if not self._stair_xs:
            ax.text(
                0.5, 0.5, "No FN data — load FNCV_ROAD.xlsm",
                transform=ax.transAxes, ha="center", va="center",
                fontsize=12, color="#7f8c8d",
                bbox=dict(boxstyle="round,pad=0.5", fc="#ecf0f1", alpha=0.85),
            )
            ax.set_xscale("linear")
            ax.set_yscale("linear")
            self._fig.tight_layout()
            self._canvas.draw_idle()
            return

        xs = list(self._stair_xs)
        ys = list(self._stair_ys)

        # ── Resolve the axis range FIRST. Everything else (bands, diagonals,
        # limits, tick locators) snaps to this single source of truth.
        import math
        if self._auto_axes_cb.isChecked():
            x_left = min(xs) * 0.7
            x_right = max(xs) * 1.4
            ys_pos = [y for y in ys if y > 0.0]
            y_bot = (min(ys_pos) * 0.1) if ys_pos else 1e-12
            y_top = (max(ys) * 10.0) if ys else 1.0
            if self._show_hse_cb.isChecked():
                x_right = max(x_right, max(xs) * 10.0, 1000.0)
                x_left  = min(x_left, 1.0)
                y_top   = max(y_top, 1.0)
        else:
            x_left  = float(self._x_start.value())
            x_right = float(self._x_stop.value())
            y_bot   = float(self._y_start.value())
            y_top   = float(self._y_stop.value())
            # Guard against inverted ranges
            if x_right <= x_left:
                x_right = x_left * 10.0
            if y_top <= y_bot:
                y_top = y_bot * 10.0

        # ALARP bands span the full plotted x-range so the colored regions
        # line up with the diagonals visually.
        if self._show_alarp_cb.isChecked():
            x_band = [x_left, x_right]
            f_up = float(self._alarp_upper.value())
            f_lo = float(self._alarp_lower.value())
            if f_up > f_lo:
                ax.fill_between(
                    x_band, [f_up, f_up], [max(y_top, 1.0)] * 2,
                    color="#e74c3c", alpha=0.12,
                    label=f"Intolerable (>{f_up:.0E}/yr)",
                )
                ax.fill_between(
                    x_band, [f_lo, f_lo], [f_up, f_up],
                    color="#f39c12", alpha=0.10,
                    label="ALARP region",
                )
                ax.fill_between(
                    x_band, [min(y_bot, 1e-14)] * 2, [f_lo, f_lo],
                    color="#27ae60", alpha=0.10,
                    label=f"Broadly acceptable (<{f_lo:.0E}/yr)",
                )

        # HSE criterion lines: F = anchor / N  (slope -1 per decade on log-log).
        # Draw them BEFORE the staircase so the data sits on top.
        if self._show_hse_cb.isChecked():
            anchor_high = float(self._hse_high.value())
            anchor_low  = float(self._hse_low.value())
            x_line = [x_left, x_right]
            for anchor, color, label in (
                (anchor_high, "#c0392b",
                 f"HSE High Level (F = {anchor_high:.0E}/N)"),
                (anchor_low,  "#16a085",
                 f"HSE Low Level (F = {anchor_low:.0E}/N)"),
            ):
                # F = anchor / N evaluated at the two range endpoints
                y_line = [anchor / x_left, anchor / x_right]
                ax.plot(
                    x_line, y_line,
                    linestyle="--", color=color, linewidth=1.6,
                    alpha=0.85, label=label,
                )

        # The staircase itself
        ax.plot(
            xs, ys,
            "-", color="#2c3e50", linewidth=2.0,
            label="F-N curve",
        )
        # Markers at the knots only (one per unique N)
        if self._fn_points:
            knot_xs = [p[0] for p in self._fn_points]
            knot_ys = [p[1] for p in self._fn_points]
            ax.plot(
                knot_xs, knot_ys,
                "o", color="#c0392b", markersize=5,
                label="Scenario knots",
            )

        ax.set_xscale("log")
        ax.set_yscale("log")
        ax.set_xlabel(self._xlabel_edit.text() or " ", fontsize=11)
        ax.set_ylabel(self._ylabel_edit.text() or " ", fontsize=11)
        ax.set_title(self._title_edit.text() or " ", fontsize=13, fontweight="bold")
        if self._show_grid_cb.isChecked():
            ax.grid(True, which="both", linewidth=0.5, alpha=0.5)
        else:
            ax.grid(False)

        # Legend: starts in the saved location (upper right by default), but
        # the user can drag it anywhere; the drag-end callback writes the new
        # position back to self._legend_loc so subsequent redraws preserve it.
        if isinstance(self._legend_loc, tuple):
            # User-dragged position: figure-fraction coords in (0,1)x(0,1)
            legend = ax.legend(
                fontsize=9, framealpha=0.85,
                loc="upper left", bbox_to_anchor=self._legend_loc,
                bbox_transform=self._fig.transFigure,
            )
        else:
            legend = ax.legend(fontsize=9, framealpha=0.85, loc=self._legend_loc)
        try:
            legend.set_draggable(True, use_blit=False)
            # When the user releases the drag, capture the legend's new
            # position in figure-fraction coords so the next redraw can
            # restore it.
            def _on_drag_release(_evt, _legend=legend):
                try:
                    bbox = _legend.get_window_extent()
                    inv  = self._fig.transFigure.inverted()
                    x0, y0 = inv.transform((bbox.x0, bbox.y1))   # use top-left
                    self._legend_loc = (float(x0), float(y0))
                except Exception:
                    pass
            self._canvas.mpl_connect("button_release_event", _on_drag_release)
        except Exception:
            # Older matplotlib without draggable() support — no-op.
            pass

        # Apply manual tick-step on log axes. The "step" is a multiplier:
        # 10 → every decade, 100 → every two decades, etc. We translate it
        # to LogLocator's `numticks` indirectly by skipping decades.
        from matplotlib.ticker import LogLocator
        if not self._auto_axes_cb.isChecked():
            x_step_decades = max(1, int(round(math.log10(float(self._x_step.value())))))
            y_step_decades = max(1, int(round(math.log10(float(self._y_step.value())))))
            ax.xaxis.set_major_locator(
                LogLocator(base=10.0, subs=(1.0,), numticks=64)
            )
            ax.yaxis.set_major_locator(
                LogLocator(base=10.0, subs=(1.0,), numticks=64)
            )
            # If step > 1 decade, thin the major ticks accordingly.
            if x_step_decades > 1:
                ax.xaxis.set_major_locator(
                    LogLocator(base=10.0, subs=(1.0,), numticks=64,
                               numdecs=x_step_decades)
                )
            if y_step_decades > 1:
                ax.yaxis.set_major_locator(
                    LogLocator(base=10.0, subs=(1.0,), numticks=64,
                               numdecs=y_step_decades)
                )

        ax.set_xlim(x_left, x_right)
        ax.set_ylim(y_bot, y_top)
        self._fig.tight_layout()
        self._canvas.draw_idle()

    # ── Slot: browse for workbook ────────────────────────────────────────
    def _on_browse_workbook(self) -> None:
        start_dir = (
            str(self._workbook_path.parent) if self._workbook_path else str(Path.cwd())
        )
        path, _ = QFileDialog.getOpenFileName(
            self,
            "Select FN-Curve workbook",
            start_dir,
            "Excel files (*.xlsm *.xlsx);;All files (*.*)",
        )
        if path:
            self.reload_from_workbook(path, silent=False)

    # ── Slot: save PNG ───────────────────────────────────────────────────
    def _on_save_png(self) -> None:
        if not _HAVE_MPL or self._canvas is None:
            QMessageBox.information(
                self, "FN-Curve", "matplotlib is not available — cannot save PNG."
            )
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Save FN curve as PNG", "fn_curve.png", "PNG files (*.png)"
        )
        if path:
            try:
                self._fig.savefig(path, dpi=200, bbox_inches="tight")
            except Exception as exc:
                QMessageBox.warning(self, "FN-Curve", f"Could not save: {exc}")


# ─────────────────────────────────────────────────────────────────────────────
# Self-test
# ─────────────────────────────────────────────────────────────────────────────
# if __name__ == "__main__":                                  # pragma: no cover
#     import sys
#     from PyQt5.QtWidgets import QApplication

#     app = QApplication(sys.argv)
#     here = Path(__file__).resolve().parent
#     wb = here / "FNCV_ROAD.xlsm"
#     w = FNCurveSubTabWidget(workbook_path=wb if wb.exists() else None)
#     w.resize(1100, 600)
#     w.setWindowTitle("FN-Curve — standalone test")
#     w.show()
#     sys.exit(app.exec_())

# """
# fn_curve_subtab_widget.py
# =========================
# FN-Curve sub-tab for Tab 7 (Risk Calculations).

# Mirrors the FNCurve2 sheet of FNCV_ROAD.xlsm:

#   Input (columns L–R, rows 3..N) is a list of accident scenarios sorted
#   by fatalities (R) in DESCENDING order. Each row has:
#       L = fire position (1..6)
#       M = scenario code (e.g. "100C+CNVC")
#       O = frequency per year   f_i
#       R = number of fatalities N_i

#   Cumulative frequency (column S) is the CCDF F(N >= n), built top-down:
#       S_3 = O_3
#       S_i = S_{i-1} + O_i           (i = 4, 5, ...)
#   Because rows are sorted by N descending, S_i is the sum of frequencies
#   of all scenarios with N >= R_i — i.e. the complementary cumulative
#   distribution function evaluated at N = R_i.

#   FN-Curve plot points (columns T,U) form a staircase. Walking the data
#   bottom-up (smallest N to largest N), each pair of consecutive source
#   rows emits TWO plot points:
#       horizontal:  (N_curr, F_curr) -> (N_next, F_curr)
#       vertical:    (N_next, F_curr) -> (N_next, F_next)
#   giving the classic descending log-log staircase used in QRA
#   societal-risk plots.

# Public API
# ----------
#     FNCurveSubTabWidget(workbook_path: Path | str | None = None,
#                         parent: QWidget | None = None)
#         Build the widget. If `workbook_path` is None, look next to this
#         file for FNCV_ROAD.xlsm.

#     .reload_from_workbook(path)        Re-read the workbook and redraw.
#     .set_scenarios(rows)               Inject scenarios programmatically:
#                                        rows is an iterable of dicts with keys
#                                        'pos', 'name', 'freq', 'fatalities'
#                                        (extra keys are preserved for the table).
#     .get_fn_points()    -> list[(N, F)] sorted by N ascending  (CCDF curve points)
#     .get_staircase()    -> (xs, ys) ready to plot on a log-log axis
# """
# from __future__ import annotations

# from pathlib import Path
# from typing import Iterable, List, Optional, Sequence, Tuple

# from PyQt5.QtCore import Qt
# from PyQt5.QtGui import QColor
# from PyQt5.QtWidgets import (
#     QAbstractItemView,
#     QCheckBox,
#     QDoubleSpinBox,
#     QFileDialog,
#     QFrame,
#     QHBoxLayout,
#     QHeaderView,
#     QLabel,
#     QMessageBox,
#     QPushButton,
#     QSplitter,
#     QTableWidget,
#     QTableWidgetItem,
#     QVBoxLayout,
#     QWidget,
# )


# # ── Matplotlib (optional but required for the chart half) ────────────────────
# try:
#     import matplotlib
#     matplotlib.use("Qt5Agg")
#     from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
#     from matplotlib.figure import Figure
#     _HAVE_MPL = True
# except Exception:                                          # pragma: no cover
#     _HAVE_MPL = False


# # ─────────────────────────────────────────────────────────────────────────────
# # Core algorithm — mirrors FNCurve2 columns L..U exactly
# # ─────────────────────────────────────────────────────────────────────────────
# def build_fn_curve(scenarios: Sequence[dict]) -> Tuple[
#     List[dict], List[Tuple[float, float]], Tuple[List[float], List[float]]
# ]:
#     """Replicate the FNCurve2 sheet construction.

#     Parameters
#     ----------
#     scenarios : sequence of dicts with keys 'fatalities' (N) and 'freq' (f).
#                 Other keys are passed through untouched.

#     Returns
#     -------
#     sorted_rows : list of input dicts, sorted by N descending, each with an
#                   added 'cumul_freq' key (the column-S value).
#     fn_points   : list of (N, F(>=N)) tuples, one per UNIQUE N value,
#                   sorted by N ascending.    These are the "knots" of the
#                   CCDF and the data plotted in column T,U of the sheet.
#     staircase   : (xs, ys) lists ready to feed matplotlib's plot() to draw
#                   the staircase the same way the sheet does.
#     """
#     # Keep only scenarios that contribute to the curve: positive frequency
#     # AND at least one fatality. The Excel sheet skips empty/zero rows the
#     # same way (rows 15-22 of FNCurve2 only carry T/U because their O is 0).
#     rows = [
#         dict(s) for s in scenarios
#         if (s.get("freq") or 0.0) > 0.0
#            and (s.get("fatalities") or 0.0) > 0.0
#     ]

#     # 1. Sort by fatalities DESCENDING, like FN_CURVE_CREATE3 in the VBA.
#     rows.sort(key=lambda r: -float(r["fatalities"]))

#     # 2. Build the CCDF column-S: top-down running sum of frequencies.
#     #    Because the rows are sorted by N descending, this running sum at
#     #    row i is exactly F(N >= N_i) = sum of freq of all events whose N >= N_i.
#     cumul = 0.0
#     for r in rows:
#         cumul += float(r["freq"])
#         r["cumul_freq"] = cumul

#     # 3. Collapse rows that share the same N into one knot, taking the
#     #    LARGEST cumulative frequency at that N (i.e. the value AFTER all
#     #    rows with that N have been added). This is what the staircase
#     #    drops to when crossing that N from the left.
#     knots: dict[float, float] = {}
#     for r in rows:
#         n = float(r["fatalities"])
#         f = r["cumul_freq"]
#         # Because rows are processed in N-descending order, the first time
#         # we see a given N has the SMALLEST cumul; subsequent occurrences of
#         # the same N have larger cumul. We want the LARGEST = the value
#         # after the last duplicate row was summed in.
#         if (n not in knots) or (f > knots[n]):
#             knots[n] = f
#     fn_points = sorted(knots.items(), key=lambda p: p[0])   # ascending N

#     # 4. Staircase: walk fn_points left-to-right (smallest N first); at each
#     #    knot the curve is HORIZONTAL out to the next knot's N, then VERTICAL
#     #    DOWN to the next knot's F. Same convention as FNCurve2 columns T,U.
#     xs: List[float] = []
#     ys: List[float] = []
#     for i, (n_curr, f_curr) in enumerate(fn_points):
#         if i == 0:
#             xs.append(n_curr)
#             ys.append(f_curr)
#         if i < len(fn_points) - 1:
#             n_next, f_next = fn_points[i + 1]
#             # Horizontal segment to next N at the current F level
#             xs.append(n_next)
#             ys.append(f_curr)
#             # Vertical drop down to the next F level at that same N
#             xs.append(n_next)
#             ys.append(f_next)
#     return rows, fn_points, (xs, ys)


# # ─────────────────────────────────────────────────────────────────────────────
# # Workbook loader — reads the FNCurve2 sheet using the SAME columns as VBA
# # ─────────────────────────────────────────────────────────────────────────────
# def _load_scenarios_from_workbook(workbook_path: Path) -> List[dict]:
#     """Read scenarios from the FNCurve2 sheet of FNCV_ROAD.xlsm.

#     Expected sheet layout (matches the workbook in the project root):

#         Row 1: headers
#         Row 2: header row 2
#         Row 3..N (left block, columns A..K): per-vehicle/category aggregates.
#                                              We DON'T plot these.
#         Row 25..(end): the per-scenario list that drives the plot.
#             A = Fire Point
#             B = Scenario code
#             C = Description
#             E = Frequency / yr
#             F = Frequency / veh-km
#             H = Fatalities  (사망자수)
#     """
#     try:
#         from openpyxl import load_workbook
#     except ImportError as exc:                              # pragma: no cover
#         raise RuntimeError(
#             "openpyxl is required to load FNCV_ROAD.xlsm — pip install openpyxl"
#         ) from exc

#     wb = load_workbook(str(workbook_path), data_only=True, keep_vba=True)
#     if "FNCurve2" not in wb.sheetnames:
#         raise RuntimeError(
#             f"Workbook {workbook_path.name!r} has no FNCurve2 sheet "
#             f"(found: {wb.sheetnames})."
#         )
#     ws = wb["FNCurve2"]

#     scenarios: List[dict] = []
#     # Scenario rows start at row 25 in the project's workbook. We probe
#     # for the first row whose A column is a number to be defensive against
#     # template revisions.
#     start_row = None
#     for row in range(20, ws.max_row + 1):
#         a = ws.cell(row=row, column=1).value
#         if isinstance(a, (int, float)):
#             start_row = row
#             break
#     if start_row is None:
#         return scenarios

#     for row in range(start_row, ws.max_row + 1):
#         pos  = ws.cell(row=row, column=1).value          # A
#         name = ws.cell(row=row, column=2).value          # B
#         desc = ws.cell(row=row, column=3).value          # C
#         freq = ws.cell(row=row, column=5).value          # E  freq /yr
#         fat  = ws.cell(row=row, column=8).value          # H  사망자수
#         if pos is None and name is None and freq is None and fat is None:
#             continue
#         try:
#             f = float(freq) if freq is not None else 0.0
#         except (TypeError, ValueError):
#             f = 0.0
#         try:
#             n = float(fat) if fat is not None else 0.0
#         except (TypeError, ValueError):
#             n = 0.0
#         scenarios.append({
#             "pos":        pos,
#             "name":       str(name) if name is not None else "",
#             "desc":       str(desc) if desc is not None else "",
#             "freq":       f,
#             "fatalities": n,
#         })
#     return scenarios


# # ─────────────────────────────────────────────────────────────────────────────
# # Widget
# # ─────────────────────────────────────────────────────────────────────────────
# class FNCurveSubTabWidget(QWidget):
#     """FN-Curve sub-tab for the Risk Calculations tab.

#     Reads scenarios from FNCV_ROAD.xlsm (FNCurve2 sheet) on construction,
#     builds the CCDF F(N>=n), shows the per-knot data table on the left
#     and the log-log staircase chart on the right.
#     """

#     DEFAULT_WORKBOOK_NAME = "FNCV_ROAD.xlsm"

#     def __init__(
#         self,
#         workbook_path: Optional[Path | str] = None,
#         parent: Optional[QWidget] = None,
#     ) -> None:
#         super().__init__(parent)

#         self._workbook_path: Optional[Path] = None
#         self._scenarios: List[dict] = []
#         self._sorted_rows: List[dict] = []
#         self._fn_points: List[Tuple[float, float]] = []
#         self._stair_xs: List[float] = []
#         self._stair_ys: List[float] = []

#         self._build_ui()

#         # Load workbook
#         if workbook_path is None:
#             workbook_path = Path(__file__).resolve().parent / self.DEFAULT_WORKBOOK_NAME
#         self.reload_from_workbook(workbook_path, silent=True)

#     # ── UI construction ───────────────────────────────────────────────────
#     def _build_ui(self) -> None:
#         outer = QVBoxLayout(self)
#         outer.setContentsMargins(6, 6, 6, 6)
#         outer.setSpacing(6)

#         # Header / status bar
#         header_row = QHBoxLayout()
#         title = QLabel("FN-Curve  —  Societal Risk  F(N ≥ n)")
#         title.setStyleSheet("font-weight:bold; font-size:13px; color:#2c3e50;")
#         header_row.addWidget(title)
#         header_row.addStretch()

#         self._reload_btn = QPushButton("📂  Load workbook…")
#         self._reload_btn.setStyleSheet(
#             "QPushButton { background:#2980b9; color:white; padding:4px 10px; border-radius:3px; }"
#             "QPushButton:hover { background:#3498db; }"
#         )
#         self._reload_btn.clicked.connect(self._on_browse_workbook)
#         header_row.addWidget(self._reload_btn)

#         self._save_png_btn = QPushButton("💾  Save PNG")
#         self._save_png_btn.setStyleSheet(
#             "QPushButton { background:#27ae60; color:white; padding:4px 10px; border-radius:3px; }"
#             "QPushButton:hover { background:#2ecc71; }"
#         )
#         self._save_png_btn.clicked.connect(self._on_save_png)
#         header_row.addWidget(self._save_png_btn)

#         outer.addLayout(header_row)

#         self._status_lbl = QLabel("Loading workbook…")
#         self._status_lbl.setStyleSheet("color:#666; font-size:11px;")
#         outer.addWidget(self._status_lbl)

#         # Splitter: data table on the left, chart on the right
#         splitter = QSplitter(Qt.Horizontal)
#         splitter.setChildrenCollapsible(False)
#         outer.addWidget(splitter, 1)

#         # ── Left: data table (mirrors columns L,M,R,O,S of the sheet) ─────
#         left_panel = QWidget()
#         left_v = QVBoxLayout(left_panel)
#         left_v.setContentsMargins(0, 0, 0, 0)
#         left_v.setSpacing(4)

#         left_v.addWidget(self._info_label(
#             "Sorted scenarios (by fatalities descending). "
#             "Cumul. Freq. = F(N ≥ n) — sum of freq of all events with N ≥ this row's N."
#         ))

#         self._table = QTableWidget(0, 5)
#         self._table.setHorizontalHeaderLabels([
#             "Pos", "Scenario", "Fatalities (N)", "Freq /yr (f)", "Cumul. Freq. F(≥N)"
#         ])
#         self._table.setEditTriggers(QAbstractItemView.NoEditTriggers)
#         self._table.setSelectionBehavior(QAbstractItemView.SelectRows)
#         self._table.setAlternatingRowColors(True)
#         self._table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
#         self._table.horizontalHeader().setStretchLastSection(True)
#         self._table.verticalHeader().setDefaultSectionSize(20)
#         left_v.addWidget(self._table, 1)

#         splitter.addWidget(left_panel)

#         # ── Right: chart + ALARP controls ─────────────────────────────────
#         right_panel = QWidget()
#         right_v = QVBoxLayout(right_panel)
#         right_v.setContentsMargins(0, 0, 0, 0)
#         right_v.setSpacing(4)

#         if _HAVE_MPL:
#             self._fig = Figure(figsize=(7, 5))
#             self._fig.patch.set_facecolor("#f8f9fa")
#             self._canvas = FigureCanvas(self._fig)
#             self._ax = self._fig.add_subplot(111)
#             right_v.addWidget(self._canvas, 1)
#         else:                                                # pragma: no cover
#             self._canvas = None
#             self._ax = None
#             no_mpl = QLabel(
#                 "matplotlib is not available — install it to see the FN curve "
#                 "(pip install matplotlib)."
#             )
#             no_mpl.setAlignment(Qt.AlignCenter)
#             no_mpl.setStyleSheet("color:#7f8c8d; font-size:13px;")
#             right_v.addWidget(no_mpl, 1)

#         # ALARP controls
#         ctrl = QHBoxLayout()
#         ctrl.setSpacing(6)
#         ctrl.addWidget(QLabel("ALARP upper:"))
#         self._alarp_upper = QDoubleSpinBox()
#         self._alarp_upper.setRange(1e-12, 1.0)
#         self._alarp_upper.setDecimals(10)
#         self._alarp_upper.setValue(1e-4)
#         self._alarp_upper.setFixedWidth(110)
#         ctrl.addWidget(self._alarp_upper)

#         ctrl.addWidget(QLabel("lower:"))
#         self._alarp_lower = QDoubleSpinBox()
#         self._alarp_lower.setRange(1e-12, 1.0)
#         self._alarp_lower.setDecimals(10)
#         self._alarp_lower.setValue(1e-6)
#         self._alarp_lower.setFixedWidth(110)
#         ctrl.addWidget(self._alarp_lower)

#         self._show_alarp_cb = QCheckBox("Show ALARP")
#         self._show_alarp_cb.setChecked(True)
#         self._show_alarp_cb.toggled.connect(self._redraw)
#         ctrl.addWidget(self._show_alarp_cb)

#         self._show_grid_cb = QCheckBox("Grid")
#         self._show_grid_cb.setChecked(True)
#         self._show_grid_cb.toggled.connect(self._redraw)
#         ctrl.addWidget(self._show_grid_cb)

#         self._alarp_upper.valueChanged.connect(self._redraw)
#         self._alarp_lower.valueChanged.connect(self._redraw)

#         redraw_btn = QPushButton("🔁  Redraw")
#         redraw_btn.setStyleSheet(
#             "QPushButton { background:#34495e; color:white; padding:3px 9px; border-radius:3px; }"
#         )
#         redraw_btn.clicked.connect(self._redraw)
#         ctrl.addWidget(redraw_btn)
#         ctrl.addStretch()
#         right_v.addLayout(ctrl)

#         splitter.addWidget(right_panel)
#         splitter.setStretchFactor(0, 2)
#         splitter.setStretchFactor(1, 3)
#         splitter.setSizes([400, 700])

#     @staticmethod
#     def _info_label(text: str) -> QLabel:
#         lbl = QLabel(text)
#         lbl.setWordWrap(True)
#         lbl.setStyleSheet("color:#555; font-size:11px; padding:2px;")
#         return lbl

#     # ── Public API ────────────────────────────────────────────────────────
#     def reload_from_workbook(
#         self,
#         workbook_path: Path | str,
#         silent: bool = False,
#     ) -> bool:
#         """Reload scenarios from the given workbook and redraw.

#         Returns True on success, False otherwise.
#         """
#         path = Path(workbook_path)
#         if not path.exists():
#             msg = (
#                 f"Workbook not found:\n{path}\n\n"
#                 f"Click 'Load workbook…' to pick the correct FNCV_ROAD.xlsm."
#             )
#             self._status_lbl.setText(f"⚠  {path.name} not found.")
#             self._status_lbl.setStyleSheet("color:#c0392b; font-size:11px;")
#             if not silent:
#                 QMessageBox.warning(self, "FN-Curve", msg)
#             return False
#         try:
#             scenarios = _load_scenarios_from_workbook(path)
#         except Exception as exc:
#             self._status_lbl.setText(f"⚠  Error loading {path.name}: {exc}")
#             self._status_lbl.setStyleSheet("color:#c0392b; font-size:11px;")
#             if not silent:
#                 QMessageBox.warning(self, "FN-Curve", str(exc))
#             return False

#         self._workbook_path = path
#         self.set_scenarios(scenarios)
#         return True

#     def set_scenarios(self, scenarios: Iterable[dict]) -> None:
#         """Inject scenarios programmatically and rebuild the curve.

#         Each scenario must be a dict with at minimum:
#             'fatalities' : float (N)
#             'freq'       : float (events / year)
#         Other keys ('pos', 'name', 'desc', …) are kept for the table.
#         """
#         self._scenarios = [dict(s) for s in scenarios]
#         self._recompute()
#         self._populate_table()
#         self._redraw()

#         n_total   = len(self._scenarios)
#         n_used    = len(self._sorted_rows)
#         n_knots   = len(self._fn_points)
#         if n_used == 0:
#             self._status_lbl.setText(
#                 f"⚠  Loaded {n_total} scenario(s) but none have non-zero "
#                 f"frequency AND fatalities — FN-curve is empty. "
#                 f"Run the EVC analysis to populate fatalities (column H of FNCurve2)."
#             )
#             self._status_lbl.setStyleSheet("color:#c0392b; font-size:11px;")
#         else:
#             self._status_lbl.setText(
#                 f"✅  {n_used} contributing scenario(s) → {n_knots} unique N "
#                 f"value(s).  Total freq F(N≥0) = {self._sorted_rows[-1]['cumul_freq']:.4e} /yr."
#             )
#             self._status_lbl.setStyleSheet("color:#27ae60; font-size:11px;")

#     def get_fn_points(self) -> List[Tuple[float, float]]:
#         """Return [(N, F(>=N)), ...] sorted by N ascending."""
#         return list(self._fn_points)

#     def get_staircase(self) -> Tuple[List[float], List[float]]:
#         """Return (xs, ys) ready to plot the staircase on a log-log axis."""
#         return list(self._stair_xs), list(self._stair_ys)

#     # ── Internal: recompute ───────────────────────────────────────────────
#     def _recompute(self) -> None:
#         sorted_rows, fn_points, (xs, ys) = build_fn_curve(self._scenarios)
#         self._sorted_rows = sorted_rows
#         self._fn_points = fn_points
#         self._stair_xs = xs
#         self._stair_ys = ys

#     # ── Internal: populate table ──────────────────────────────────────────
#     def _populate_table(self) -> None:
#         self._table.setRowCount(0)

#         def _ci(text, bg=None, align=Qt.AlignCenter):
#             it = QTableWidgetItem(str(text))
#             it.setTextAlignment(align)
#             if bg is not None:
#                 it.setBackground(bg)
#             return it

#         zebra = QColor("#fbfcfd")
#         for r in self._sorted_rows:
#             ri = self._table.rowCount()
#             self._table.insertRow(ri)
#             self._table.setItem(ri, 0, _ci(r.get("pos", "")))
#             self._table.setItem(ri, 1, _ci(r.get("name", "")))
#             self._table.setItem(ri, 2, _ci(f"{float(r['fatalities']):.2f}"))
#             self._table.setItem(ri, 3, _ci(f"{float(r['freq']):.4e}"))
#             self._table.setItem(
#                 ri, 4,
#                 _ci(f"{float(r['cumul_freq']):.4e}", bg=zebra),
#             )

#     # ── Internal: redraw chart ────────────────────────────────────────────
#     def _redraw(self) -> None:
#         if not _HAVE_MPL or self._ax is None:
#             return
#         ax = self._ax
#         ax.clear()

#         if not self._stair_xs:
#             ax.text(
#                 0.5, 0.5, "No FN data — load FNCV_ROAD.xlsm",
#                 transform=ax.transAxes, ha="center", va="center",
#                 fontsize=12, color="#7f8c8d",
#                 bbox=dict(boxstyle="round,pad=0.5", fc="#ecf0f1", alpha=0.85),
#             )
#             ax.set_xscale("linear")
#             ax.set_yscale("linear")
#             self._fig.tight_layout()
#             self._canvas.draw_idle()
#             return

#         xs = list(self._stair_xs)
#         ys = list(self._stair_ys)

#         # ALARP bands
#         if self._show_alarp_cb.isChecked():
#             x_lo = min(xs) * 0.5 if min(xs) > 0 else 0.1
#             x_hi = max(xs) * 2.0
#             x_band = [x_lo, x_hi]
#             f_up = float(self._alarp_upper.value())
#             f_lo = float(self._alarp_lower.value())
#             if f_up > f_lo:
#                 ax.fill_between(
#                     x_band, [f_up, f_up], [1.0, 1.0],
#                     color="#e74c3c", alpha=0.12,
#                     label=f"Intolerable (>{f_up:.0E}/yr)",
#                 )
#                 ax.fill_between(
#                     x_band, [f_lo, f_lo], [f_up, f_up],
#                     color="#f39c12", alpha=0.10,
#                     label="ALARP region",
#                 )
#                 ax.fill_between(
#                     x_band, [1e-14, 1e-14], [f_lo, f_lo],
#                     color="#27ae60", alpha=0.10,
#                     label=f"Broadly acceptable (<{f_lo:.0E}/yr)",
#                 )

#         # The staircase itself
#         ax.plot(
#             xs, ys,
#             "-", color="#2c3e50", linewidth=2.0,
#             label="F-N curve",
#         )
#         # Markers at the knots only (one per unique N)
#         if self._fn_points:
#             knot_xs = [p[0] for p in self._fn_points]
#             knot_ys = [p[1] for p in self._fn_points]
#             ax.plot(
#                 knot_xs, knot_ys,
#                 "o", color="#c0392b", markersize=5,
#                 label="Scenario knots",
#             )

#         ax.set_xscale("log")
#         ax.set_yscale("log")
#         ax.set_xlabel("Number of Fatalities  N", fontsize=11)
#         ax.set_ylabel("Cumulative Frequency  F(N ≥ n)  [events / year]", fontsize=11)
#         ax.set_title("F-N Curve  (Societal Risk)", fontsize=13, fontweight="bold")
#         if self._show_grid_cb.isChecked():
#             ax.grid(True, which="both", linewidth=0.5, alpha=0.5)
#         else:
#             ax.grid(False)
#         ax.legend(fontsize=9, loc="lower left", framealpha=0.85)
#         # Tight x-range hugging the data
#         ax.set_xlim(min(xs) * 0.7, max(xs) * 1.4)
#         self._fig.tight_layout()
#         self._canvas.draw_idle()

#     # ── Slot: browse for workbook ────────────────────────────────────────
#     def _on_browse_workbook(self) -> None:
#         start_dir = (
#             str(self._workbook_path.parent) if self._workbook_path else str(Path.cwd())
#         )
#         path, _ = QFileDialog.getOpenFileName(
#             self,
#             "Select FN-Curve workbook",
#             start_dir,
#             "Excel files (*.xlsm *.xlsx);;All files (*.*)",
#         )
#         if path:
#             self.reload_from_workbook(path, silent=False)

#     # ── Slot: save PNG ───────────────────────────────────────────────────
#     def _on_save_png(self) -> None:
#         if not _HAVE_MPL or self._canvas is None:
#             QMessageBox.information(
#                 self, "FN-Curve", "matplotlib is not available — cannot save PNG."
#             )
#             return
#         path, _ = QFileDialog.getSaveFileName(
#             self, "Save FN curve as PNG", "fn_curve.png", "PNG files (*.png)"
#         )
#         if path:
#             try:
#                 self._fig.savefig(path, dpi=200, bbox_inches="tight")
#             except Exception as exc:
#                 QMessageBox.warning(self, "FN-Curve", f"Could not save: {exc}")


# # ─────────────────────────────────────────────────────────────────────────────
# # Self-test
# # ─────────────────────────────────────────────────────────────────────────────
# if __name__ == "__main__":                                  # pragma: no cover
#     import sys
#     from PyQt5.QtWidgets import QApplication

#     app = QApplication(sys.argv)
#     here = Path(__file__).resolve().parent
#     wb = here / "FNCV_ROAD.xlsm"
#     w = FNCurveSubTabWidget(workbook_path=wb if wb.exists() else None)
#     w.resize(1100, 600)
#     w.setWindowTitle("FN-Curve — standalone test")
#     w.show()
#     sys.exit(app.exec_())


# """
# fn_curve_subtab_widget.py

# Risk Calculations sub-tab that recreates FNCurve2 table from FNCV_ROAD.xlsm
# and plots the FN curve (N vs F).
# """

# from pathlib import Path
# from typing import Any, Dict, List, Optional

# from PyQt5.QtCore import Qt
# from PyQt5.QtWidgets import (
#     QLabel,
#     QMessageBox,
#     QPushButton,
#     QTableWidget,
#     QTableWidgetItem,
#     QVBoxLayout,
#     QWidget,
# )


# class FNCurveSubTabWidget(QWidget):
#     def __init__(self, parent=None):
#         super().__init__(parent)
#         self._table_rows: List[Dict[str, Any]] = []
#         self._build_ui()
#         self.reload_from_excel()

#     def _build_ui(self):
#         root = QVBoxLayout(self)
#         root.setContentsMargins(6, 6, 6, 6)
#         root.setSpacing(6)

#         title = QLabel("FN-Curve - FNCV_ROAD.xlsm (FNCurve2)")
#         title.setStyleSheet(
#             "font-size:13px;font-weight:bold;color:#1a5276;"
#             "padding:4px 8px;background:#d5e8f8;border-radius:4px;"
#         )
#         root.addWidget(title)

#         self.status_label = QLabel("Loading workbook data...")
#         self.status_label.setStyleSheet("font-size:11px;color:#555;padding:2px 4px;")
#         root.addWidget(self.status_label)

#         self.table = QTableWidget(0, 10)
#         self.table.setHorizontalHeaderLabels([
#             "Fire Point",
#             "Scenario No",
#             "Description",
#             "Frequency /yr",
#             "Frequency /veh-km",
#             "Return Year",
#             "Fatalities",
#             "Cumulative Freq.",
#             "N",
#             "F",
#         ])
#         self.table.setAlternatingRowColors(True)
#         self.table.setEditTriggers(QTableWidget.NoEditTriggers)
#         self.table.setSelectionBehavior(QTableWidget.SelectRows)
#         self.table.horizontalHeader().setStretchLastSection(True)
#         root.addWidget(self.table, 1)

#         self._fig = None
#         self._ax = None
#         self._canvas = None
#         try:
#             import matplotlib
#             matplotlib.use("Agg")
#             import matplotlib.pyplot as plt
#             from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas

#             self._fig, self._ax = plt.subplots(figsize=(9, 4.2))
#             self._canvas = FigureCanvas(self._fig)
#             self._canvas.setMinimumHeight(260)
#             root.addWidget(self._canvas)
#         except Exception:
#             no_plot = QLabel("matplotlib not available - FN chart cannot be displayed.")
#             no_plot.setStyleSheet("font-size:11px;color:#7f8c8d;padding:4px;")
#             root.addWidget(no_plot)

#         reload_btn = QPushButton("Reload FN Data")
#         reload_btn.setStyleSheet(
#             "QPushButton{background:#2980b9;color:white;font-weight:bold;"
#             "border-radius:3px;padding:4px 10px;}"
#             "QPushButton:hover{background:#1f6692;}"
#         )
#         reload_btn.clicked.connect(self.reload_from_excel)
#         root.addWidget(reload_btn, alignment=Qt.AlignRight)

#     def reload_from_excel(self):
#         try:
#             rows = self._read_fncurve2_rows()
#             self._table_rows = rows
#             self._populate_table(rows)
#             self._draw_chart(rows)
#             self.status_label.setText(f"Loaded {len(rows)} rows from FNCurve2.")
#         except Exception as exc:
#             self.status_label.setText(f"Load failed: {exc}")
#             QMessageBox.warning(self, "FN-Curve", f"Failed to load FNCurve2 data.\n\n{exc}")

#     def _read_fncurve2_rows(self) -> List[Dict[str, Any]]:
#         try:
#             from openpyxl import load_workbook
#         except Exception as exc:
#             raise RuntimeError("openpyxl is required to read FNCV_ROAD.xlsm") from exc

#         wb_path = Path(__file__).with_name("FNCV_ROAD.xlsm")
#         if not wb_path.exists():
#             raise FileNotFoundError(f"Workbook not found: {wb_path}")

#         wb = load_workbook(wb_path, data_only=True, read_only=True)
#         if "FNCurve2" not in wb.sheetnames:
#             raise RuntimeError("Sheet 'FNCurve2' not found in workbook")

#         ws = wb["FNCurve2"]

#         header_row = self._find_header_row(ws)
#         sub_header_row = header_row + 1
#         data_start = sub_header_row + 1

#         col_map = self._resolve_columns(ws, header_row, sub_header_row)

#         rows: List[Dict[str, Any]] = []
#         blank_streak = 0
#         for r in range(data_start, min(ws.max_row + 1, data_start + 3000)):
#             rec = {
#                 "fire_point": ws.cell(r, col_map["fire_point"]).value,
#                 "scenario_no": ws.cell(r, col_map["scenario_no"]).value,
#                 "description": ws.cell(r, col_map["description"]).value,
#                 "frequency_yr": ws.cell(r, col_map["frequency_yr"]).value,
#                 "frequency_vkm": ws.cell(r, col_map["frequency_vkm"]).value,
#                 "return_year": ws.cell(r, col_map["return_year"]).value,
#                 "fatalities": ws.cell(r, col_map["fatalities"]).value,
#                 "cumulative_freq": ws.cell(r, col_map["cumulative_freq"]).value,
#                 "n": ws.cell(r, col_map["n"]).value,
#                 "f": ws.cell(r, col_map["f"]).value,
#             }

#             # Skip pre-table captions and separator rows until numeric data starts.
#             has_numeric_signal = any(
#                 self._to_float(rec.get(k)) is not None
#                 for k in ("frequency_yr", "fatalities", "cumulative_freq", "n", "f")
#             )
#             if not has_numeric_signal:
#                 blank_streak += 1
#                 if blank_streak >= 20:
#                     break
#                 continue

#             key_empty = (
#                 rec["fire_point"] is None
#                 and rec["scenario_no"] is None
#                 and rec["description"] is None
#                 and rec["frequency_yr"] is None
#                 and rec["fatalities"] is None
#                 and rec["n"] is None
#                 and rec["f"] is None
#             )
#             if key_empty:
#                 blank_streak += 1
#                 if blank_streak >= 8:
#                     break
#                 continue

#             blank_streak = 0
#             rows.append(rec)

#         wb.close()
#         return rows

#     def _find_header_row(self, ws) -> int:
#         def _norm(v: Any) -> str:
#             return str(v or "").strip().lower().replace("\n", " ")

#         # First pass: robust semantic detection (English/Korean fragments).
#         for r in range(1, min(120, ws.max_row + 1)):
#             values = [_norm(ws.cell(r, c).value) for c in range(1, min(120, ws.max_column + 1))]
#             has_fire = any(("fire point" in v) or ("fire" in v and "point" in v) or ("화재" in v) for v in values)
#             has_freq = any(("frequency" in v) or ("빈도" in v) for v in values)
#             has_fatal = any(("fatal" in v) or ("사망" in v) for v in values)
#             if has_fire and has_freq and has_fatal:
#                 return r

#         # Second pass: detect row with N/F subheaders and assume header is the row above.
#         for r in range(2, min(200, ws.max_row + 1)):
#             values = [_norm(ws.cell(r, c).value) for c in range(1, min(140, ws.max_column + 1))]
#             for c in range(0, len(values) - 1):
#                 if values[c] == "n" and values[c + 1] == "f":
#                     return r - 1

#         # Last fallback: common workbook pattern around upper table area.
#         return 1

#     def _resolve_columns(self, ws, header_row: int, sub_header_row: int) -> Dict[str, int]:
#         def _norm(v: Any) -> str:
#             return str(v or "").strip().lower().replace("\n", " ")

#         max_scan = min(160, ws.max_column + 1)
#         top = [_norm(ws.cell(header_row, c).value) for c in range(1, max_scan)]
#         sub = [_norm(ws.cell(sub_header_row, c).value) for c in range(1, max_scan)]

#         def _find_top(fragment: str) -> Optional[int]:
#             for i, v in enumerate(top, start=1):
#                 if fragment in v:
#                     return i
#             return None

#         col_fire = _find_top("fire point") or _find_top("fire") or _find_top("화재")
#         col_freq_yr = _find_top("frequency /yr") or _find_top("frequency/yr") or _find_top("frequency") or _find_top("빈도")
#         col_freq_vkm = _find_top("frequency /veh") or _find_top("frequency/veh") or _find_top("veh-km")
#         col_return = _find_top("return year") or _find_top("return") or _find_top("회귀")
#         col_fatal = _find_top("fatal") or _find_top("사망")
#         col_cum = _find_top("cumulative") or _find_top("누적")

#         col_desc = None
#         for i, v in enumerate(sub, start=1):
#             if v == "description":
#                 col_desc = i
#                 break
#         if col_desc is None:
#             col_desc = (col_fire or 1) + 2

#         col_no = None
#         for i, v in enumerate(sub, start=1):
#             if v == "no" and abs(i - col_desc) <= 2:
#                 col_no = i
#                 break
#         if col_no is None:
#             col_no = max(1, col_desc - 1)

#         search_start = (col_cum or col_desc or 1) + 1
#         col_n = None
#         col_f = None
#         for i in range(search_start, max_scan):
#             v_sub = sub[i - 1]
#             v_top = top[i - 1]
#             if (v_sub == "n" or v_top == "n") and col_n is None:
#                 col_n = i
#             elif (v_sub == "f" or v_top == "f") and col_f is None:
#                 col_f = i

#         # Global fallback: locate exact N/F anywhere on header rows.
#         if col_n is None:
#             for i in range(1, max_scan):
#                 if sub[i - 1] == "n" or top[i - 1] == "n":
#                     col_n = i
#                     break
#         if col_f is None and col_n is not None:
#             for i in range(col_n + 1, min(col_n + 4, max_scan)):
#                 if sub[i - 1] == "f" or top[i - 1] == "f":
#                     col_f = i
#                     break

#         if col_n is None:
#             col_n = (col_cum or col_desc or 8) + 1
#         if col_f is None:
#             col_f = col_n + 1

#         # Position-based fallbacks around the N/F pair for mixed/merged headers.
#         if col_cum is None:
#             col_cum = max(1, col_n - 1)
#         if col_fatal is None:
#             col_fatal = max(1, col_cum - 1)
#         if col_return is None:
#             col_return = max(1, col_fatal - 1)
#         if col_freq_vkm is None:
#             col_freq_vkm = max(1, col_return - 1)
#         if col_freq_yr is None:
#             col_freq_yr = max(1, col_freq_vkm - 1)
#         if col_desc is None:
#             col_desc = max(1, col_freq_yr - 1)
#         if col_no is None:
#             col_no = max(1, col_desc - 1)
#         if col_fire is None:
#             col_fire = max(1, col_no - 1)

#         required = {
#             "fire_point": col_fire,
#             "scenario_no": col_no,
#             "description": col_desc,
#             "frequency_yr": col_freq_yr,
#             "frequency_vkm": col_freq_vkm,
#             "return_year": col_return,
#             "fatalities": col_fatal,
#             "cumulative_freq": col_cum,
#             "n": col_n,
#             "f": col_f,
#         }
#         return required

#     def _to_float(self, v: Any) -> Optional[float]:
#         try:
#             if v is None:
#                 return None
#             return float(v)
#         except Exception:
#             return None

#     def _fmt(self, v: Any) -> str:
#         if v is None:
#             return ""
#         if isinstance(v, float):
#             if abs(v) >= 1000 or (abs(v) > 0 and abs(v) < 0.001):
#                 return f"{v:.4E}"
#             return f"{v:.4f}".rstrip("0").rstrip(".")
#         return str(v)

#     def _populate_table(self, rows: List[Dict[str, Any]]):
#         self.table.setRowCount(0)
#         for rec in rows:
#             r = self.table.rowCount()
#             self.table.insertRow(r)
#             vals = [
#                 rec["fire_point"],
#                 rec["scenario_no"],
#                 rec["description"],
#                 rec["frequency_yr"],
#                 rec["frequency_vkm"],
#                 rec["return_year"],
#                 rec["fatalities"],
#                 rec["cumulative_freq"],
#                 rec["n"],
#                 rec["f"],
#             ]
#             for c, v in enumerate(vals):
#                 item = QTableWidgetItem(self._fmt(v))
#                 item.setTextAlignment(Qt.AlignCenter)
#                 self.table.setItem(r, c, item)

#     def _draw_chart(self, rows: List[Dict[str, Any]]):
#         if self._ax is None or self._canvas is None:
#             return

#         ns: List[float] = []
#         fs: List[float] = []
#         for rec in rows:
#             try:
#                 n = float(rec.get("n"))
#                 f = float(rec.get("f"))
#             except Exception:
#                 continue
#             if n > 0 and f > 0:
#                 ns.append(n)
#                 fs.append(f)

#         self._ax.clear()
#         self._ax.set_title("FN Curve")
#         self._ax.set_xlabel("N (Fatalities)")
#         self._ax.set_ylabel("F (Cumulative Frequency /yr)")
#         self._ax.set_xscale("log")
#         self._ax.set_yscale("log")
#         self._ax.grid(True, which="both", alpha=0.45)

#         if ns and fs:
#             pairs = sorted(zip(ns, fs), key=lambda x: x[0])
#             xs, ys = zip(*pairs)
#             self._ax.plot(xs, ys, marker="o", linewidth=1.5, markersize=4, color="#2E86C1")

#         self._fig.tight_layout()
#         self._canvas.draw()
