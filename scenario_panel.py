"""
scenario_panel.py
====================================================================
"Scenario" panel for the General Input Information tab.

Left 20% : one radio per sheet of scenarios.xlsx, and a compact "Control"
           table (Vehicle | ACCR | N-Serious | FG1 | FG2) loaded from the
           selected sheet.  Editing the Control recomputes the node values
           in the diagram via the event-tree relationship
               leaf_freq = vehicle_CaseYr x P(path)
           and every connector percentage is derived from the node values.
Right 80%: the event-tree diagram reconstructed from the selected sheet,
           in the style of the Standard Scenario sub-tab.
"""

from __future__ import annotations

import os
import re
import math
import copy
from collections import Counter

from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QGroupBox, QLabel, QRadioButton,
    QButtonGroup, QTableWidget, QTableWidgetItem, QPushButton, QScrollArea,
    QHeaderView, QSizePolicy, QAbstractScrollArea, QCheckBox,
)
from PyQt5.QtCore import Qt, QRectF, QPointF, pyqtSignal
from PyQt5.QtGui import QPainter, QColor, QPen, QBrush, QFont

try:
    import openpyxl as _openpyxl
    _HAS_OPENPYXL = True
except Exception:  # pragma: no cover
    _openpyxl = None
    _HAS_OPENPYXL = False

_FALLBACK_SHEETS = [
    "Longitudinal-All Vehicles",
    "Longitudinal-Small vehicles onl",
    "Transverse_Small Vehicles",
]

_VEHICLE_KW = [
    "전차종", "승용차", "버스", "화물", "승합", "위험물", "특수", "수송",
    "대형", "소형",
    "passenger", "car", "bus", "truck", "cargo", "special", "hazard", "vehicle",
    "hgv", "lorry", "van",
]
_FIRE_KW = {
    "not serious": "Not Serious",
    "단독화재": "단독화재",
    "2대연속화재": "2대연속화재",
    "연속화재": "연속화재",
    "화재확산": "화재확산",
}
_STATE_KW = {"normal": "Normal", "congest": "Congest", "congestion": "Congest"}

# Korean -> English for display (longest / most specific first).  Applied only
# when rendering; the raw names are kept for vehicle classification.
_KO_EN = [
    ("중대형 화물", "Heavy Cargo Truck"),
    ("위험물수송", "Hazardous Cargo"),
    ("2대연속화재", "2-Veh Fire"),
    ("전차종", "All Vehicles"),
    ("승용차", "Passenger Car"),
    ("단독화재", "Single Fire"),
    ("연속화재", "Consecutive Fire"),
    ("화재확산", "Fire Spread"),
    ("버스", "Bus"),
    ("화물", "Cargo Truck"),
    ("위험물", "Hazardous"),
    ("수송", "Transport"),
    ("이하", ""),
    ("PASSENGER CAR", "Passenger Car"),
    ("Samll Bus", "Small Bus"),
]


def _en(s):
    """Translate any Korean terms in a display string to English."""
    if not s:
        return s
    for ko, en in _KO_EN:
        if ko in s:
            s = s.replace(ko, en)
    return s.strip()


def _num(v):
    return v if (isinstance(v, (int, float)) and not isinstance(v, bool) and math.isfinite(v)) else None


def _is_vehicle(s):
    low = s.lower()
    if any(ch.isdigit() for ch in s):
        return False
    return any(k.lower() in low for k in _VEHICLE_KW)


def _fire_label(s):
    low = s.strip().lower()
    for k, lab in _FIRE_KW.items():
        if k.lower() in low:
            return lab
    return None


def classify_vehicle(name):
    """Map a sheet vehicle name to a Control row index (0 PC, 1 BUS, 2 GV, 3 ST)."""
    low = (name or "").lower()
    if "위험물" in name or "특수" in name or "hazard" in low or "special" in low:
        return 3
    if "버스" in name or "bus" in low:
        return 1
    if "승용" in name or "passenger" in low or ("car" in low and "cargo" not in low):
        return 0
    if "화물" in name or "대형" in name or "cargo" in low or "truck" in low:
        return 2
    return None


# ====================================================================
#  Parser
# ====================================================================
def parse_scenario_tree(ws):
    max_r = min(ws.max_row, 400)
    max_c = min(ws.max_column, 40)

    freq_col = ry_col = rate_col = scen_col = None
    header_row = 1
    for r in range(1, 7):
        for c in range(1, max_c + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str):
                t = v.strip().lower()
                if "frequency per yr" in t:
                    freq_col, header_row = c, r
                elif "return year" in t:
                    ry_col = c
                elif "사고발생율" in v:
                    rate_col = c
                elif "scenario no" in t and scen_col is None:
                    scen_col = c
    if freq_col is None:
        return []
    if rate_col is None:
        rate_col = freq_col - 1

    state_cnt, fire_cnt = Counter(), Counter()
    for r in range(header_row + 1, max_r + 1):
        for c in range(1, freq_col):
            v = ws.cell(r, c).value
            if isinstance(v, str):
                low = v.strip().lower()
                if low in _STATE_KW:
                    state_cnt[c] += 1
                if "mw" in low and any(ch.isdigit() for ch in v):
                    fire_cnt[c] += 1
                if _fire_label(v):
                    fire_cnt[c] += 1
    state_col = state_cnt.most_common(1)[0][0] if state_cnt else None
    fire_col = fire_cnt.most_common(1)[0][0] if fire_cnt else None
    prob_col = (fire_col - 1) if fire_col and fire_col > 1 else None

    name_hits = Counter()
    for r in range(header_row + 1, max_r + 1):
        for c in range(1, (fire_col or 4)):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.strip() and _is_vehicle(v.strip()):
                name_hits[c] += 1
                break
    if not name_hits:
        return []
    name_col = name_hits.most_common(1)[0][0]
    val_col = name_col + 1

    veh_rows = []
    for r in range(header_row + 1, max_r + 1):
        v = ws.cell(r, name_col).value
        if isinstance(v, str) and v.strip() and _is_vehicle(v.strip()):
            veh_rows.append((r, v.strip()))

    def metrics(name_row):
        accr = caseyr = ry = None
        for rr in range(name_row, min(name_row + 5, max_r + 1)):
            lab = ws.cell(rr, name_col).value
            val = _num(ws.cell(rr, val_col).value)
            if isinstance(lab, str) and val is not None:
                ll = lab.strip().lower()
                if "accr" in ll:
                    accr = val
                elif "건" in lab or ("/" in lab and "년" in lab) or "case" in ll:
                    caseyr = val
                elif ll == "ry" or "return" in ll:
                    ry = val
        if accr is None:
            nums = []
            for rr in range(name_row, min(name_row + 6, max_r + 1)):
                val = _num(ws.cell(rr, val_col).value)
                if val is not None:
                    nums.append(val)
            if len(nums) >= 3:
                accr, caseyr, ry = nums[0], nums[1], nums[2]
            elif nums:
                accr = nums[0]
        return accr, caseyr, ry

    vehicles = []
    for (r, n) in veh_rows:
        a, cy, ry = metrics(r)
        vehicles.append({"name": n, "row": r, "accr": a, "caseyr": cy, "ry": ry, "leaves": []})

    flat_leaves = []
    ns_rows = []
    cur_fire = cur_size = cur_state = None
    cur_nodevals = []
    capturing = False
    for r in range(header_row + 1, max_r + 1):
        if fire_col:
            fv = ws.cell(r, fire_col).value
            if isinstance(fv, str) and fv.strip():
                lab = _fire_label(fv)
                if lab:
                    cur_fire = lab
                    cur_size = None
                    cur_nodevals = []
                    capturing = False
                    if lab == "Not Serious":
                        ns_rows.append(r)
                low = fv.strip().lower()
                if "mw" in low and any(ch.isdigit() for ch in fv):
                    cur_size = fv.strip().replace(" ", "")
                    cur_nodevals = []
                    capturing = True
            elif capturing:
                nv = _num(fv)
                if nv is not None and len(cur_nodevals) < 3:
                    cur_nodevals.append(nv)
        if state_col:
            sv = ws.cell(r, state_col).value
            if isinstance(sv, str) and sv.strip().lower() in _STATE_KW:
                cur_state = _STATE_KW[sv.strip().lower()]
        fq = _num(ws.cell(r, freq_col).value)
        if fq is not None and fq > 0:
            sno = ws.cell(r, scen_col).value if scen_col else None
            flat_leaves.append({
                "row": r, "fire": cur_fire, "size": cur_size, "state": cur_state,
                "nodevals": list(cur_nodevals),
                "rate": _num(ws.cell(r, rate_col).value) if rate_col else None,
                "freq": fq,
                "ry": _num(ws.cell(r, ry_col).value) if ry_col else None,
                "scenario_no": str(sno).strip() if isinstance(sno, str) and sno.strip() else None,
            })
            cur_state = None

    ns_sorted = sorted(set(ns_rows))

    def assign(row):
        starts = [n for n in ns_sorted if n <= row]
        if not starts:
            return None
        bstart = max(starts)
        ends = [n for n in ns_sorted if n > bstart]
        bend = min(ends) if ends else 10 ** 9
        cands = [v for v in vehicles if bstart <= v["row"] < bend]
        if cands:
            return cands[0]
        return min(vehicles, key=lambda v: abs(v["row"] - row)) if vehicles else None

    if ns_sorted:
        for lf in flat_leaves:
            v = assign(lf["row"])
            if v is not None:
                v["leaves"].append(lf)
    else:
        for lf in flat_leaves:
            if vehicles:
                min(vehicles, key=lambda v: abs(v["row"] - lf["row"]))["leaves"].append(lf)

    # group leaves into branches; pair Normal+Congest
    for v in vehicles:
        lvs = sorted(v["leaves"], key=lambda l: l["row"])
        branches = []
        i = 0
        while i < len(lvs):
            lf = lvs[i]
            if not lf.get("state"):
                branches.append({"label": lf["fire"] or "Not Serious", "size": lf["size"],
                                 "is_ns": (lf["fire"] == "Not Serious" or not lf["size"]),
                                 "nodevals": lf["nodevals"], "leaves": [lf]})
                i += 1
                continue
            grp = [lf]
            j = i + 1
            while j < len(lvs) and lvs[j].get("state") and lvs[j]["state"] != "Normal":
                grp.append(lvs[j]); j += 1
            size = next((g["size"] for g in grp if g["size"]), None)
            fire = next((g["fire"] for g in grp if g["fire"] and g["fire"] != "Not Serious"), None)
            nodevals = next((g["nodevals"] for g in grp if g["nodevals"]), [])
            branches.append({"label": fire or (size or "fire"), "size": size,
                             "is_ns": False, "nodevals": nodevals, "leaves": grp})
            i = j
        v["branches"] = branches
        del v["leaves"]

    # baseline snapshots for recompute. Use the sum of the leaf scenarios as
    # the vehicle total so the Control percentages and the connector
    # percentages are always mutually consistent.
    out = []
    for v in vehicles:
        if not v["branches"]:
            continue
        leaf_sum = sum(l["freq"] for b in v["branches"] for l in b["leaves"])
        base_cy = leaf_sum if leaf_sum > 0 else (v["caseyr"] or 0.0)
        v["caseyr"] = base_cy
        v["ry"] = (1.0 / base_cy) if base_cy > 0 else v["ry"]
        v["base_accr"] = v["accr"]
        v["base_caseyr"] = base_cy
        for b in v["branches"]:
            b["base_nodevals"] = list(b.get("nodevals") or [])
            for l in b["leaves"]:
                l["base_freq"] = l["freq"]
                l["base_rate"] = l["rate"]
        out.append(v)

    with_accr = [v for v in out if v["accr"] is not None]
    return with_accr if with_accr else out


# ====================================================================
#  Recompute + connector percentages
# ====================================================================
def recompute_vehicle(v, accr, ns_pct, fg_list):
    base_accr = v.get("base_accr")
    base_cy = v.get("base_caseyr") or 0.0
    ratio = (accr / base_accr) if (base_accr) else 1.0
    cy = base_cy * ratio
    ns_branches = [b for b in v["branches"] if b["is_ns"]]
    fire_branches = [b for b in v["branches"] if not b["is_ns"]]
    base_ns = sum(l["base_freq"] for b in ns_branches for l in b["leaves"])
    base_fire = sum(l["base_freq"] for b in fire_branches for l in b["leaves"])

    if ns_pct is None:
        ns_frac = (base_ns / base_cy) if base_cy else 0.0
    else:
        ns_frac = max(0.0, min(1.0, ns_pct / 100.0))
    fire_frac = max(0.0, 1.0 - ns_frac)

    total_ns = cy * ns_frac
    for b in ns_branches:
        bb = sum(l["base_freq"] for l in b["leaves"]) or 1.0
        for l in b["leaves"]:
            l["freq"] = total_ns * (l["base_freq"] / base_ns) if base_ns else 0.0
            l["ry"] = (1.0 / l["freq"]) if l["freq"] > 0 else None
            l["rate"] = (l["base_rate"] or 0.0) * ratio if l["base_rate"] is not None else None
        b["nodevals"] = _scaled_nodevals(b, ratio, sum(l["freq"] for l in b["leaves"]))

    total_fire = cy * fire_frac
    nf = len(fire_branches)
    if nf:
        weights = []
        for i, b in enumerate(fire_branches):
            if fg_list and i < len(fg_list) and fg_list[i] is not None:
                weights.append(max(0.0, fg_list[i] / 100.0))
            else:
                bb = sum(l["base_freq"] for l in b["leaves"])
                weights.append((bb / base_fire) if base_fire else 1.0 / nf)
        s = sum(weights) or 1.0
        weights = [w / s for w in weights]
        for b, w in zip(fire_branches, weights):
            bt = total_fire * w
            bb = sum(l["base_freq"] for l in b["leaves"]) or 1.0
            for l in b["leaves"]:
                l["freq"] = bt * (l["base_freq"] / bb)
                l["ry"] = (1.0 / l["freq"]) if l["freq"] > 0 else None
                l["rate"] = (l["base_rate"] or 0.0) * ratio if l["base_rate"] is not None else None
            b["nodevals"] = _scaled_nodevals(b, ratio, bt)

    v["accr"] = accr
    v["caseyr"] = cy
    v["ry"] = (1.0 / cy) if cy > 0 else None


def _scaled_nodevals(branch, ratio, total):
    base = branch.get("base_nodevals") or []
    rate = (base[0] * ratio) if base else None
    out = []
    if rate is not None:
        out.append(rate)
    out.append(total)
    out.append((1.0 / total) if total > 0 else 0.0)
    return out


def compute_connector_pcts(vehicles):
    for v in vehicles:
        vtotal = sum(l["freq"] for b in v["branches"] for l in b["leaves"]) or 0.0
        fire_branches = [b for b in v["branches"] if not b["is_ns"]]
        fire_total = sum(l["freq"] for b in fire_branches for l in b["leaves"]) or 0.0
        nfire = len(fire_branches)
        for b in v["branches"]:
            btotal = sum(l["freq"] for l in b["leaves"]) or 0.0
            if b["is_ns"]:
                b["pct"] = (btotal / vtotal) if vtotal else None
            elif nfire >= 2 and fire_total:
                b["pct"] = btotal / fire_total
            else:
                b["pct"] = (btotal / vtotal) if vtotal else None
            for l in b["leaves"]:
                if b.get("size") and len(b["leaves"]) > 1 and btotal:
                    l["pct"] = l["freq"] / btotal
                else:
                    l["pct"] = None


# ====================================================================
#  Diagram
# ====================================================================
class ScenarioDiagram(QWidget):
    vehicleToggled = pyqtSignal()      # a vehicle check-box was toggled
    PAD = 26
    CB_LANE = 34          # left lane reserved for the per-vehicle check-boxes
    GAP = 60
    GAP_VB = 116          # wider vehicle->branch gap to fit the fire junction
    W_VEH = 170
    W_BR = 152
    W_ST = 116
    W_LEAF = 168
    H_LEAF = 52
    H_STATE = 44
    H_BR_FIRE = 92
    H_BR_FLAT = 46
    H_VEH = 78
    V_GAP = 18

    _C_BG = QColor("#0f1722")
    _C_VEH = QColor("#1f6f8b")
    _C_BR = QColor("#26566b")
    _C_BR_NS = QColor("#3a4a5a")
    _C_ST = QColor("#22313f")
    _C_LEAF = QColor("#1c2c3a")
    _C_LEAFB = QColor("#3a9ec2")
    _C_LINE = QColor("#6b7f95")
    _C_TXT = QColor("#eaf1f8")
    _C_SUB = QColor("#a9bdd4")
    _C_PCT = QColor("#f0c674")

    def __init__(self, parent=None):
        super().__init__(parent)
        self._title = ""
        self._note = "Select a scenario."
        self._boxes = []
        self._segments = []
        self._pcts = []
        self._box_veh = []        # vehicle index per box / segment / pct
        self._seg_veh = []
        self._pct_veh = []
        self._veh_names = []      # display name per vehicle index
        self._checkboxes = []     # QCheckBox per vehicle
        self._enabled = {}        # display name -> bool (default True)
        self.setMinimumHeight(360)
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Minimum)

    def _on_vehicle_toggled(self, name, checked):
        self._enabled[name] = checked
        self.update()
        self.vehicleToggled.emit()

    def _alpha_for(self, vidx):
        if 0 <= vidx < len(self._veh_names):
            if not self._enabled.get(self._veh_names[vidx], True):
                return 0.22
        return 1.0

    @staticmethod
    def _f(v, dec=4):
        if v is None:
            return "-"
        av = abs(v)
        if av != 0 and (av < 1e-3 or av >= 1e6):
            return "%.2e" % v
        return ("%." + str(dec) + "f") % v

    @staticmethod
    def _fint(v):
        if v is None:
            return "-"
        if abs(v) >= 1e6:
            return "%.2e" % v
        return "{:,.1f}".format(v)

    def set_data(self, title, vehicles, note=""):
        self._title = title
        self._note = note
        self._build(vehicles)
        self.update()

    def _spine(self, x1, parent_cy, x2, targets):
        """Classic tree connector: a single stub from the parent to one
        vertical spine, then a separate horizontal branch from the spine to
        each child.  targets: list of (child_y, pct_text)."""
        if not targets:
            return
        spine_x = x1 + 20
        ys = [t[0] for t in targets] + [parent_cy]
        self._segments.append((x1, parent_cy, spine_x, parent_cy))
        self._segments.append((spine_x, min(ys), spine_x, max(ys)))
        for (cy, pct) in targets:
            self._segments.append((spine_x, cy, x2, cy))
            if pct:
                self._pcts.append(((spine_x + x2) / 2.0, cy, pct))

    @staticmethod
    def _pctlabel(p):
        return ("%.0f%%" % (p * 100)) if p is not None else ""

    def _vehicle_spine(self, vrect, vcy, br_infos, x_br, cascade_to=None):
        """Vehicle connector.  Normally a binary split into Not Serious (ns%)
        and Fire (the automatic complement, 100-ns%).

        When ``cascade_to`` is given (a fire-node rect of the vehicle above),
        the upper connector is routed up to that node instead of to this
        vehicle's own Not Serious branch, and the Not Serious branch is left
        unconnected -- a fire-escalation cascade between vehicle classes."""
        x1 = vrect.right()
        x_vs = x1 + 28
        x_fj = x_br - 44
        x_casc = x_br - 18
        ns = [(bcy, br) for (_, bcy, _, _, br) in br_infos if br["is_ns"]]
        fire = [(bcy, br) for (_, bcy, _, _, br) in br_infos if not br["is_ns"]]
        ns_sum = sum((br.get("pct") or 0.0) for _, br in ns)
        fire_pct = max(0.0, 1.0 - ns_sum)

        # stub from the vehicle out to the spine lane
        self._segments.append((x1, vcy, x_vs, vcy))

        def riser(xr, y0, y1):
            if abs(y0 - y1) > 0.5:
                self._segments.append((xr, y0, xr, y1))

        if cascade_to is not None:
            # upper connector escalates to the previous vehicle's fire node;
            # routed in its own lane (x_casc) and landed on the box's lower-left
            # edge -- offset below the node's own centre connector so it meets
            # the node itself, not the existing connector line.
            ty = cascade_to.center().y() + cascade_to.height() * 0.28
            self._segments.append((x_vs, vcy, x_casc, vcy))
            riser(x_casc, vcy, ty)
            self._segments.append((x_casc, ty, cascade_to.left(), ty))
            ns_label = self._pctlabel(ns[0][1].get("pct")) if ns else ""
            if ns_label:
                self._pcts.append(((x_vs + x_casc) / 2.0, vcy, ns_label))
        else:
            # Not Serious branch(es): own riser + horizontal, always connected
            for (bcy, br) in ns:
                riser(x_vs, vcy, bcy)
                self._segments.append((x_vs, bcy, x_br, bcy))
                self._pcts.append(((x_vs + x_br) / 2.0, bcy, self._pctlabel(br.get("pct"))))

        # Fire side
        if len(fire) == 1:
            bcy = fire[0][0]
            riser(x_vs, vcy, bcy)
            self._segments.append((x_vs, bcy, x_br, bcy))
            self._pcts.append(((x_vs + x_br) / 2.0, bcy, self._pctlabel(fire_pct)))
        elif len(fire) >= 2:
            fcy = sum(c for c, _ in fire) / len(fire)
            riser(x_vs, vcy, fcy)
            self._segments.append((x_vs, fcy, x_fj, fcy))
            self._pcts.append(((x_vs + x_fj) / 2.0, fcy, self._pctlabel(fire_pct)))
            for (bcy, br) in fire:
                riser(x_fj, fcy, bcy)
                self._segments.append((x_fj, bcy, x_br, bcy))
                self._pcts.append(((x_fj + x_br) / 2.0, bcy, self._pctlabel(br.get("pct"))))

    def _build(self, vehicles):
        self._boxes = []
        self._segments = []
        self._pcts = []
        self._box_veh = []
        self._seg_veh = []
        self._pct_veh = []
        self._veh_names = []
        for cb in self._checkboxes:
            cb.setParent(None)
            cb.deleteLater()
        self._checkboxes = []
        if not vehicles:
            self.setMinimumHeight(320)
            return

        x_veh = self.PAD + self.CB_LANE          # leave a lane for the check-boxes
        x_br = x_veh + self.W_VEH + self.GAP_VB
        x_st = x_br + self.W_BR + self.GAP
        x_leaf = x_st + self.W_ST + self.GAP

        y = self.PAD + 8
        prev_fire_rect = None
        for vidx, v in enumerate(vehicles):
            b0, s0, p0 = len(self._boxes), len(self._segments), len(self._pcts)
            br_infos = []
            br_centers = []
            for br in v["branches"]:
                is_fire = bool(br["size"])
                leaf_items = []
                for lf in br["leaves"]:
                    lrect = QRectF(x_leaf, y, self.W_LEAF, self.H_LEAF)
                    leaf_lines = [
                        lf.get("scenario_no") or _en(br["label"]),
                        "Case/Yr  " + self._f(lf["freq"]),
                        "RY  " + self._fint(lf["ry"]),
                    ]
                    self._boxes.append((lrect, leaf_lines, self._C_LEAF, self._C_LEAFB, True))
                    lcy = lrect.center().y()
                    srect = None
                    if is_fire and lf.get("state"):
                        srect = QRectF(x_st, lcy - self.H_STATE / 2, self.W_ST, self.H_STATE)
                        self._boxes.append((srect, [lf["state"], self._f(lf["rate"])],
                                            self._C_ST, self._C_ST.lighter(140), True))
                    leaf_items.append((lcy, srect, lf))
                    y += self.H_LEAF + self.V_GAP

                bcy = sum(it[0] for it in leaf_items) / len(leaf_items)
                if br["size"]:
                    bh = self.H_BR_FIRE
                    nv = br["nodevals"]
                    label_en = _en(br["label"])
                    size_en = _en(br["size"])
                    lines = [label_en]
                    if size_en and size_en != label_en:
                        lines.append(size_en)
                    if len(nv) >= 1:
                        lines.append(self._f(nv[0]))
                    if len(nv) >= 2:
                        lines.append("Case/Yr  " + self._f(nv[1]))
                    if len(nv) >= 3:
                        lines.append("RY  " + self._fint(nv[2]))
                    fill, border = self._C_BR, self._C_BR.lighter(135)
                else:
                    bh = self.H_BR_FLAT
                    rate0 = br["leaves"][0]["rate"] if br["leaves"] else None
                    lines = [_en(br["label"]), self._f(rate0)]
                    fill, border = self._C_BR_NS, self._C_BR_NS.lighter(140)
                brect = QRectF(x_br, bcy - bh / 2, self.W_BR, bh)
                self._boxes.append((brect, lines, fill, border, True))
                br_centers.append(bcy)
                br_infos.append((brect, bcy, is_fire, leaf_items, br))
                y += self.V_GAP

            vcy = sum(br_centers) / len(br_centers) if br_centers else (self.PAD + self.H_VEH)
            vrect = QRectF(x_veh, vcy - self.H_VEH / 2, self.W_VEH, self.H_VEH)
            vlines = [_en(v["name"]), "ACCR  " + self._f(v["accr"]),
                      "Case/Yr  " + self._f(v["caseyr"]), "RY  " + self._fint(v["ry"])]
            self._boxes.append((vrect, vlines, self._C_VEH, self._C_VEH.lighter(135), True))

            # vehicle -> branches.  Cargo (GV) and hazardous (ST) vehicles
            # cascade their upper connector up to the previous vehicle's fire
            # node and leave their own Not Serious unconnected.
            cur_fire_rect = next((bi[0] for bi in br_infos if bi[2]), None)
            cls = classify_vehicle(v["name"])
            cascade_to = prev_fire_rect if (cls in (2, 3) and prev_fire_rect is not None) else None
            self._vehicle_spine(vrect, vcy, br_infos, x_br, cascade_to=cascade_to)
            if cur_fire_rect is not None:
                prev_fire_rect = cur_fire_rect

            # branch -> states/leaves
            for (brect, bcy, is_fire, leaf_items, br) in br_infos:
                has_states = is_fire and any(it[1] for it in leaf_items)
                if has_states:
                    self._spine(brect.right(), bcy, x_st,
                                [(it[1].center().y(),
                                  ("%.0f%%" % (it[2]["pct"] * 100)) if it[2].get("pct") is not None else "")
                                 for it in leaf_items])
                    for (lcy, srect, lf) in leaf_items:
                        self._segments.append((srect.right(), lcy, x_leaf, lcy))
                elif len(leaf_items) > 1:
                    self._spine(brect.right(), bcy, x_leaf,
                                [(it[0], "") for it in leaf_items])
                else:
                    lcy = leaf_items[0][0]
                    self._segments.append((brect.right(), lcy, x_leaf, lcy))
            y += self.V_GAP * 2

            # tag everything appended this iteration with the vehicle index
            name = _en(v["name"])
            self._veh_names.append(name)
            self._box_veh += [vidx] * (len(self._boxes) - b0)
            self._seg_veh += [vidx] * (len(self._segments) - s0)
            self._pct_veh += [vidx] * (len(self._pcts) - p0)

            # per-vehicle include check-box, in the lane just left of the box
            # (where the red markers are), styled to stand out on the dark bg.
            cb = QCheckBox(self)
            cb.setChecked(self._enabled.get(name, True))
            cb.setGeometry(int(x_veh - 30), int(vcy - 16), 22, 22)
            cb.setCursor(Qt.PointingHandCursor)
            cb.setToolTip("Include %s in the scenario and its HRR" % name)
            cb.setStyleSheet(
                "QCheckBox{background:transparent;spacing:0px;}"
                "QCheckBox::indicator{width:18px;height:18px;border:2px solid "
                "#9fb6cc;border-radius:4px;background:#0f1722;}"
                "QCheckBox::indicator:checked{background:#36c08a;"
                "border:2px solid #eaf4ff;}"
                "QCheckBox::indicator:hover{border:2px solid #f0c674;}")
            cb.toggled.connect(
                lambda checked, nm=name: self._on_vehicle_toggled(nm, checked))
            cb.raise_()
            cb.show()
            self._checkboxes.append(cb)

        self.setMinimumWidth(int(x_leaf + self.W_LEAF + self.PAD))
        self.setMinimumHeight(int(y + self.PAD))

    def paintEvent(self, _e):
        p = QPainter(self)
        p.setRenderHint(QPainter.Antialiasing, True)
        p.fillRect(self.rect(), self._C_BG)
        if not self._boxes:
            p.setPen(self._C_SUB); p.setFont(QFont("", 10))
            p.drawText(self.rect(), Qt.AlignCenter, self._note or "No diagram.")
            p.end(); return
        p.setPen(self._C_SUB)
        tf = QFont(); tf.setPointSize(10); tf.setBold(True); p.setFont(tf)
        p.drawText(QRectF(self.PAD, 2, self.width() - self.PAD, 18),
                   Qt.AlignLeft | Qt.AlignVCenter, self._title)
        for i, (x1, y1, x2, y2) in enumerate(self._segments):
            p.setOpacity(self._alpha_for(self._seg_veh[i] if i < len(self._seg_veh) else -1))
            p.setPen(QPen(self._C_LINE, 1.4))
            p.drawLine(QPointF(x1, y1), QPointF(x2, y2))
        p.setOpacity(1.0)
        pf = QFont(); pf.setPointSize(8); pf.setBold(True)
        for i, (cx, cy, pct) in enumerate(self._pcts):
            p.setOpacity(self._alpha_for(self._pct_veh[i] if i < len(self._pct_veh) else -1))
            p.setPen(self._C_PCT); p.setFont(pf)
            p.drawText(QRectF(cx - 20, cy - 15, 40, 13), Qt.AlignCenter, pct)
        p.setOpacity(1.0)
        for i, (rect, lines, fill, border, head_bold) in enumerate(self._boxes):
            p.setOpacity(self._alpha_for(self._box_veh[i] if i < len(self._box_veh) else -1))
            p.setBrush(QBrush(fill)); p.setPen(QPen(border, 1.5))
            p.drawRoundedRect(rect, 7, 7)
            n = len(lines); lh = rect.height() / max(n, 1)
            for j, line in enumerate(lines):
                if j == 0:
                    p.setPen(self._C_TXT)
                    f = QFont(); f.setPointSize(9 if rect.width() > 150 else 8); f.setBold(head_bold)
                else:
                    p.setPen(self._C_SUB)
                    f = QFont(); f.setPointSize(8)
                p.setFont(f)
                p.drawText(QRectF(rect.x() + 7, rect.y() + j * lh, rect.width() - 12, lh),
                           Qt.AlignLeft | Qt.AlignVCenter, line)
        p.setOpacity(1.0)
        p.end()


# ====================================================================
#  Control table
# ====================================================================
class ControlTable(QWidget):
    valuesChanged = pyqtSignal()
    HEADERS = ["Vehicle", "ACCR", "N-Serious", "FG1", "FG2"]

    def __init__(self, parent=None):
        super().__init__(parent)
        self._loading = False
        v = QVBoxLayout(self)
        v.setContentsMargins(0, 0, 0, 0)
        v.setSpacing(6)
        self.table = QTableWidget(0, len(self.HEADERS))
        self.table.setHorizontalHeaderLabels(self.HEADERS)
        self.table.verticalHeader().setVisible(False)
        self.table.setSizeAdjustPolicy(QAbstractScrollArea.AdjustToContents)
        self.table.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.table.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.table.setStyleSheet(
            "QTableWidget{background:#ffffff;color:#1a252f;gridline-color:#cdd9e6;"
            "font-size:11px;border:1px solid #95a5a6;}"
            "QHeaderView::section{background:#2c4a63;color:white;font-weight:bold;"
            "padding:3px 6px;border:1px solid #1f3a52;}")
        for c in range(len(self.HEADERS)):
            self.table.horizontalHeader().setSectionResizeMode(c, QHeaderView.ResizeToContents)
        self.table.itemChanged.connect(self._on_item_changed)
        v.addWidget(self.table)

        bar = QHBoxLayout()
        self.btn_recalc = QPushButton("Recalculate")
        self.btn_recalc.setStyleSheet("QPushButton{background:#2980b9;color:white;"
                                      "font-weight:bold;border-radius:5px;padding:4px 10px;}")
        self.btn_recalc.clicked.connect(lambda: self.valuesChanged.emit())
        self.btn_reset = QPushButton("Reset")
        self.btn_reset.setStyleSheet("QPushButton{background:#7f8c8d;color:white;"
                                     "font-weight:bold;border-radius:5px;padding:4px 10px;}")
        self.btn_reset.clicked.connect(self._reset)
        bar.addWidget(self.btn_recalc); bar.addWidget(self.btn_reset); bar.addStretch(1)
        v.addLayout(bar)
        self.setSizePolicy(QSizePolicy.Maximum, QSizePolicy.Maximum)
        self._baseline = []

    def _set(self, r, c, text, editable):
        it = self.table.item(r, c)
        it.setText(text)
        if editable:
            it.setFlags(Qt.ItemIsEnabled | Qt.ItemIsEditable | Qt.ItemIsSelectable)
            it.setForeground(QColor("#1a252f")); it.setBackground(QColor("#ffffff"))
        else:
            it.setFlags(Qt.ItemIsEnabled)
            it.setForeground(QColor("#9aa7b4")); it.setBackground(QColor("#eef1f4"))

    def load(self, rows):
        """rows: list of dict(label, accr, ns, fg1, fg2, nfire), one per vehicle."""
        self._loading = True
        self._baseline = [dict(d) for d in rows]
        self.table.setRowCount(len(rows))
        for r, data in enumerate(rows):
            if self.table.item(r, 0) is None:
                it = QTableWidgetItem()
                it.setFlags(Qt.ItemIsEnabled)
                it.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
                self.table.setItem(r, 0, it)
                for c in range(1, 5):
                    cell = QTableWidgetItem("-")
                    cell.setTextAlignment(Qt.AlignCenter)
                    cell.setFlags(Qt.ItemIsEnabled)
                    self.table.setItem(r, c, cell)
            self.table.item(r, 0).setText(data.get("label", ""))
            self._set(r, 1, "%.4f" % data["accr"] if data.get("accr") is not None else "-",
                      data.get("accr") is not None)
            self._set(r, 2, "%.2f" % data["ns"] if data.get("ns") is not None else "-",
                      data.get("ns") is not None)
            nf = data.get("nfire", 0)
            self._set(r, 3, "%.2f" % data["fg1"] if data.get("fg1") is not None else "-", nf >= 1)
            self._set(r, 4, "%.2f" % data["fg2"] if data.get("fg2") is not None else "-", nf >= 2)
        self._loading = False
        self.table.resizeColumnsToContents()
        self._fit_width()
        self._fit_height()

    def _fit_width(self):
        w = self.table.frameWidth() * 2 + self.table.verticalHeader().width()
        for c in range(self.table.columnCount()):
            w += self.table.columnWidth(c)
        self.table.setFixedWidth(w + 4)

    def _fit_height(self):
        """Pin the table tall enough to show every row + the header, so no
        rows are hidden behind a scrollbar."""
        h = self.table.horizontalHeader().height() + self.table.frameWidth() * 2
        for r in range(self.table.rowCount()):
            h += self.table.rowHeight(r)
        self.table.setFixedHeight(h + 2)

    def _reset(self):
        self.load(self._baseline)
        self.valuesChanged.emit()

    def _on_item_changed(self, _it):
        if not self._loading:
            self.valuesChanged.emit()

    def values(self):
        out = {}
        for r in range(self.table.rowCount()):
            def num(c):
                try:
                    return float(self.table.item(r, c).text())
                except (TypeError, ValueError, AttributeError):
                    return None
            out[r] = dict(accr=num(1), ns=num(2), fg1=num(3), fg2=num(4))
        return out


# ====================================================================
#  Scenario panel
# ====================================================================
class ScenarioPanel(QWidget):
    scenarioSelected = pyqtSignal()      # emitted when the active sheet changes

    def __init__(self, workbook_path=None, parent=None):
        super().__init__(parent)
        self._wb_path = workbook_path or self._find_workbook()
        self._cache = {}
        self._baseline = []       # parsed vehicles for current sheet
        self._mapping = {}        # control row -> vehicle index
        self._sheets = self._load_sheet_names()

        root = QHBoxLayout(self)
        root.setContentsMargins(6, 6, 6, 6)
        root.setSpacing(10)

        left = QVBoxLayout(); left.setSpacing(8)
        gsel = QGroupBox("Scenarios")
        gsel.setStyleSheet("QGroupBox{font-weight:bold;}")
        gv = QVBoxLayout(gsel)
        self.btn_group = QButtonGroup(self)
        for i, name in enumerate(self._sheets):
            rb = QRadioButton(name)
            rb.setStyleSheet("QRadioButton{color:#1a252f;font-size:11px;padding:2px;}")
            self.btn_group.addButton(rb, i)
            gv.addWidget(rb)
            if i == 0:
                rb.setChecked(True)
        gv.addStretch(1)
        self.btn_group.idClicked.connect(self._on_select)
        left.addWidget(gsel)

        gctrl = QGroupBox("Control")
        gctrl.setStyleSheet("QGroupBox{font-weight:bold;}")
        cv = QVBoxLayout(gctrl); cv.setContentsMargins(8, 6, 8, 8)
        self.control = ControlTable()
        self.control.valuesChanged.connect(self._apply_control)
        cv.addWidget(self.control, 0, Qt.AlignLeft | Qt.AlignTop)
        gctrl.setSizePolicy(QSizePolicy.Maximum, QSizePolicy.Maximum)
        left.addWidget(gctrl)

        # Traffic Ventilation host (the launcher reparents the app's
        # "Scenario Configuration" group in here, renamed "Traffic Ventilation")
        self.tv_host = QWidget()
        self._tv_layout = QVBoxLayout(self.tv_host)
        self._tv_layout.setContentsMargins(0, 0, 0, 0)
        self._tv_placeholder = QLabel(
            "Traffic Ventilation configuration appears here when loaded inside "
            "the QRA System.")
        self._tv_placeholder.setWordWrap(True)
        self._tv_placeholder.setStyleSheet(
            "color:#5a6b7d;font-size:11px;padding:6px;"
            "border:1px dashed #9fb6cc;border-radius:6px;")
        self._tv_layout.addWidget(self._tv_placeholder)
        left.addWidget(self.tv_host)
        left.addStretch(1)
        left_box = QWidget(); left_box.setLayout(left)

        gdia = QGroupBox("Node Connection Diagram")
        gdia.setStyleSheet("QGroupBox{font-weight:bold;}")
        dv = QVBoxLayout(gdia)
        self.diagram = ScenarioDiagram()
        self.diagram.vehicleToggled.connect(self.scenarioSelected.emit)
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setWidget(self.diagram)
        scroll.setMinimumHeight(440)
        scroll.setStyleSheet("QScrollArea{border:1px solid #24405f;border-radius:6px;}")
        dv.addWidget(scroll)

        root.addWidget(left_box, 20)
        root.addWidget(gdia, 80)

        if self._sheets:
            self._on_select(0)

    def mount_traffic_ventilation(self, group_widget):
        """Place the reparented 'Traffic Ventilation' group (originally the
        app's 'Scenario Configuration') under the Control panel.  Reparenting
        keeps self.fire_positions_input / vent_* / traffic_* references and
        signals valid."""
        if self._tv_placeholder is not None:
            self._tv_placeholder.hide()
            self._tv_layout.removeWidget(self._tv_placeholder)
            self._tv_placeholder.deleteLater()
            self._tv_placeholder = None
        if group_widget is not None:
            self._tv_layout.addWidget(group_widget)
            group_widget.show()
        return True

    # ---- workbook ----
    @staticmethod
    def _find_workbook():
        here = os.path.dirname(os.path.abspath(__file__))
        for name in ("scenarios.xlsx", "Scenarios.xlsx"):
            p = os.path.join(here, name)
            if os.path.exists(p):
                return p
        return os.path.join(here, "scenarios.xlsx")

    def _load_sheet_names(self):
        if _HAS_OPENPYXL and os.path.exists(self._wb_path):
            try:
                wb = _openpyxl.load_workbook(self._wb_path, read_only=True, data_only=True)
                names = list(wb.sheetnames); wb.close()
                if names:
                    return names
            except Exception:
                pass
        return list(_FALLBACK_SHEETS)

    def _parse(self, sheet_name):
        if sheet_name in self._cache:
            return copy.deepcopy(self._cache[sheet_name])
        vehicles = []
        if _HAS_OPENPYXL and os.path.exists(self._wb_path):
            try:
                wb = _openpyxl.load_workbook(self._wb_path, data_only=True)
                if sheet_name in wb.sheetnames:
                    vehicles = parse_scenario_tree(wb[sheet_name])
                wb.close()
            except Exception:
                vehicles = []
        self._cache[sheet_name] = vehicles
        return copy.deepcopy(vehicles)

    # ---- selection / control ----
    def _build_control_rows(self, vehicles):
        """One control row per vehicle in the selected sheet, in sheet order."""
        rows = []
        self._mapping = {}
        for idx, v in enumerate(vehicles):
            ns_branches = [b for b in v["branches"] if b["is_ns"]]
            fire_branches = [b for b in v["branches"] if not b["is_ns"]]
            cy = v.get("base_caseyr") or sum(l["base_freq"] for b in v["branches"] for l in b["leaves"])
            base_ns = sum(l["base_freq"] for b in ns_branches for l in b["leaves"])
            base_fire = sum(l["base_freq"] for b in fire_branches for l in b["leaves"])
            ns_pct = (base_ns / cy * 100.0) if cy else None
            fg1 = fg2 = None
            if fire_branches:
                t0 = sum(l["base_freq"] for l in fire_branches[0]["leaves"])
                fg1 = (t0 / base_fire * 100.0) if base_fire else 100.0
            if len(fire_branches) >= 2:
                t1 = sum(l["base_freq"] for l in fire_branches[1]["leaves"])
                fg2 = (t1 / base_fire * 100.0) if base_fire else 0.0
            rows.append(dict(label=_en(v["name"]), accr=v.get("accr"), ns=ns_pct,
                             fg1=fg1, fg2=fg2, nfire=len(fire_branches)))
            self._mapping[idx] = idx
        return rows

    def _on_select(self, idx):
        if not self._sheets:
            return
        idx = max(0, min(idx, len(self._sheets) - 1))
        self._sheet_name = self._sheets[idx]
        self._baseline = self._parse(self._sheet_name)
        # a fresh scenario starts with every vehicle included
        self.diagram._enabled.clear()
        rows = self._build_control_rows(self._baseline)
        self.control.load(rows)
        self._apply_control()
        self.scenarioSelected.emit()

    def selected_hrr_mw(self):
        """Distinct fire-size HRR values (MW), sorted ascending, of the
        currently included vehicles in the selected scenario.  Vehicles whose
        diagram check-box is unticked are excluded, so their HRR drops out of
        the fds_inputs folders."""
        en = getattr(self.diagram, "_enabled", {})
        vehicles = [v for v in (self._baseline or [])
                    if en.get(_en(v["name"]), True)]
        return self._hrr_from_vehicles(vehicles)

    def all_hrr_mw(self):
        """Union of distinct fire-size HRR values (MW) across ALL scenario
        sheets, sorted ascending -- e.g. [5, 10, 15, 20, 30, 100]."""
        sizes = set()
        for name in self._sheets:
            try:
                sizes.update(self._hrr_from_vehicles(self._parse(name)))
            except Exception:
                continue
        return sorted(sizes)

    @staticmethod
    def _hrr_from_vehicles(vehicles):
        sizes = set()
        for v in (vehicles or []):
            for b in v.get("branches", []):
                s = b.get("size")
                if not s:
                    continue
                m = re.search(r"(\d+(?:\.\d+)?)", str(s))
                if m:
                    val = float(m.group(1))
                    sizes.add(int(val) if val.is_integer() else val)
        return sorted(sizes)

    def _apply_control(self):
        vehicles = copy.deepcopy(self._baseline)
        if vehicles:
            vals = self.control.values()
            for r, vidx in self._mapping.items():
                if vidx >= len(vehicles):
                    continue
                cv = vals.get(r, {})
                accr = cv.get("accr")
                if accr is None:
                    accr = vehicles[vidx].get("base_accr") or vehicles[vidx].get("accr") or 0.0
                fg_list = [cv.get("fg1"), cv.get("fg2")]
                recompute_vehicle(vehicles[vidx], accr, cv.get("ns"), fg_list)
            compute_connector_pcts(vehicles)
            note = ""
        elif not _HAS_OPENPYXL:
            note = "Install 'openpyxl' and place scenarios.xlsx next to the app."
        elif not os.path.exists(self._wb_path):
            note = "scenarios.xlsx not found next to the app."
        else:
            note = "No event-tree data could be parsed from this sheet."
        self.diagram.set_data(getattr(self, "_sheet_name", ""), vehicles, note)