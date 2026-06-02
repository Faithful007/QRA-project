"""
regen_tunnel_json.py
Regenerates tunnel_traffic_sheet.json from FNCV_ROAD.xlsm with:
  - All cells (cached values + formulas)
  - All colors preserved
  - All named ranges pointing to this sheet
  - Merged-cell info
  - Column widths
"""

import json
import re
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.utils import get_column_letter, column_index_from_string
except ImportError:
    print("openpyxl not found. Run: pip install openpyxl")
    sys.exit(1)

SHEET_NAME = '터널교통량등제원'
XLSM_PATH  = Path(__file__).with_name('FNCV_ROAD.xlsm')
JSON_PATH  = Path(__file__).with_name('tunnel_traffic_sheet.json')

BLUE_EDITABLE = 'FF00B0F0'

# ── Helpers ──────────────────────────────────────────────────────────────────

def normalise_color(rgb: str) -> str | None:
    """Normalise ARGB colour string; return None for transparent/no-fill."""
    if not rgb or rgb in ('00000000', '000000', 'FFFFFFFF', '00FFFFFF'):
        return None
    if len(rgb) == 6:
        return 'FF' + rgb.upper()
    return rgb.upper()


def col_idx_to_letter(n: int) -> str:
    letters = ''
    while n > 0:
        n, rem = divmod(n - 1, 26)
        letters = chr(65 + rem) + letters
    return letters


# ── Load workbooks ────────────────────────────────────────────────────────────

print(f"Loading {XLSM_PATH} …")
wb_formula = openpyxl.load_workbook(str(XLSM_PATH), data_only=False)
wb_data    = openpyxl.load_workbook(str(XLSM_PATH), data_only=True)

ws_f = wb_formula[SHEET_NAME]
ws_d = wb_data[SHEET_NAME]

print(f"Sheet '{SHEET_NAME}': {ws_f.max_row} rows × {ws_f.max_column} cols")

# ── Named ranges ──────────────────────────────────────────────────────────────

# Pattern: optional quote + sheet name + optional quote + ! + $COL$ROW
_name_pat = re.compile(
    r"'?" + re.escape(SHEET_NAME) + r"'?!\$?([A-Z]{1,3})\$?(\d+)",
    re.IGNORECASE
)

named_refs: dict   = {}   # name -> {sheet, ref (coord)}
coord_to_names: dict = {}  # coord -> [name, ...]

for dn_name in wb_formula.defined_names:
    defn = wb_formula.defined_names[dn_name]
    ref  = (defn.attr_text or '').strip()
    m = _name_pat.match(ref)
    if m:
        col_l, row_s = m.group(1).upper(), m.group(2)
        coord = f"{col_l}{row_s}"
        named_refs[dn_name] = {'sheet': SHEET_NAME, 'ref': coord}
        coord_to_names.setdefault(coord, []).append(dn_name)

print(f"Named ranges pointing to this sheet: {len(named_refs)}")

# ── Merged cells ──────────────────────────────────────────────────────────────

merged_cells: list = []
merged_coords: set = set()   # all coords that are part of any merge

for mc_range in ws_f.merged_cells.ranges:
    min_row = mc_range.min_row
    max_row = mc_range.max_row
    min_col = mc_range.min_col
    max_col = mc_range.max_col
    top_left = f"{col_idx_to_letter(min_col)}{min_row}"
    merged_cells.append({
        'range': str(mc_range),
        'min_row': min_row, 'max_row': max_row,
        'min_col': min_col, 'max_col': max_col,
        'top_left': top_left,
        'row_span': max_row - min_row + 1,
        'col_span': max_col - min_col + 1,
    })
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            coord = f"{col_idx_to_letter(c)}{r}"
            merged_coords.add(coord)

print(f"Merged cell ranges: {len(merged_cells)}")

# ── Column widths ─────────────────────────────────────────────────────────────

col_widths: dict = {}   # 1-based col index -> width (Excel units)
for col_letter, dim in ws_f.column_dimensions.items():
    if dim.width:
        idx = column_index_from_string(col_letter)
        col_widths[idx] = round(dim.width, 2)

# Default width if not explicitly set
DEFAULT_COL_WIDTH = 8.43

# ── Build cell list ───────────────────────────────────────────────────────────

cells: list = []

for row_cells in ws_f.iter_rows():
    for cell_f in row_cells:
        if cell_f.value is None:
            continue

        row = cell_f.row
        col = cell_f.column
        coord = cell_f.coordinate   # e.g. "D8"

        # Cached value from data_only workbook
        cell_d = ws_d.cell(row=row, column=col)
        cached_value = cell_d.value

        # Formula / literal value
        formula = None
        value   = cached_value

        raw_val = cell_f.value
        if isinstance(raw_val, str) and raw_val.startswith('='):
            formula = raw_val
            value   = cached_value        # use cached value as display default
        else:
            value = raw_val

        # Background colour
        try:
            fg = cell_f.fill.fgColor
            rgb = fg.rgb if fg and fg.type == 'rgb' else None
            color = normalise_color(rgb)
        except Exception:
            color = None

        is_editable = (color == BLUE_EDITABLE)

        cells.append({
            'row':           row,
            'col':           col,
            'coordinate':    coord,
            'value':         value,
            'formula':       formula,
            'cached_value':  cached_value,
            'color':         color,
            'is_editable':   is_editable,
            'in_merge':      coord in merged_coords,
            'manual_override': False,   # always reset on regen
        })

# Force specific cells to always be editable (blue) regardless of their Excel
# background colour.  These are cells the user must be able to edit freely
# without unlocking admin mode.
FORCE_EDITABLE_CELLS = {'J3', 'K3'}
for _c in cells:
    if _c['coordinate'] in FORCE_EDITABLE_CELLS:
        _c['is_editable'] = True
        _c['color'] = BLUE_EDITABLE

print(f"Total cells: {len(cells)}")
editable_count = sum(1 for c in cells if c['is_editable'])
formula_count  = sum(1 for c in cells if c.get('formula'))
print(f"  Editable (blue): {editable_count}")
print(f"  Formula cells:   {formula_count}")

# ── Build named_values snapshot ───────────────────────────────────────────────
# Resolve each named ref to its cached cell value so the formula evaluator
# has a fallback when cell lookup fails.

cells_by_coord = {c['coordinate']: c for c in cells}

named_values: dict = {}
for name, info in named_refs.items():
    coord = info['ref']
    cell  = cells_by_coord.get(coord)
    if cell:
        named_values[name] = cell['value']

print(f"Named values resolved: {len(named_values)}")

# ── Assemble JSON ─────────────────────────────────────────────────────────────

output = {
    'sheet_name':   SHEET_NAME,
    'max_row':      ws_f.max_row,
    'max_col':      ws_f.max_column,
    'named_refs':   named_refs,
    'named_values': named_values,
    'merged_cells': merged_cells,
    'col_widths':   col_widths,
    'cells':        cells,
}

json_str = json.dumps(output, ensure_ascii=False, indent=2, default=str)
JSON_PATH.write_text(json_str, encoding='utf-8')
print(f"\nWritten to {JSON_PATH}")
print(f"File size: {JSON_PATH.stat().st_size / 1024:.1f} KB")
